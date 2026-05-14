"""
Excel 02_A2 sheet'inde 'Markaya İletildi mi?' satirini yumusat.
"""
from openpyxl import load_workbook
from pathlib import Path

XLSX = Path("/Users/Erdo/Desktop/Claude Projects/Turkcel/TURKCELL_TRAFIK_FIRSATLARI.xlsx")
wb = load_workbook(XLSX)

# 02_A2 sheet
ws = wb["02_A2_Tatil_Bayram_Fikirleri"]
# Tüm hücrelerde markaya iletildi -> daha önce konuşulmuş
for row in ws.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            if "Markaya İletildi mi" in cell.value:
                cell.value = cell.value.replace("Markaya İletildi mi", "Daha Önce Konuşuldu mu")
            if "Markaya iletildi" in cell.value:
                cell.value = cell.value.replace("Markaya iletildi", "Daha önce konuşuldu")
            if "markaya iletilmiş" in cell.value:
                cell.value = cell.value.replace("markaya iletilmiş", "daha önce konuşulmuş")
            if "Markaya iletilmiş" in cell.value:
                cell.value = cell.value.replace("Markaya iletilmiş", "Daha önce konuşulmuş")

# 00_Yonetici_Ozeti - daha temiz wording
ws = wb["00_Yonetici_Ozeti"]
for row in ws.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            v = cell.value
            v = v.replace("Aylık Adreslenebilir Hacim", "Aylık Toplam Arama Hacmi")
            v = v.replace("DfS Doğrulanmış KW", "Anahtar Kelime Verisi İncelenmiş")
            v = v.replace("DfS doğrulandı", "veri üzerinden incelendi")
            v = v.replace("MARKAYA İLETİLMİŞ", "DAHA ÖNCE KONUŞULMUŞ")
            v = v.replace("Markaya iletilmiş", "Daha önce konuşulmuş")
            v = v.replace("markaya iletilmiş", "daha önce konuşulmuş")
            cell.value = v

wb.save(XLSX)
print(f"Updated Excel: {XLSX}")
