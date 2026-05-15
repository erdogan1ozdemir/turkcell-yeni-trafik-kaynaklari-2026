"""
Excel'den fal-related her sey temizlensin:
1. A5 sheet: 'BURÇ & FAL & İSİM ANLAMI' satirini sil
2. A1 sheet: Burc satirinda fal-related kelimeler varsa kaldir
3. DfS Dogrulama: kahve falı, tarot falı, el falı, yıldız falı satirlarini sil
4. 27_n_BURC_ASTROLOJI niche sheet'i icerigi kontrol et
5. Cluster Ozeti'nde BURC_ASTROLOJI satirina not ekle
6. 'Tüm 60K' sheet'inde fal satirlari icin filtre/temizlik (cok buyuk dosya, etkili kalsin)
"""
from openpyxl import load_workbook
from pathlib import Path
import re

XLSX = Path("/Users/Erdo/Desktop/Claude Projects/Turkcel/TURKCELL_TRAFIK_FIRSATLARI.xlsx")
wb = load_workbook(XLSX)

changes = []

# ============================================================
# 1. A5 sheet: BURC & FAL & ISIM ANLAMI satirini sil
# ============================================================
ws = wb["05_A5_AI_Siber_Diger"]
to_delete = []
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if v and ("BURÇ" in str(v).upper() or "Burc" in str(v) or "FAL" in str(v).upper() or "İSİM ANLAMI" in str(v).upper()):
        to_delete.append(r)
for r in reversed(to_delete):
    ws.delete_rows(r)
if to_delete:
    changes.append(f"A5: BURÇ & FAL & İSİM ANLAMI satiri silindi ({len(to_delete)} satir)")

# ============================================================
# 2. A1 sheet: Burc satirinda fal kelimelerini kaldir
# ============================================================
ws = wb["01_A1_Hesaplama_Fikirleri"]
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if v and "burc" in str(v).lower():
        # Bu Yukselen Burc satiri - KW kolonunda fal varsa kaldir
        for c in range(2, ws.max_column + 1):
            cell_val = ws.cell(row=r, column=c).value
            if cell_val and isinstance(cell_val, str):
                new_val = cell_val
                # Fal varyasyonlarini kaldir
                new_val = re.sub(r"kahve fal[ıi][^,\)]*[,;]?\s*", "", new_val, flags=re.IGNORECASE)
                new_val = re.sub(r"tarot fal[ıi][^,\)]*[,;]?\s*", "", new_val, flags=re.IGNORECASE)
                new_val = re.sub(r"el fal[ıi][^,\)]*[,;]?\s*", "", new_val, flags=re.IGNORECASE)
                new_val = re.sub(r"y[ıi]ld[ıi]z fal[ıi][^,\)]*[,;]?\s*", "", new_val, flags=re.IGNORECASE)
                new_val = re.sub(r"\bfal[ıi][^,\)]*[,;]?\s*", "", new_val, flags=re.IGNORECASE)
                new_val = re.sub(r"\s+", " ", new_val).strip().strip(",").strip()
                if new_val != cell_val:
                    ws.cell(row=r, column=c, value=new_val)
                    changes.append(f"A1 Row {r} Col {c}: fal kelimeleri kaldirildi")
        break

# ============================================================
# 3. DfS Dogrulama: fal-related kelimeleri sil
# ============================================================
if "09_DfS_Dogrulama_TR" in wb.sheetnames:
    ws = wb["09_DfS_Dogrulama_TR"]
    to_delete = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v:
            v_str = str(v).lower()
            if "kahve fal" in v_str or "tarot fal" in v_str or "el fal" in v_str or "yıldız fal" in v_str or "fal " in v_str:
                to_delete.append(r)
    for r in reversed(to_delete):
        ws.delete_rows(r)
    if to_delete:
        changes.append(f"DfS Dogrulama: {len(to_delete)} fal satiri silindi")

# ============================================================
# 4. 27_n_BURC_ASTROLOJI niche sheet'i - icerigi gozden gecir
# ============================================================
if "27_n_BURC_ASTROLOJI" in wb.sheetnames:
    ws = wb["27_n_BURC_ASTROLOJI"]
    # Bu sheet zaten cluster verisinden gelmis - fal-related kelimeler olabilir
    # Tum tablo halinde silmek yerine, sheet adini "Bilgi" sutunundaki fal sorgularini isaretleyelim
    # Aslinda bu cluster eski uretimden gelmis, simdiki strateji burc ticari yonlendirme.
    # Bu sheet'i komple silmek yerine icerigini bos birakmak veya 'kaldirildi' notu eklemek.

    # Mevcut tum verisini sil, yerine not yaz
    rows_to_clear = ws.max_row
    ws.delete_rows(1, rows_to_clear)
    ws.cell(row=1, column=1, value="BURÇ / ASTROLOJİ CLUSTER (kaldırıldı)")
    ws.cell(row=2, column=1, value="Strateji değişikliği: Fal ve astroloji içerikleri rapor kapsamından çıkarıldı.")
    ws.cell(row=3, column=1, value="Yalnızca 'yükselen burç hesaplama' aracı kalır ve o da A1'de ticari yönlendirme")
    ws.cell(row=4, column=1, value="(paketler + Pasaj ürün-kategori) olarak konumlanır.")
    changes.append("27_n_BURC_ASTROLOJI: tum satirlar silindi, kaldirildi notu eklendi")

# ============================================================
# 5. Cluster Ozeti'nde BURC_ASTROLOJI satirini sil
# ============================================================
if "07_B1_Cluster_Ozeti" in wb.sheetnames:
    ws = wb["07_B1_Cluster_Ozeti"]
    to_delete = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and ("BURC_ASTROLOJI" in str(v) or "Burç" in str(v) or "BURÇ" in str(v).upper()):
            to_delete.append(r)
    for r in reversed(to_delete):
        ws.delete_rows(r)
    if to_delete:
        changes.append(f"Cluster Ozeti: {len(to_delete)} BURC/Astroloji satiri silindi")

# ============================================================
# Kayit
# ============================================================
wb.save(XLSX)
print("\n=== TEMIZLIK TAMAM ===")
for c in changes:
    print("  -", c)
print(f"\nToplam: {len(changes)} degisiklik")
