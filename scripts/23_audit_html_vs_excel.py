"""
HTML'deki idea-card'lari Excel'deki ilgili fikir sheet'leriyle karsilastir.
Tutarsizliklari raporla.
"""
import re
from pathlib import Path
from openpyxl import load_workbook

HTML = Path("/Users/Erdo/Desktop/Claude Projects/Turkcel/index.html")
XLSX = Path("/Users/Erdo/Desktop/Claude Projects/Turkcel/TURKCELL_TRAFIK_FIRSATLARI.xlsx")

# HTML'i oku ve her section icin idea-card basliklarini cikar
text = HTML.read_text(encoding="utf-8")

def extract_section_cards(text, section_id):
    """Belirli bir section icinden tum idea-card h4 basliklarini cikar"""
    sec_start = text.find(f'id="{section_id}"')
    if sec_start < 0:
        return []
    next_sec = text.find('<section class="section"', sec_start + 10)
    section_text = text[sec_start:next_sec] if next_sec > 0 else text[sec_start:]
    # Tum h4 basliklarini bul
    pattern = re.compile(r'<div class="idea-card"[^>]*data-marka="([^"]*)"[^>]*>.*?<h4>(.*?)</h4>', re.DOTALL)
    return pattern.findall(section_text)

# Her section'dan kartlari cikar
sections = {
    'a6 (#a6)': extract_section_cards(text, 'a6'),
    'a1 (#a1)': extract_section_cards(text, 'a1'),
    'a3 (#a3)': extract_section_cards(text, 'a3'),
    'a4 (#a4)': extract_section_cards(text, 'a4'),
    'a5 (#a5)': extract_section_cards(text, 'a5'),
}

print("=" * 80)
print("HTML idea-card SAYIM")
print("=" * 80)
for k, cards in sections.items():
    print(f"\n{k}: {len(cards)} kart")
    for i, (marka, title) in enumerate(cards, 1):
        marka_str = "(*)" if marka == "evet" else "    "
        print(f"  {i:2d}. {marka_str} {title[:80]}")

# Excel'i oku
print("\n" + "=" * 80)
print("EXCEL SHEET SAYIM")
print("=" * 80)

wb = load_workbook(XLSX)
sheet_map = {
    'a6': '06_A6_2026_Plan_Perspektif',
    'a1': '01_A1_Hesaplama_Fikirleri',
    'a2': '02_A2_Tatil_Bayram_Fikirleri',
    'a3': '03_A3_OzelGun_Mesaj_Fikirleri',
    'a4': '04_A4_Telekom_Sorgu_Fikirleri',
    'a5': '05_A5_AI_Siber_Diger',
}

for key, sname in sheet_map.items():
    if sname not in wb.sheetnames:
        print(f"\n{key} ({sname}): SHEET YOK")
        continue
    ws = wb[sname]
    row_count = ws.max_row - 1
    print(f"\n{key} ({sname}): {row_count} satir veri")
    # Ilk kolondaki ilk 20 fikir basligini goster
    for r in range(2, min(ws.max_row + 1, 22)):
        v = ws.cell(row=r, column=1).value
        if v:
            v_str = str(v)
            print(f"  Row {r:2d}: {v_str[:90]}")
