"""
A1 sheet'teki hacim ifadelerini DfS Dogrulama gercek toplamiyla esitle:
- Dil ve Sozluk: 500K+ -> ~430K
- Birim Cevirici: 250K+ -> ~175K
- Zaman Araclari: 200K+ -> ~95K
"""
from openpyxl import load_workbook
from pathlib import Path

XLSX = Path("/Users/Erdo/Desktop/Claude Projects/Turkcel/TURKCELL_TRAFIK_FIRSATLARI.xlsx")
wb = load_workbook(XLSX)

ws = wb["01_A1_Hesaplama_Fikirleri"]

targets = {
    "dil ve sozluk": ("500K+", "~430K"),
    "dil-ve-sozluk": ("500K+", "~430K"),
    "birim cevirici": ("250K+", "~175K"),
    "birim-cevirici": ("250K+", "~175K"),
    "zaman araclari": ("200K+", "~95K"),
    "zaman-araclari": ("200K+", "~95K"),
}

changes = []
for r in range(2, ws.max_row + 1):
    title = str(ws.cell(row=r, column=1).value or "").lower()
    for key, (old, new) in targets.items():
        if key in title:
            for c in range(2, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v and isinstance(v, str) and old in v:
                    new_v = v.replace(old, new)
                    ws.cell(row=r, column=c, value=new_v)
                    changes.append(f"Row {r} Col {c} ({key}): {old} -> {new}")
            break

wb.save(XLSX)
print("=== Hacim Duzeltmeleri ===")
for c in changes:
    print("  -", c)
print(f"\nToplam {len(changes)} hucre guncellendi.")
