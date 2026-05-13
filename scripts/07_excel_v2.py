"""
Excel v2: Yol haritasi cikti, niche cluster'lar eklendi, fikir bazli kelime listesi sheet'leri.
"""
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

CSV_PATH = "/Users/Erdo/Desktop/Claude Projects/Turkcel/www.turkcell.com.tr-content-gap-subdomains-t_2026-05-13_14-14-31.csv"
OUTPUT_DIR = "/Users/Erdo/Desktop/Claude Projects/Turkcel/output"
EXCEL_PATH = f"{OUTPUT_DIR}/TURKCELL_TRAFIK_FIRSATLARI.xlsx"

df_raw = pd.read_csv(CSV_PATH, low_memory=False)
competitors = ["turktelekom.com.tr", "vodafone.com.tr", "mediamarkt.com.tr",
               "vatanbilgisayar.com", "turk.net", "pttcell.com.tr"]
df = df_raw[df_raw["www.turkcell.com.tr/: URL"].isna()].copy()

def best_comp_info(row):
    best_pos = 999
    best_comp = None
    best_url = None
    best_traf = 0
    for comp in competitors:
        pos = row.get(f"{comp}/: Organic Position", None)
        if pd.notna(pos) and pos < best_pos:
            best_pos = pos
            best_comp = comp
            best_url = row.get(f"{comp}/: URL", None)
            best_traf = row.get(f"{comp}/: Organic Traffic", 0)
    return pd.Series([best_comp, best_pos if best_pos < 999 else None, best_url,
                      best_traf if pd.notna(best_traf) else 0])

df[["best_competitor", "best_position", "best_url", "best_traffic"]] = df.apply(best_comp_info, axis=1)
df = df[df["best_competitor"].notna()].copy()

# Niche cluster regex (06 ile ayni)
NICHE_PATTERNS = {
    "01_HESAPLAMA_DUNYA": [r"\bsayac\b", r"\bsay[ıi]c[ıi]\b", r"\bhesap\w+", r"\bhesaplay\w+",
        r"\b(yüzde|yuzde|oran|fark|toplam|farkı) hesap", r"\b(kdv|otv|ötv|stopaj|damga vergisi|gelir vergisi|kurumlar vergisi)\b",
        r"\b(brüt|net|maaş|maas|asgari) ?(ücret|hesap)", r"\b(kıdem|kidem|ihbar) tazmin",
        r"\b(emekli|emeklilik) (yaşı|hesap|maaş)", r"\b(yıllık izin|yillik izin|izin g[üu]nü|izin g[üu]nleri)",
        r"\b(fazla mesai|gece mesai|hafta sonu mesai)", r"\b(yıl sonu notu|not ortalama|gpa|aort|üni\.? ort)",
        r"\b(bmi|vücut kitle|ideal kilo|metabolizma|kalori|protein|su tüketim)",
        r"\b(hamilelik|gebelik|ovulasyon|regl|aşı takvim|menapoz)",
        r"\b(doğum tarihi|dogum tarihi|kaç haftalık|kac haftalik)\b",
        r"\b(burç|burc|yükselen|astroloji|yıldız haritası|doğum haritası)",
        r"\b(kredi|taksit|faiz|vade) hesap", r"\b(iban|swift) (hesap|sorgu|olu[şs]tur)",
        r"\b(döviz|altın|gümüş|euro|dolar) (kuru|fiyat|hesap|çevir)",
        r"\b(benzin|motorin|lpg|elektrik|do[ğg]algaz) (fiyat|hesap|tüketim)",
        r"\b(yol|mesafe|km|rota) hesap", r"\b(saat fark|zaman fark|gmt|utc)",
        r"\b(yıl|gün|hafta|ay) (sayac|hesap)", r"\b(geri sayım|geri sayim|countdown)",
        r"\b(rastgele|şans|piyango|loto)", r"\b(faktöriyel|yuvarla|yüzdelik dilim)",
        r"\b(kira|aidat|tapu|emlak vergi) hesap", r"\bdaire hesap", r"\bnoter (ücret|hesap)",
        r"\btoki (başvuru|hesap)"],
    "02_BIRIM_DONUSUM": [r"\b(inç|inc|inch) (kaç|kac) ?(cm|ekran)",
        r"\b(cm|metre|km|mm|mil) (kaç|kac)", r"\b(gb|mb|tb|kb) (kaç|kac)",
        r"\b(kg|gram|libre|pound|ton) (kaç|kac)", r"\b(litre|ml|gallon|fincan|bardak|kase) (kaç|kac)",
        r"\b(\d+) ?(inç|cm|kg|gram|metre|km|ml|litre|tl|usd|eur|gb|mb)",
        r"\b(celsius|fahrenheit|kelvin|santigrat)",
        r"\b(saat|dakika|saniye) (çevir|hesap|kaç)", r"\bderece (çevir|kaç)"],
    "03_TATIL_BAYRAM_OZEL_GUN": [r"\bresmî tatil\b", r"\bresmi tatil\b", r"\btatil mi\b",
        r"\b(kurban|ramazan|şeker|berat|mevlit|miraç|regaip) (bayram|kandil)",
        r"\barefe\b", r"\bayrefe\b", r"\barife\b",
        r"\b(yılbaş[ıi]|yilbasi|yeni yıl|yeni yil)",
        r"\b(cumhuriyet|zafer|gençlik|egemenlik|barış|çocuk) bayram",
        r"\b(29 ekim|23 nisan|19 mayıs|30 ağustos|10 kasım|18 mart)",
        r"\b(yarı yıl|yariyil|yari yil|sömestr|somestr|yaz tatili|kış tatili)",
        r"\b(black friday|kara cuma|cyber monday|valentine)",
        r"\b(anneler|babalar|sevgililer|öğretmenler|çocuklar|kadınlar) g[uü]n[uü]",
        r"\b(8 mart|14 şubat|24 kasım|1 mayıs)",
        r"\b(nevruz|hıdırellez|hidirellez)",
        r"\b(geri sayım|kaç gün kaldı) (ramazan|kurban|bayram|yılbaş|tatil)"],
    "04_MESAJ_SOZ_KART": [r"\b(mesaj|söz|sozler|şiir|siir|söy|yazi|notu) ?(ı|i|ları|leri|larım)?\b.*\b(günü|bayram|sevgili|anne|baba|eş|kardeş|arkadaş|öğretmen|patron)",
        r"\b(en güzel|guzel|duygusal|romantik|kısa|uzun|anlamlı|anlamli|komik|esprili) (mesaj|söz|sözler|şiir)",
        r"\b(günaydın|iyi geceler|tatlı uykular|hayırlı sabahlar) (mesaj|söz)",
        r"\b(taziye|geçmiş olsun|şifa|başsağlığı) (mesaj|söz)",
        r"\b(düğün|nişan|kına|söz|bebek|bilek) (kart|davetiye|mesaj)",
        r"\b(doğum günü|yas günü) (hediyesi|mesaj|kart)"],
    "05_TANIM_NEDIR": [r"\bnedir\b", r"\bne demek\b", r"\bne anlama gelir\b", r"\bne işe yarar\b",
        r"\bne ise yarar\b", r"\baçılımı\b", r"\bacilimi\b", r"\bkısaltma\b"],
    "06_NASIL_YAPILIR": [r"\bnas[ıi]l\b"],
    "07_NE_ZAMAN": [r"\bne zaman\b", r"\bhangi g[uü]n\b", r"\bhangi tarih\b", r"\bhangi ay\b"],
    "08_KAC": [r"\bkaç\b", r"\bkac\b"],
    "09_SOSYAL_MEDYA_APP": [r"\b(whatsapp|wp|whatsapp web)\b", r"\b(instagram|insta|ig)\b",
        r"\b(facebook|fb)\b", r"\b(twitter|x|x giriş)\b", r"\b(telegram|telegram web)\b",
        r"\b(snapchat|snap)\b", r"\b(tiktok|tik tok)\b", r"\b(linkedin|reddit|pinterest|discord)\b",
        r"\b(youtube|yt)\b", r"\b(netflix|disney\+|exxen|tabii|blutv|gain)\b",
        r"\b(spotify|apple music|fizy|tidal)\b", r"\b(bip|whatsapp|signal|wechat)\b",
        r"\b(zoom|teams|skype|meet)\b", r"\b(canva|figma|photoshop|illustrator)\b",
        r"\b(chatgpt|gemini|claude|grok|copilot|deepseek)\b"],
    "10_TELEFON_CIHAZ": [r"\biphone\b", r"\bsamsung\b", r"\bxiaomi\b", r"\bhuawei\b", r"\boppo\b",
        r"\bvivo\b", r"\bredmi\b", r"\bhonor\b", r"\brealme\b", r"\boneplus\b",
        r"\b(akıllı saat|akilli saat|smart watch)", r"\b(kulaklık|kulaklik|airpods|earbuds)",
        r"\b(şarj|sarj) (cihaz|aleti|adaptör)", r"\b(power bank|powerbank|taşınabilir şarj)",
        r"\b(kılıf|kilif|ekran koruyucu|cam)", r"\b(sim|esim|nano sim|micro sim)",
        r"\b(imei|seri no|model no)"],
    "11_TELEFON_AYAR_SORUN": [r"\b(güvenli arama|safe search|family link)",
        r"\b(nfc|airdrop|airtag|airpods)", r"\b(ekran görüntüsü|screenshot|ekran kaydı|screen record)",
        r"\b(face id|touch id|parmak izi|şifre|sifre)",
        r"\b(yedek|backup|icloud|google drive|onedrive)",
        r"\b(format|sıfırla|fabrika ayar)", r"\b(güncelleme|update|ios|android)",
        r"\b(virüs|virus|malware|antivirus)", r"\b(konum|location|gps)"],
    "12_INTERNET_MODEM_WIFI": [r"\b(internet|wi[- ]?fi|modem|router)",
        r"\b(hız testi|hiz testi|speed test|ping)", r"\b(192\.168\.|10\.0\.|172\.16\.)",
        r"\bdns\b", r"\bip adres", r"\bping\b", r"\bvpn\b",
        r"\b(fiber|adsl|vdsl|kablo internet)"],
    "13_EDEVLET_SORGULAMA": [r"\be-devlet\b", r"\bedevlet\b", r"\btc kimlik\b", r"\bnüfus\b",
        r"\b(pasaport|ehliyet|sgk|emekli sandığı)",
        r"\b(borç sorgu|borc sorgu|trafik cezası|hgs|ogs|mtv)",
        r"\b(plaka sorgu|araç sorgu|arac sorgu|muayene tarihi)",
        r"\b(barkod|qr|karekod) sorgu", r"\b(adli sicil|sabıka kayıt|ikametgâh)",
        r"\b(askerlik|bedelli|sevk|tecil)", r"\b(başvuru|basvuru|form|dilekçe|dilekce)",
        r"\b(noter|tapu|gemlik|nüfus müdürlüğü)",
        r"\b(vergi numarası|gelir vergisi|kurumlar vergisi)",
        r"\b(posta kodu|alan kodu|plaka kodu|telefon kodu)"],
    "14_OTOMOTIV": [r"\b(araba|otomobil|araç|arac|otomotiv|otomotİv)",
        r"\b(motor|motorsiklet|scooter|bisiklet)",
        r"\b(plaka|trafik|muayene|sigorta|kasko)",
        r"\b(sürücü|surucu|ehliyet) (kurs|sınav|belge)",
        r"\b(benzinli|dizel|hibrit|elektrikli) (araba|otomobil)",
        r"\b(yedek parça|yedek parca|fren|debriyaj|amortisör)"],
    "15_EGITIM_SINAV": [r"\b(yks|tyt|ayt|lgs|kpss|ales|yds|yökdil)",
        r"\b(ösym|osym|meb|yök|yok)",
        r"\b(sınav takvim|sinav takvim|sınav tarih|tercih kılavuz)",
        r"\b(üniversite|universite|lise|ortaokul|ilkokul) (puan|tercih|sınav)",
        r"\b(ders çalışma|ders calisma|konu anlatım|özet|ozet|soru çözüm)",
        r"\b(eba|btk akademi|udemy|coursera)", r"\b(diploma|sertifika|burs)"],
    "16_BURC_ASTROLOJI": [r"\b(burç|burc|astroloji|yıldız|zodyak|horoskop)",
        r"\b(koç|boğa|ikizler|yengeç|aslan|başak|terazi|akrep|yay|oğlak|kova|balık) burc",
        r"\b(yükselen|yukselen|ay|güneş|gunes) burc",
        r"\b(doğum haritası|dogum haritasi)",
        r"\b(uyumlu burç|uyumsuz burç|burç eşleşmesi)"],
    "17_SPOR": [r"\b(maç|mac|skor|fikstür|puan durumu|lig|şampiyona)",
        r"\b(fenerbahçe|fenerbahce|galatasaray|beşiktaş|besiktas|trabzonspor)",
        r"\b(milli takım|euro|dünya kupası|şampiyonlar ligi|avrupa ligi)",
        r"\b(futbol|basketbol|voleybol|hentbol|tenis|yüzme)",
        r"\b(formula|f1|nascar)", r"\b(antrenman|kondisyon|fitness|yoga|pilates)"],
    "18_OYUN_REHBER": [r"\b(gta|pubg|lol|valorant|fortnite|minecraft|fifa|efootball)",
        r"\b(playstation|ps[345]|xbox|nintendo|switch)",
        r"\b(steam|epic games|origin|battle\.net)",
        r"\b(roblox|among us|free fire|call of duty|cod)",
        r"\bkonsol\b", r"\boyun (kolu|seti|simülasyon)"],
    "19_DIZI_FILM": [r"\b(dizi|film|sezon|bölüm|bolum) (konusu|oyuncu|özet|özet|yayın)",
        r"\b(yapımcı|yönetmen|senaryo|kanal)",
        r"\bbilim kurgu\b", r"\bromantik komedi\b", r"\bgerilim\b"],
    "20_KLAVYE_SEMBOL": [r"\b(et işareti|et isareti|@|#|&|%|\\$)",
        r"\b(noktalama|virgül|nokta|ünlem|soru işareti|tırnak|ayraç)",
        r"\b(klavye|tuş|tus|combo|kısayol|kisayol)",
        r"\b(emoji|emojikler|ifade)", r"\b(büyük harf|kucuk harf|caps lock)"],
    "21_KELIME_YAZIM_DIL": [r"\b(eşanlamlı|esanlamli|zıt anlamlı|zit anlamli|terim sözlük)",
        r"\b(noktalı virgül|ki bağlacı|de da bağlacı|herşey her şey)",
        r"\b(ingilizce|almanca|fransızca|arapça|rusça|çince) (öğren|ogren|cümle)",
        r"\b(çeviri|ceviri|sözlük|sozluk|kelime anlam)",
        r"\b(yazım|yazim) (kuralları|kontrol|kılavuzu|kilavuzu)"],
    "22_HAYVANLAR": [r"\b(köpek|kopek|kedi|kuş|kus|balık|balik|kuş bakımı)",
        r"\b(mama|yem|kafes|tasma)", r"\b(yavru|yetişkin|yaşlı) (kedi|köpek)"],
    "23_BITKILER_BAHCE": [r"\b(çiçek|cicek|bitki|fidan|tohum|sera|saksı)",
        r"\b(orkide|kaktüs|gül|menekşe|begonya|sakız|aloe)",
        r"\b(toprak|gübre|sulama|budama|yetiştirme)"],
    "24_RUYA_TABIRI": [r"\b(rüya|ruya) ?(tabir|görmek|gormek|yorum)", r"\b(rüyada|ruyada)"],
    "25_ASTRONOMI": [r"\b(ay tutulması|güneş tutulması|gunes tutulmasi|kuyruklu yıldız)",
        r"\b(gezegen|asteroid|takım yıldız|samanyolu)",
        r"\b(uzay|nasa|spacex|mars|venüs|jüpiter|satürn|uranüs|neptün)",
        r"\b(meteor|şahap|göktaşı|göktaşı yağmuru)"],
    "26_KONSER_ETKINLIK": [r"\b(konser|festival|etkinlik|tiyatro|opera|bale)",
        r"\b(bilet|ticket|biletix|passo)", r"\b(sergi|müze|fuar)"],
    "27_SEYAHAT_ULASIM": [r"\b(uçak|ucak) (bileti|firma|kalkış|iniş)",
        r"\b(otobüs|otobus) (bileti|firma|sefer)",
        r"\b(tren|hızlı tren|yht) (bileti|sefer)",
        r"\b(otel|tatil köyü|all inclusive|kamp)",
        r"\b(vize|pasaport|gümrük|seyahat sigortası)",
        r"\b(yurt dışı|yurtdışı|yurtiçi) (seyahat|tatil|gezi)",
        r"\b(navlun|kargo|posta) (takip|fiyat)"],
    "28_IS_BASVURU_CV": [r"\b(cv|özgeçmiş|ozgecmis) (hazırlama|hazirlama|şablon|sablon|örnek|ornek)",
        r"\b(iş başvuru|is basvuru|işe alım|mülakat|mulakat) (sorular|tüyolar)",
        r"\b(staj|kariyer|iş ilanları|is ilanlari)",
        r"\b(linkedin|linkedlin) (profil|optimizasyon)",
        r"\b(motivasyon mektup|kapak yazısı|niyet mektup)"],
    "29_FINANS_BANKA_YATIRIM": [r"\b(banka|bankacılık|bankacilik|atm|kart)",
        r"\b(kredi notu|findeks|sicil|borç sorgu|borc sorgu)",
        r"\b(yatırım|yatirim|borsa|hisse senedi|bist|tahvil|fon|portföy)",
        r"\b(kripto|bitcoin|ethereum|nft|blockchain|metaverse)",
        r"\b(altın fiyatı|gümüş fiyatı|döviz kuru|dolar tl|euro tl)",
        r"\b(faiz oran|enflasyon|tüfe|enflasyon hesap|reeskont)",
        r"\b(emekli|individual emeklilik|bes)"],
    "30_HUKUK_KANUN": [r"\b(kanun|yasa|tüzük|yönetmelik)",
        r"\b(boşanma|nafaka|velayet|alacak)", r"\b(iş hukuku|sözleşme|fesih)",
        r"\b(miras|tereke|veraset)", r"\b(ceza|tazminat|şikayet|tüketici)"],
    "31_EV_DEKOR_MOBILYA": [r"\b(mobilya|koltuk|yatak|masa|sandalye|gardırop)",
        r"\b(perde|halı|hali|kilim)",
        r"\b(boya|duvar kağıdı|duvar kagidi|tadilat|kalıcı)",
        r"\b(banyo|mutfak|salon|yatak odası) (dekor|tasarım)"],
    "32_BEAUTY_KOZMETIK": [r"\b(makyaj|fondoten|ruj|maskara|allık|göz kalemi)",
        r"\b(cilt bakım|nemlendir|krem|serum|ton|spf)",
        r"\b(saç bakım|saç boya|saç şekillendir|saç tipi)",
        r"\b(tırnak|manikür|pedikür|ojesi)", r"\b(parfüm|deodorant|koku)"],
    "33_COCUK_BEBEK": [r"\bbebek bakım", r"\bemzir", r"\bemzirme",
        r"\b(çocuk gelişimi|cocuk gelisimi|montessori|ana sınıfı)",
        r"\b(beslenme önerileri|mama tarif|pure)"],
    "34_SAGLIK_VUCUT": [r"\b(belirti|semptom|hastalık|hastalik|teşhis|teshis)",
        r"\b(ağrı|agri|sancı|sanci)",
        r"\b(soğuk algınlığı|grip|nezle|sinüzit|alerji)",
        r"\b(diyabet|hipertansiyon|kolesterol|tansiyon)",
        r"\b(hamilelik|gebelik|loğusalık|emzirme|bebek)",
        r"\b(diş|dis|kulak|göz|gozluk) (ağrı|temizlik|implant|muayene)",
        r"\b(astım|astim|akciğer|kalp|böbrek|karaciğer|safra)",
        r"\b(panik atak|anksiyete|depresyon|stres|uyku)",
        r"\b(vitamin|mineral|takviye|d3|b12|c vitamini)",
        r"\b(diyet|kilo verme|kilo alma|protein)"],
    "35_YEMEK_TARIF": [r"\btarif(i)?\b", r"\bnasıl yap[ıi]l[ıi]r\b",
        r"\b(corba|çorba|salata|tatlı|tatli|kek|pasta|börek|borek|pilav|et yemeği|tavuk yemeği)",
        r"\b(makarna|mantı|köfte|kebap|kavurma|musakka|dolma|sarma)",
        r"\b(çay|cay|kahve|smoothie|kokteyl|içecek)"],
}

def categorize_multi(keyword):
    kw = str(keyword).lower()
    for cname, patterns in NICHE_PATTERNS.items():
        if any(re.search(p, kw) for p in patterns):
            return cname
    return "ZZ_DIGER"

df["niche"] = df["Keyword"].apply(categorize_multi)
df_sorted = df.sort_values("Volume", ascending=False)

# Excel olusturma
wb = Workbook()
wb.remove(wb.active)

hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="1F4E78")
sub_hdr_font = Font(bold=True, size=11)
sub_hdr_fill = PatternFill("solid", fgColor="D9E1F2")
prio_fill = PatternFill("solid", fgColor="C6EFCE")
mid_fill = PatternFill("solid", fgColor="FFEB9C")
warn_fill = PatternFill("solid", fgColor="FCE4D6")
border = Border(left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"))

def write_df(ws, dataf, start_row=1, freeze=True, col_widths=None):
    for r_idx, row in enumerate(dataframe_to_rows(dataf, index=False, header=True), start_row):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if r_idx == start_row:
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
    if freeze:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    if col_widths:
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

# ========= 00 OZET =========
ws = wb.create_sheet("00_Yonetici_Ozeti")
rows = [
    ["TURKCELL.COM.TR - YENİ TRAFİK FIRSATLARI v2 (Hacim filtresi yok)", "", ""],
    ["Tarih", "13 Mayıs 2026", ""],
    ["Veri", "Ahrefs Content Gap (75.000 kw) + DataForSEO doğrulama (TR/tr)", ""],
    ["", "", ""],
    ["TEMEL SAYILAR", "", ""],
    ["Toplam Turkcell-yok rakip-var fırsat", 60715, ""],
    ["Volume ≥ 10", 60715, "Hepsi"],
    ["Volume ≥ 100", 48087, ""],
    ["Volume ≥ 500", 12826, ""],
    ["Volume ≥ 1000", 7247, ""],
    ["", "", ""],
    ["EN BÜYÜK TEK-SAYFA FIRSATLAR (DfS doğrulanmış)", "Aylık Hacim", "KD"],
    ["hesap makinesi", 2740000, "—"],
    ["kurban bayramı ne zaman", 1220000, 11],
    ["internet hız testi", 1000000, 34],
    ["anneler günü ne zaman", 673000, 8],
    ["yüzde hesaplama (MEVCUT)", 673000, 12],
    ["babalar günü ne zaman", 450000, "—"],
    ["29 ekim cumhuriyet bayramı", 450000, "—"],
    ["asgari ücret 2026 (+%210K trend)", 368000, 9],
    ["yaş hesaplama (MEVCUT)", 301000, 9],
    ["KDV hesaplama", 201000, 19],
    ["kıdem tazminatı hesaplama", 201000, 27],
    ["yükselen burç hesaplama", 201000, 10],
    ["doğum günü mesajları", 201000, "—"],
    ["güzel sözler", 165000, 11],
    ["sgk işe giriş", 135000, 20],
    ["günaydın mesajları", 135000, 4],
    ["chatgpt türkçe", 135000, 16],
    ["anlamlı sözler", 110000, 4],
    ["hgs bakiye sorgulama", 110000, 19],
    ["", "", ""],
    ["RAPOR YAPISI", "", ""],
    ["Sheet 01: A1_Hesaplama_Fikirleri", "Kendi araştırma A1 - calculator + KW listesi", ""],
    ["Sheet 02: A2_Tatil_Bayram_Fikirleri", "A2 - resmi tatil & bayram hub", ""],
    ["Sheet 03: A3_Ozel_Gun_Mesaj_Fikirleri", "A3 - özel gün + mesaj/söz/kart", ""],
    ["Sheet 04: A4_Telekom_Sorgu_Fikirleri", "A4 - telekom yardım + sorgulama", ""],
    ["Sheet 05: A5_AI_Siber_Diger_Fikirleri", "A5/A6 - AI rehberi, siber güvenlik, dizi/film, CV", ""],
    ["Sheet 06: B1_Cluster_Ozeti", "35 niche cluster özet tablo", ""],
    ["Sheet 07: B2_Rakip_URL_Patternleri", "Vodafone/TT/turk.net trafik kaynakları", ""],
    ["Sheet 08: DfS_Dogrulama", "DataForSEO Mayıs 2026 doğrulanmış 100+ KW", ""],
    ["Sheet 09+: niche_* clusterler", "Her cluster için top 200 keyword listesi (low volume dahil)", ""],
    ["Son sheet: TUM_60K_FIRSATLAR", "Tüm 60.715 fırsat tek tabloda", ""],
]
for r_idx, row in enumerate(rows, 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == 1:
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.fill = hdr_fill
        elif val in ["TEMEL SAYILAR", "EN BÜYÜK TEK-SAYFA FIRSATLAR (DfS doğrulanmış)", "RAPOR YAPISI"]:
            cell.font = sub_hdr_font
            cell.fill = sub_hdr_fill
ws.column_dimensions["A"].width = 60
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 40
ws.merge_cells("A1:C1")

# ========= A1: HESAPLAMA FİKİRLERİ =========
ws = wb.create_sheet("01_A1_Hesaplama_Fikirleri")
a1 = pd.DataFrame([
    # FIKIR 1 - HESAP MAKINESI HUB
    ["FIKIR 1: Genel Calculator Hub", "/hesaplama-araclari/", "hesap makinesi", "2.740.000", "düşük", "informational", "★★★",
     "hesap makinesi, bilimsel hesap makinesi (10K), fonksiyonlu hesap makinesi (1.5K), casio hesap makinesi (6.9K), casio bilimsel hesap makinesi (600), hesap makinesi fiyatları (1.1K), hesaplama makinesi (600), hesap makinesi casio (500), casıo hesap makinesi (600)",
     "Vatan Bilgisayar /hesap-makinesi tek sayfa 31.802 trafik"],
    # FIKIR 2 - YAS GENISLETME
    ["FIKIR 2: /yas-hesaplama genişletme (MEVCUT)", "/yas-hesaplama (mevcut)", "yaş hesaplama", "301.000", "9", "informational", "★★★",
     "yaş hesaplama, kaç yaşındayım, doğum tarihinden yaş, kedi yaşı hesaplama (12K), köpek yaşı hesaplama, iki tarih arası gün, gün gün yaş hesaplama, ay ay yaş, bebek yaş hesaplama",
     "Mevcut sayfa; kedi/köpek yaşı widget ekle"],
    # FIKIR 3 - YUZDE GENISLETME
    ["FIKIR 3: /yuzde-hesaplama genişletme (MEVCUT)", "/yuzde-hesaplama (mevcut)", "yüzde hesaplama", "673.000", "12", "informational", "★★★",
     "yüzde hesaplama, yüzde artış, yüzde azalış, yüzde indirim, yüzde değişim, yüzde fark, yüzde kazanç, yüzde formülü, kdv yüzdesi, kar yüzdesi, yüzdelik dilim",
     "Mevcut sayfa; senaryo bazlı genişlet"],
    # FIKIR 4 - BRUT NET
    ["FIKIR 4: /brutten-nete-maas-hesaplama (Vodafone modeli)", "/brutten-nete-maas-hesaplama/", "brüt net hesaplama", "8.100 (küme 236K)", "30", "informational", "★★★",
     "brütten nete maaş (27K), brüt maaş (25K), netten brüte maaş (20K), netten brüt (15K), net maaş (14K), brütten net (13K), brütten nete (13K), aylık maaş (1.5K), brüt maaştan net (1.1K), maaş hesaplama 2026 (2.1K), asgari ücret 2026 (368K!)",
     "Vodafone tek sayfa 7.365 trafik, 110 keyword"],
    # FIKIR 5 - KDV
    ["FIKIR 5: /kdv-hesaplama/", "/kdv-hesaplama/", "kdv hesaplama", "201.000", "19", "informational", "★★★",
     "kdv hesaplama, kdv formülü, kdv dahil, kdv hariç, kdv oranları 2026, kdv nedir, %20 kdv, %10 kdv, %8 kdv, %1 kdv, tevkifat hesaplama, kdv tevkifatı, ötv nedir (8.2K)",
     "İşletme/freelancer kitle, Paycell köprüsü"],
    # FIKIR 6 - KIDEM IHBAR
    ["FIKIR 6: /kidem-ve-ihbar-tazminati-hesaplama/", "/kidem-ihbar-tazminati/", "kıdem tazminatı hesaplama", "201.000 + 27.100", "27 / 25", "informational/commercial", "★★★",
     "kıdem tazminatı hesaplama, kıdem tazminatı tavanı 2026, kıdem hesaplama, ihbar tazminatı hesaplama, ihbar süresi, iş sözleşmesi fesih, haklı fesih, kıdem ihbar farkı",
     "HR/iş hayatı kitle"],
    # FIKIR 7 - CALISAN HAKLARI MINI-HUB
    ["FIKIR 7: /calisan-haklari-hesaplayicilari/ mini-hub", "/calisan-haklari/", "yıllık izin hesaplama", "9.900 + 9.900 + 49.500 + 135K", "4 / —", "informational", "★★★",
     "yıllık izin (9.9K), yıllık izin kaç gün, fazla mesai (9.9K), gece zammı, bayram mesaisi, sgk hizmet dökümü (49.5K), sgk işe giriş (135K), sgk işe giriş bildirgesi (11K), emeklilik yaşı (3.6K)",
     "Hub altında 8-10 alt-calculator"],
    # FIKIR 8 - KREDI TAKSIT
    ["FIKIR 8: /kredi-ve-taksit-hesaplama/", "/kredi-hesaplama/", "kredi hesaplama", "1.220.000", "45 (zor)", "informational", "★",
     "kredi hesaplama, taksit hesaplama (1.9K), cayma bedeli hesaplama (3K, KD 0), vade farkı hesaplama (2.4K), iskonto hesaplama (1K), youtube gelir hesaplama (2.1K), kira artış oranı (6.6K), tüfe hesaplama, enflasyon hesaplama, mevduat faizi",
     "Yüksek hacim zor ama uzun kuyruk + telekom-özel cayma bedeli"],
    # FIKIR 9 - VERGI
    ["FIKIR 9: /vergi-hesaplama-hub/", "/vergi-hesaplama/", "kdv hesaplama (ana)", "201.000 + alt", "19+", "informational", "★★",
     "kdv hesaplama (201K), gelir vergisi, gelir vergisi dilimleri 2026, damga vergisi, stopaj, kurumlar vergisi, vergi iadesi (140), mtv sorgulama (22.2K), motorlu taşıt, vergi mükellefiyet (1.7K), vergi levhası ücreti (600)",
     "KOBİ/işletme blog"],
    # FIKIR 10 - SAGLIK
    ["FIKIR 10: /saglik-hesaplayicilari/ mini-hub", "/saglik-hesaplayicilari/", "bmi/kalori/regl/ovulasyon/hamilelik", "100K+ toplam", "0-12", "informational", "★★",
     "bmi hesaplama (8.1K), kalori hesaplama (74K), ideal kilo, hamilelik haftası (2.4K), gebelik haftası, kaç haftalık hamileyim, ovulasyon hesaplama (6.6K), yumurtlama günü, regl hesaplama (18.1K), regl takvimi, adet takvimi, tahmini doğum tarihi, su tüketimi",
     "Kadın+sağlık kitle"],
    # FIKIR 11 - ASTROLOJI
    ["FIKIR 11: /burc-ve-astroloji-hesaplayicilari/", "/burc-hesaplama/", "yükselen burç hesaplama", "201.000", "10", "informational", "★★",
     "yükselen burç hesaplama, burç hesaplama, ay burcu, güneş burcu, çin burcu, doğum haritası çıkarma, burç uyumu, uyumlu burçlar (260), kahve falı (18.1K), tarot falı, el falı, yıldız falı (390), günlük/haftalık/aylık burç yorumu, 12 burç özellikleri",
     "Eğlence cluster, mega hacim"],
    # FIKIR 12 - YAZMA ARACLARI
    ["FIKIR 12: /yazma-araclari/ mini-hub", "/yazma-araclari/", "karakter sayacı / kelime sayacı", "14.800 + 18.100", "3 / 12", "informational", "★★★",
     "karakter sayacı (14.8K), karakter sayma, kelime sayacı (18.1K), kelime sayma, büyük harf küçük harf çevirici, türkçe karakter düzeltici, klavye sembolleri, et işareti @ (72K!), klavyede @ nasıl yapılır (5.6K+3.5K+3.1K), bilgisayarda @ (5.6K), noktalama işaretleri (40.5K), ters yazma, qr kod oluşturucu, şifre oluşturucu, lorem ipsum, sayı yazıyla",
     "Sosyal medya+öğrenci kitle, en düşük KD"],
    # FIKIR 13 - DIL SOZLUK
    ["FIKIR 13: /dil-ve-sozluk-araclari/", "/sozluk/", "noktalama işaretleri / iyelik eki / isim anlamı", "40.5K + 14.8K + 6.6K", "4 / 5 / 3", "informational", "★★",
     "noktalama işaretleri (40.5K), iyelik eki (14.8K), tdk sözlük (368K), yazım kılavuzu (1.9K), isim anlamı (6.6K), kısaltma sözlüğü, eşanlamlı/zıt anlamlı, atasözleri ve deyimler, ki bağlacı yazımı, de da bağlacı, çeviri (cluster 181K), ingilizce öğrenme, sıfırdan ingilizce (800)",
     "Eğitim cluster düşük KD"],
    # FIKIR 14 - BIRIM CEVIRICI
    ["FIKIR 14: /birim-cevirici/ mini-hub", "/birim-cevirici/", "inç-cm + döviz çevirici", "250K+ küme", "0-32", "informational", "★★",
     "55 inç kaç cm (21K), 65 inç kaç cm (17K), 43 inç kaç cm (10K), 50 inç kaç ekran (13K), 55 inç kaç ekran (13K), tv inç hesaplama (1K), iphone 11 kaç cm (5K), 75 inç kaç cm (4.8K), 140 ekran kaç inç (2.9K), kg libre, dolar tl çevirici, döviz çevirici (74K), celsius fahrenheit, mil km, gb mb kaç, 1 gb kaç mb (telekom)",
     "GB-MB telekom-özel; TV inç MediaMarkt dominant ama Turkcell TV+ köprüsü"],
    # FIKIR 15 - ZAMAN ARACLARI
    ["FIKIR 15: /zaman-araclari/ mini-hub", "/zaman-araclari/", "bugün ayın kaçı + saat farkı + geri sayım", "200K+ küme", "—", "informational", "★★",
     "bugün ayın kaçı (90.5K), bugün hangi gün, saat kaç türkiye (2.9K), şu an saat kaç, abd saat farkı (390), türkiye new york saat farkı, kurban bayramına kaç gün kaldı, ramazan'a kaç gün kaldı (13K), bayrama kaç gün kaldı (2.6K), bu hafta kaçıncı hafta, yılın kaçıncı günü, hicri takvim 2026, miladi hicri çevirici",
     "Geri sayım widget'ları spike lokomotifi"],
], columns=["Fikir", "URL_Onerisi", "Ana_Keyword", "Aylik_Hacim", "KD", "Intent", "Oncelik",
            "Oynanabilecek_KW_Listesi", "Notlar"])
write_df(ws, a1, col_widths={"A": 40, "B": 30, "C": 28, "D": 16, "E": 10, "F": 16, "G": 10, "H": 80, "I": 38})
# Renk
for r_idx in range(2, len(a1) + 2):
    p = ws.cell(row=r_idx, column=7).value
    if p == "★★★":
        for c in range(1, 10):
            ws.cell(row=r_idx, column=c).fill = prio_fill
    elif p == "★★":
        for c in range(1, 10):
            ws.cell(row=r_idx, column=c).fill = mid_fill

# ========= A2: TATIL BAYRAM FIKIRLERI =========
ws = wb.create_sheet("02_A2_Tatil_Bayram_Fikirleri")
a2 = pd.DataFrame([
    ["A2.1 RESMI TATIL HUB", "/resmi-tatiller/", "resmi tatiller 2026", "22.200 (+%8303 trend)", "3", "informational", "★★★",
     "resmi tatiller 2026, resmi tatil günleri (6.6K), 2026 15 tatil ne zaman (14K), 9 gün tatil 2026, 10 gün tatil, bağlı tatil 2026, kasım tatili 2025 (16K), yarı yıl tatili 2026 (11K), sömestr ne zaman (11K), yaz tatili 2026 (11K), okul tatilleri 2026, meb tatil takvimi, bugün tatil mi (2K), pazartesi okullar tatil mi (2.1K)",
     "Vodafone /resmi-tatiller 30.896 trafik"],
    ["A2.2 KURBAN BAYRAMI 2026", "/kurban-bayrami-2026/", "kurban bayramı ne zaman", "1.220.000 (Mar-May spike)", "11", "informational", "★★★",
     "kurban bayramı ne zaman, 2025 kurban bayramı (77K), kurba bayram ne zaman 2025 (41K), 2026 kurbanlık bayramı diyanet (7.3K), kurban bayramı hangi gün, kurban bayramı tarihi, kurban bayramı hangi ayda (3.4K), kurban bayramı kaçında (1.4K), kurban bayramı ayın kaçına (1.9K), kurban bayramı tatil mi, arefe günü tatil mi (5.8K+2.5K), arefe ne zaman (5.4K), arefe hangi gün (2.4K), kurban bayramı geri sayım",
     "Vodafone 1. sırada her varyantta"],
    ["A2.3 RAMAZAN 2026", "/ramazan-2026/", "ramazan ne zaman başlıyor 2026", "18.000 (Şubat-Mart spike)", "0", "informational", "★★★",
     "ramazan ne zaman başlıyor 2026, 2026 ramazan ne zaman başlıyor (8.7K), 2025 ramazan'a kaç gün kaldı (13K), ramazan ne zaman bitiyor, ramazan kaç gün sürer, ramazan ayı 2026, iftar vakti istanbul/ankara/izmir (il bazlı), imsak vakti il bazlı (823K!), namaz vakitleri il bazlı (13.6M!), oruç tutma niyeti, ramazan duaları",
     "Diyanet baskın ama il bazlı sayfalar fırsat"],
    ["A2.4 DINI GUNLER KANDILLER", "/dini-gunler-kandiller/", "kandil günleri 2026", "~50K küme", "düşük", "informational", "★★",
     "mevlit kandili ne zaman, berat kandili ne zaman, kadir gecesi, regaip kandili, miraç kandili, kandil günleri 2026, kandil mesajları, kandil duaları",
     ""],
    ["A2.5 29 EKIM CUMHURIYET BAYRAMI", "/29-ekim-cumhuriyet-bayrami/", "29 ekim cumhuriyet bayramı", "450.000 (Ekim 5M!)", "—", "informational", "★★★",
     "29 ekim cumhuriyet bayramı, 29 ekim, 29 ekim 1923, 29 ekim mesajları (880, Ekim 12K), 29 ekim sözleri, 29 ekim şiirleri, 29 ekim hangi gün (3.6K), 29 ekim tatil mi (10K), 29 ekim resmi tatil mi (9.5K), 29 ekim cumhuriyet bayramı tatil mi (1.8K), cumhuriyet bayramı tarihçesi, atatürk cumhuriyet sözleri",
     "Ekim ayında 5M arama"],
    ["A2.6 DIGER MILLI BAYRAMLAR", "/milli-bayramlar/", "23 nisan, 19 mayıs, 30 ağustos, 18 mart, 10 kasım", "1M+ spike ayı", "—", "informational", "★★",
     "23 nisan ulusal egemenlik ve çocuk bayramı, 23 nisan mesajları, 23 nisan şiirleri; 19 mayıs atatürk'ü anma, 19 mayıs mesajları; 30 ağustos zafer bayramı, 30 ağustos mesajları; 18 mart resmi tatil mi (6.8K), 18 mart çanakkale şehitlerini anma; on kasım resmi tatil mi (1.2K), 10 kasım anma töreni; 1 mayıs emek dayanışma",
     "Her bayramda spike ay"],
    ["A2.7 YILBASI VE BLACK FRIDAY", "/yilbasi-2026/, /black-friday/", "yılbaşı, black friday", "31K + 13K + spike", "0-11", "informational", "★★",
     "yılbaşı ne zaman, yılbaşı 2026, 31 aralık tatil mi (13K, KD 0), yılbaşı hangi güne (1.7K), yılbaşı filmleri (4.4K), yılbaşı mesajları, yılbaşı menüsü, yılbaşı hediyesi, black friday ne zaman (31K), black friday 2026, kara cuma ne zaman (13K), cyber monday",
     "Aralık başı + Kasım sonu"],
], columns=["Fikir", "URL_Onerisi", "Ana_Keyword", "Aylik_Hacim", "KD", "Intent", "Oncelik", "Oynanabilecek_KW_Listesi", "Notlar"])
write_df(ws, a2, col_widths={"A": 35, "B": 28, "C": 30, "D": 22, "E": 10, "F": 14, "G": 10, "H": 80, "I": 28})
for r_idx in range(2, len(a2) + 2):
    p = ws.cell(row=r_idx, column=7).value
    if p == "★★★":
        for c in range(1, 10):
            ws.cell(row=r_idx, column=c).fill = prio_fill
    elif p == "★★":
        for c in range(1, 10):
            ws.cell(row=r_idx, column=c).fill = mid_fill

# ========= A3: OZEL GUN MESAJ FIKIRLERI =========
ws = wb.create_sheet("03_A3_OzelGun_Mesaj_Fikirleri")
a3 = pd.DataFrame([
    ["A3.1 ANNELER GÜNÜ 2026", "/anneler-gunu-2026/", "anneler günü ne zaman", "673.000 (Mayıs 5M!)", "8", "informational", "★★★",
     "anneler günü ne zaman, anneler günü 2026, anneler günü hangi tarih, anneler günü hangi ay, anneler günü mesajları (74K, Mayıs 1M), anneler günü mesajı (92K), anneler günü sözleri (7.5K), anneler günü sözleri kısa (3K), anneler günü mesajı kısa (2.2K), anneler günü yazısı (2.4K), kayınvalideye anneler günü mesajı (2.4K), öğretmene anneler günü mesajı (1.9K), anneler günü hediyesi, anne hediyesi, anneler günü şiiri, anneler günü kartı",
     "Mayıs ayında 5M! Hub + mesaj galerisi + hediye fikirleri"],
    ["A3.2 BABALAR GÜNÜ 2026", "/babalar-gunu-2026/", "babalar günü ne zaman", "450.000 (Haziran 4M!)", "—", "informational", "★★★",
     "babalar günü ne zaman, babalar günü 2026, babalar günü mesajları (4.3K), babalar günü mesajı (3.7K), babalar günü sözleri (15K), babalar günü sözleri kısa (2.1K), babalar günü için yazı (1.9K), babalar günü şiiri, duygusal babalar günü mesajları (2.4K), babalar gunu mesaji (2.3K), eşe babalar günü mesajı (2K), babalar günü'ne özel notlar (2K), babalar günü hediyesi, baba hediyesi",
     "Haziran ayında 4M"],
    ["A3.3 SEVGİLİLER GÜNÜ 14 ŞUBAT", "/sevgililer-gunu/", "sevgililer günü mesajları", "33.100 (Şubat 368K)", "—", "informational", "★★★",
     "14 şubat ne zaman, sevgililer günü hangi gün, sevgililer günü mesajları (33.1K, Şubat 368K), sevgililer günü sözleri, sevgililer günü şiirleri, sevgiliye şiir (6.6K), aşk şiirleri, sevgililer günü hediyesi, sevgilime mesaj, romantik mesajlar, aşk mesajları",
     "Şubat 368K spike"],
    ["A3.4 ÖĞRETMENLER GÜNÜ 24 KASIM", "/ogretmenler-gunu/", "öğretmenler günü için güzel sözler", "5.600 (Kasım spike)", "1", "informational", "★★",
     "öğretmenler günü ne zaman, 24 kasım, öğretmenler günü için güzel sözler (5.6K, KD 1), unutulmaz öğretmen sözleri (6.6K, KD 0), öğretmenler günü notu (2.4K), öğretmenler günü için yazı (1.9K), öğretmen sözleri kısa (2K), öğretmenler günü mesajı, öğretmenler günü şiiri, öğretmenlere hediye",
     "Kasım ayı spike, KD çok düşük"],
    ["A3.5 8 MART KADINLAR GÜNÜ", "/8-mart-kadinlar-gunu/", "8 mart mesajları", "~100K+ tahmini", "—", "informational", "★★",
     "8 mart ne zaman, kadınlar günü hangi gün, 8 mart kadınlar günü, dünya kadınlar günü, 8 mart mesajları, kadınlar günü sözleri, kadınlar günü şiirleri, kadınlar günü hediyesi",
     "Mart ayı spike"],
    ["A3.6 23 NİSAN ÇOCUK BAYRAMI", "/23-nisan-cocuk-bayrami/", "23 nisan", "~200K+ tahmini", "—", "informational", "★★",
     "23 nisan ne zaman, 23 nisan mesajları, 23 nisan şiirleri, 23 nisan ulusal egemenlik, çocuk bayramı mesajları, çocuk şiirleri, 23 nisan hediyesi",
     "Nisan spike"],
    ["A3.7 DOĞUM GÜNÜ MEGA-HUB", "/dogum-gunu-mesajlari/", "doğum günü mesajları", "201.000 (yıl boyu)", "—", "informational", "★★★",
     "doğum günü mesajları (201K), doğum günü mesajı, doğum günü hediyesi (26K), yaş günü mesajları (880), eşe/anneye/babaya/arkadaşa/sevgiliye/kardeşe doğum günü mesajı, çocuğuma doğum günü, komik/duygusal doğum günü mesajları, iyiki doğdun mesajları, doğum günü şarkıları, doğum günü hesaplama (8.1K)",
     "Yıl boyu sabit hacim, BiP entegrasyonu doğal"],
    ["A3.8 GUNLUK MESAJ GALERILERI", "/gunluk-mesajlar/", "günaydın mesajları", "135.000 + 40.5K + 165K + 110K", "4 / 4 / 11", "informational", "★★★",
     "günaydın mesajları (135K, KD 4!), günaydın sözleri, günaydın resimleri, iyi geceler mesajları (40.5K), iyi geceler sözleri, tatlı uykular, hayırlı sabahlar mesajları, güzel sözler (165K, KD 11), anlamlı sözler (110K, KD 4!), duygusal sözler, motivasyon sözleri, taziye mesajları, başsağlığı mesajları, geçmiş olsun mesajları, şifa mesajları",
     "EN DÜŞÜK KD! 135K+40K+165K+110K = 450K+/ay"],
    ["A3.9 BAYRAM MESAJLARI HUB", "/bayram-mesajlari/", "bayram mesajları + ramazan/kurban", "14.8K + 74K (spike)", "—", "informational", "★★★",
     "bayram mesajları (14.8K, Mart 110K), ramazan bayramı mesajları (74K, Mart 823K!), kurban bayramı mesajları, bayram için güzel sözler, bayram tebrik mesajı, bayramda söylenecek dualar, yeni yıl mesajları, yılbaşı mesajları, kandil mesajları, 29 ekim mesajları (880, Ekim 12K)",
     "Spike aylar Mart/Ekim/Aralık"],
], columns=["Fikir", "URL_Onerisi", "Ana_Keyword", "Aylik_Hacim", "KD", "Intent", "Oncelik", "Oynanabilecek_KW_Listesi", "Notlar"])
write_df(ws, a3, col_widths={"A": 35, "B": 28, "C": 30, "D": 22, "E": 10, "F": 14, "G": 10, "H": 80, "I": 32})
for r_idx in range(2, len(a3) + 2):
    p = ws.cell(row=r_idx, column=7).value
    if p == "★★★":
        for c in range(1, 10):
            ws.cell(row=r_idx, column=c).fill = prio_fill
    elif p == "★★":
        for c in range(1, 10):
            ws.cell(row=r_idx, column=c).fill = mid_fill

# ========= A4: TELEKOM + SORGU FIKIRLERI =========
ws = wb.create_sheet("04_A4_Telekom_Sorgu_Fikirleri")
a4 = pd.DataFrame([
    ["A4.1 MOBİL YARDIM HUB", "/mobil-yardim/", "30+ how-to sayfa", "200K+ küme", "0-30", "informational", "★★★",
     "güvenli arama nasıl kapatılır (20K, KD 0), nfc özelliği nasıl açılır (12K), nfc nedir (11K), nfc özelliği nedir (11K), konum nasıl atılır (10K), konum paylaşma, canlı konum, kendi numaramı nasıl öğrenebilirim (3.6K), esim nedir (6.6K), esim nasıl kurulur, esim destekleyen telefonlar, 5g nedir (9.9K), 5g destekleyen telefonlar, mobil imza nasıl alınır, e-imza (10K), elektronik imza (2.6K), numara taşıma, imei sorgulama (5.4K), cayma bedeli (3K), telefona format (5.4K), ekran görüntüsü (3.3K), iphone nasıl kapatılır (3.7K), güncelleme nasıl yapılır (2.9K)",
     "Vodafone/turk.net dominant; Turkcell doğal alan"],
    ["A4.2 INTERNET MODEM HUB", "/superonline-yardim/", "modem & wifi & hız", "1M+ küme", "0-34", "informational", "★★★",
     "192.168.1.1 (106K), modem arayüzü, modem arayüzüne giriş (varyant), wifi şifresi değiştirme (18.1K), modem şifresi sıfırlama, internet hız testi (1M, KD 34), hız testi, download hızı, upload hızı, ping testi, wifi sinyal güçlendirme, dns nedir, dns ayarlama, google dns, vpn nedir, vpn nasıl kullanılır (3.5K), vpn nasıl açılır (3.4K), bedava vpn, Superonline modem ayarları",
     "Superonline ürün otoritesi"],
    ["A4.3 SORGULAMA REHBERİ", "/sorgulama-rehberi/", "iban + hgs + plaka + pasaport + posta", "300K+ küme", "3-40", "informational/navigational", "★★★",
     "iban sorgulama (33.1K), iban hesaplama (170), iban öğrenme, iban nedir, hgs bakiye sorgulama (110K), hgs sorgulama, hgs nasıl yüklenir, mtv sorgulama (22.2K), mtv 2026, mtv hesaplama, plaka sorgu (74K, KD 4), plakadan araç sorgulama, trafik cezası sorgulama, posta kodu sorgulama (14.8K), posta kodu (102K), posta kodu öğrenme (12K), pasaport ücreti 2026 (18.1K, KD 3), pasaport ücretleri 2025 (23K), pasaport fiyatları 2025 (13K), yeşil pasaport ücreti (27.1K, KD 3), bordo pasaport (12.1K), 10 yıllık pasaport ücreti 2026 (9K), alan kodları (260, KD 3, il bazlı sayfalar)",
     "Turkcell brand güveni = bu cluster'ın doğal sahibi"],
    ["A4.4 SGK REHBERİ HUB", "/sgk-rehberi/", "sgk işe giriş + hizmet dökümü", "200K+ küme", "20-63", "navigational", "★★★",
     "sgk işe giriş (135K, KD 20), sgk işe giriş bildirgesi (11K), sgk hizmet dökümü (49.5K), sgk başlangıç tarihi, sgk 4a 4b 4c farkı, 4a sigortalı kim, emekli sandığı sorgulama, bağkur sorgulama, emeklilik yaşı hesaplama (3.6K), sigorta primi hesaplama",
     "HR/işveren kitle; e-Devlet bridge"],
    ["A4.5 KIRA & KONUT HUB", "/kira-konut/", "kira artış oranı + kontrat örneği", "80K+ küme", "düşük", "informational", "★★",
     "kira artış oranı hesaplama (6.6K), kira artışı yüzde kaç, tüfe oranı 2026, kira kontratı örneği (8.1K), kira sözleşmesi, dilekçe örneği (74K), dilekçe nasıl yazılır, tahliye dilekçesi, mazeret dilekçesi, taahhütname örneği, ev sahibine ihtar",
     "KOBİ blog için niş"],
    ["A4.6 CEZAEVI ILETIŞIM REHBERI", "/cezaevi-iletisim/", "cezaevi kontör", "6.600 (+%9329 trend!)", "düşük", "informational", "★★★",
     "cezaevi kontör (6.6K), cezaevi kontör yükleme, cezaevi telefon, cezaevi dakika alma (700), mahkum görüşmesi, mahkumla telefon görüşmesi, telefondan mahkum görüşme, cezaevine para yatırma, e-mektup gönderme cezaevi, cezaevi adres listesi",
     "Turkcell EN DOĞAL otorite; mega trend (+%9329 yıllık)"],
], columns=["Fikir", "URL_Onerisi", "Ana_Keyword", "Aylik_Hacim", "KD", "Intent", "Oncelik", "Oynanabilecek_KW_Listesi", "Notlar"])
write_df(ws, a4, col_widths={"A": 35, "B": 28, "C": 32, "D": 22, "E": 10, "F": 20, "G": 10, "H": 90, "I": 36})
for r_idx in range(2, len(a4) + 2):
    p = ws.cell(row=r_idx, column=7).value
    if p == "★★★":
        for c in range(1, 10):
            ws.cell(row=r_idx, column=c).fill = prio_fill
    elif p == "★★":
        for c in range(1, 10):
            ws.cell(row=r_idx, column=c).fill = mid_fill

# ========= A5: AI SIBER DIGER FIKIRLERI =========
ws = wb.create_sheet("05_A5_AI_Siber_Diger")
a5 = pd.DataFrame([
    ["A5.1 SİBER GÜVENLİK HUB", "/siber-guvenlik-rehberi/", "siber güvenlik + phishing + deepfake + 2FA", "400K+ küme", "10-49", "informational", "★★★",
     "siber güvenlik (9.9K, KD 10), phishing (368K!), oltalama saldırısı, e-mail dolandırıcılığı, siber zorbalık, siber suç, siber dolandırıcılık, deepfake nedir (8.1K), 2fa nedir, iki adımlı doğrulama, iki faktörlü kimlik doğrulama (1.9K), mcafee nedir (11K), mcafee kaldırma, airtag nedir (10K), takip cihazı, wi-fi açık ağlarda güvenlik, sosyal mühendislik, veri sızıntısı sorgulama, antivirüs program önerileri, çocuklar için internet güvenliği",
     "Turkcell+McAfee+BiP güvenlik servisleriyle direkt köprü"],
    ["A5.2 DEPREM AFET GÜVENLİĞİ", "/deprem-ve-afet-guvenligi/", "deprem güvenlik bilgileri", "158.000", "3", "informational", "★★★",
     "deprem güvenlik bilgileri (158K), deprem çantası içeriği, deprem çantasında ne olmalı, deprem sırasında ne yapmalı, depremde nasıl davranılır, deprem sonrası ne yapmalı, artçı deprem, android deprem uyarı sistemi nasıl aktif edilir (Ahrefs 4.3K), en iyi deprem uygulamaları (Ahrefs 4.2K), afad uyarı, kandilli rasathanesi, son depremler, deprem kuşağı türkiye, deprem fay hatları, yangın söndürme tüpü, 112 ne zaman aranır",
     "turk.net 1. (2.571 trafik); Turkcell brand güvenilirliği = bu cluster'ın doğal sahibi"],
    ["A5.3 YAPAY ZEKA REHBERİ", "/yapay-zeka-rehberi/", "chatgpt/gemini/claude/deepseek", "1M+ küme", "16-100", "informational", "★★",
     "yapay zeka (823K, KD 50 zor), chatgpt türkçe (135K, KD 16), chatgpt nedir, chatgpt nasıl kullanılır, chatgpt türkçe ücretsiz soru sor, chatgpt 4 ücretsiz, gemini ai (450K, KD 100 zor), gemini nasıl kullanılır (3.3K), claude ai (74K, KD 65), claude türkçe, claude vs chatgpt, deepseek (368K, KD 69), deepseek nedir, deepseek vs chatgpt, grok nedir (19K, KD 29), meta ai (17K), kling ai, capcut nedir (127K), canva nedir (14K), midjourney nedir (54K), google lens (131K, KD 31), google form oluşturma (12.1K, KD 14), prompt nedir, prompt yazma",
     "Yüksek KD ama trend mega; long-tail kapısı açık"],
    ["A5.4 SOSYAL MEDYA REHBERİ", "/sosyal-medya-rehberi/", "whatsapp + instagram + tiktok rehberleri", "1M+ küme", "0-50", "informational", "★★",
     "whatsapp silinen mesajları geri getirme (16K), whatsapp silinen mesajları görme (14K), whatsapp web giriş (201K), whatsapp işletme hesabı (900), whatsapp business; instagram hesabı silme/dondurma (20K+), instagram silinen mesajlar (8.2K); x twitter giriş, x sorun (88K); telegram web (170K); threads nedir (22K); tiktok rehberi; snapchat hesap silme; linkedin profil; facebook eski hesap; discord rehberi",
     "turk.net dominant; BiP köprü içerik"],
    ["A5.5 DİZİ FİLM REHBERİ (Turkcell TV+)", "/dizi-rehberi/", "dizi oyuncuları + sezon bilgileri", "500K+ küme", "0-23", "informational", "★",
     "kızılcık şerbeti oyuncuları (105K), eşref rüya oyuncuları (79K), taşacak bu deniz oyuncuları (80K), wednesday 2. sezon (51K, 9K varyant), valorant mobile ne zaman (5.4K), gta 6 ne zaman çıkacak (16K), iphone 16 ne zaman (20K), playstation 6 ne zaman (5.3K), samsung s26 ultra (8.8K), supernatural (51K), netflixin en çok izlenen (Ahrefs)",
     "Her popüler dizi için sayfa; Turkcell TV+ köprü"],
    ["A5.6 OYUN REHBERİ", "/oyun-rehberi/", "X ne zaman çıkacak + X kaç GB", "150K+ küme", "0-50", "informational", "★",
     "gta 6 (73K), ps5 (328K), ps5 fiyat (88K), playstation (136K), playstation 5 (272K), pubg kaç gb (2.4K), lol kaç gb (5.7K), fortnite kaç gb (4K), nintendo switch (88K), valorant mobile, oyun konsolu rehberi, oyun direksiyon seti (76K), sistem gereksinimleri",
     "Tarife önerilerine bridge (büyük dosya = büyük paket)"],
    ["A5.7 İŞ BAŞVURU & CV ARACI", "/cv-hazirla/", "cv hazırlama (online araç)", "186.000+", "33-64", "transactional", "★★★",
     "cv hazırlama (186K), cv hazırlama ücretsiz (30K), ücretsiz cv hazırlama (20K), cv hazırlama ücretsiz pdf (7.2K), cv hazırlama pdf (3.1K), pdf cv hazırlama (400), özgeçmiş hazırlama (3K), kariyer cv hazırlama (700), online cv hazırlama (600), telefondan cv hazırlama (1.2K), bedava cv hazırlama (500), cv hazırlama programı (1.9K), motivasyon mektubu örnek, niyet mektubu örnek, mülakat soruları, iş görüşmesi soruları, linkedin profil optimizasyonu",
     "Vodafone tam bu modeli kullanıyor (/hakkimizda/insan-kaynaklari/cv-hazirlama-araci 6.295 trafik)"],
    ["A5.8 ÇİÇEK / HEDIYE / ETKİNLİK", "/cicek-hediye-etkinlik/", "taze çiçek + sinema bilet + biletix", "300K+ küme", "0-67", "transactional/navigational", "★",
     "taze çiçek (155K, TT 8.), çiçeksepet (1.7K), çiçek sepet indir (500), çiçek sepeti kampanyalı ürünler (450), biletix (227K, TT 9.), biletix etkinlik (170), biletix konser (5.5K), sinema bilet fiyatları (18.1K), sinema bileti al (4.4K), vialand bilet (11K), bilet.com (3.2K), uçak bileti kampanya (1.4K), müzekart öğrenci (1.4K), müzekart nereden alınır (1.2K)",
     "Turkcell Prime ile köprü"],
    ["A5.9 BURÇ & FAL & İSİM ANLAMI (eğlence)", "/burc-fal-isim/", "yükselen burç + fal + isim anlamı", "300K+ küme", "0-36", "informational", "★",
     "yükselen burç hesaplama (201K, KD 10), burç uyumu, uyumlu burçlar (260), kahve falı (18.1K, KD 12), tarot falı (KD 36), el falı, yıldız falı (390), günlük burç yorumu, isim anlamı (6.6K, KD 3), isim sözlüğü, bebek ismi önerileri, kız ismi, erkek ismi, soyadı anlamları (90)",
     "Eğlence niche, lifestyle blog için"],
], columns=["Fikir", "URL_Onerisi", "Ana_Keyword", "Aylik_Hacim", "KD", "Intent", "Oncelik", "Oynanabilecek_KW_Listesi", "Notlar"])
write_df(ws, a5, col_widths={"A": 38, "B": 30, "C": 32, "D": 22, "E": 12, "F": 22, "G": 10, "H": 90, "I": 38})
for r_idx in range(2, len(a5) + 2):
    p = ws.cell(row=r_idx, column=7).value
    if p == "★★★":
        for c in range(1, 10):
            ws.cell(row=r_idx, column=c).fill = prio_fill
    elif p == "★★":
        for c in range(1, 10):
            ws.cell(row=r_idx, column=c).fill = mid_fill

# ========= B1: CLUSTER OZETI =========
ws = wb.create_sheet("06_B1_Cluster_Ozeti")
cluster_summary = df.groupby("niche").agg(
    Keyword_Sayisi=("Keyword", "count"),
    Toplam_Aylik_Hacim=("Volume", "sum"),
    Ort_Hacim=("Volume", "mean"),
    Max_Hacim=("Volume", "max"),
).reset_index().sort_values("Toplam_Aylik_Hacim", ascending=False)
cluster_summary["Ort_Hacim"] = cluster_summary["Ort_Hacim"].astype(int)
cluster_summary["Toplam_Aylik_Hacim"] = cluster_summary["Toplam_Aylik_Hacim"].astype(int)
cluster_summary["Max_Hacim"] = cluster_summary["Max_Hacim"].astype(int)
# Brand uyumu manual mapping
brand_uyum = {
    "01_HESAPLAMA_DUNYA": "★★★", "02_BIRIM_DONUSUM": "★★",
    "03_TATIL_BAYRAM_OZEL_GUN": "★★★", "04_MESAJ_SOZ_KART": "★★★",
    "05_TANIM_NEDIR": "★★", "06_NASIL_YAPILIR": "★★★",
    "07_NE_ZAMAN": "★★★", "08_KAC": "★★",
    "09_SOSYAL_MEDYA_APP": "★★", "10_TELEFON_CIHAZ": "★",
    "11_TELEFON_AYAR_SORUN": "★★★", "12_INTERNET_MODEM_WIFI": "★★★",
    "13_EDEVLET_SORGULAMA": "★★★", "14_OTOMOTIV": "★",
    "15_EGITIM_SINAV": "★★", "16_BURC_ASTROLOJI": "★",
    "17_SPOR": "★", "18_OYUN_REHBER": "★",
    "19_DIZI_FILM": "★", "20_KLAVYE_SEMBOL": "★★",
    "21_KELIME_YAZIM_DIL": "★★", "22_HAYVANLAR": "★",
    "23_BITKILER_BAHCE": "★", "24_RUYA_TABIRI": "★",
    "25_ASTRONOMI": "—", "26_KONSER_ETKINLIK": "★★",
    "27_SEYAHAT_ULASIM": "★★", "28_IS_BASVURU_CV": "★★★",
    "29_FINANS_BANKA_YATIRIM": "★★", "30_HUKUK_KANUN": "★★★",
    "31_EV_DEKOR_MOBILYA": "★", "32_BEAUTY_KOZMETIK": "★",
    "33_COCUK_BEBEK": "★", "34_SAGLIK_VUCUT": "★★",
    "35_YEMEK_TARIF": "★", "ZZ_DIGER": "filtrelenmeli",
}
cluster_summary["Brand_Uyumu"] = cluster_summary["niche"].map(brand_uyum)
write_df(ws, cluster_summary, col_widths={"A": 32, "B": 14, "C": 22, "D": 14, "E": 14, "F": 16})

# ========= B2: RAKIP URL PATTERN =========
ws = wb.create_sheet("07_B2_Rakip_URL_Patternleri")
def extract_path_prefix(url, depth=2):
    if pd.isna(url):
        return None
    m = re.match(r"https?://[^/]+(/[^?#]*)", str(url))
    if not m:
        return None
    path = m.group(1)
    parts = [p for p in path.split("/") if p]
    return "/" + "/".join(parts[:depth]) if parts else "/"

all_patterns = []
for comp in competitors:
    url_col = f"{comp}/: URL"
    traf_col = f"{comp}/: Organic Traffic"
    sub = df_raw[df_raw[url_col].notna() & df_raw[traf_col].notna()].copy()
    sub = sub[sub["Volume"] >= 100]
    sub["path_prefix"] = sub[url_col].apply(lambda x: extract_path_prefix(x, 2))
    stats = sub.groupby("path_prefix").agg(
        kw=("Keyword", "count"),
        traffic=(traf_col, "sum"),
        volume=("Volume", "sum"),
    ).sort_values("traffic", ascending=False).head(15).reset_index()
    stats["Rakip"] = comp
    all_patterns.append(stats[["Rakip", "path_prefix", "kw", "traffic", "volume"]])
patterns_df = pd.concat(all_patterns, ignore_index=True)
patterns_df.columns = ["Rakip", "URL_Path", "Keyword_Sayisi", "Aylik_Trafik", "Toplam_Hacim"]
patterns_df["Aylik_Trafik"] = patterns_df["Aylik_Trafik"].astype(int)
patterns_df["Toplam_Hacim"] = patterns_df["Toplam_Hacim"].astype(int)
write_df(ws, patterns_df, col_widths={"A": 26, "B": 60, "C": 14, "D": 14, "E": 14})

# ========= B3: DfS DOGRULAMA =========
ws = wb.create_sheet("08_DfS_Dogrulama_TR")
dfs_validation = pd.DataFrame([
    ["hesap makinesi", 2740000, "—", "informational", "Vatan baskın, hub fırsatı"],
    ["kredi hesaplama", 1220000, 45, "informational", "Yüksek KD"],
    ["kurban bayramı ne zaman", 1220000, 11, "informational", "Mart-Mayıs spike 6M"],
    ["internet hız testi", 1000000, 34, "informational", "Mevcutsa kontrol"],
    ["namaz vakitleri", 13600000, 15, "informational", "Diyanet baskın"],
    ["yapay zeka", 823000, 50, "informational", "Yüksek KD"],
    ["imsak vakti", 823000, 32, "informational", "Ramazan'da 2.2M+"],
    ["anneler günü ne zaman", 673000, 8, "informational", "Mayıs 5M spike"],
    ["yüzde hesaplama (MEVCUT)", 673000, 12, "informational", "Mevcut sayfa"],
    ["babalar günü ne zaman", 450000, "—", "informational", "Haziran 4M"],
    ["29 ekim cumhuriyet bayramı", 450000, "—", "informational", "Ekim 5M"],
    ["gemini ai", 450000, 100, "informational", "Çok zor"],
    ["asgari ücret 2026 (+%210K trend)", 368000, 9, "informational", "★★★"],
    ["deepseek", 368000, 69, "informational", "Zor"],
    ["tdk sözlük", 368000, 24, "navigational", "TDK kendi var"],
    ["phishing", 368000, "—", "informational", "Mart 4M (siber)"],
    ["yaş hesaplama (MEVCUT)", 301000, 9, "informational", "Mevcut"],
    ["KDV hesaplama", 201000, 19, "informational", "★★★"],
    ["kıdem tazminatı hesaplama", 201000, 27, "informational", "★★★"],
    ["yükselen burç hesaplama", 201000, 10, "informational", "★★★"],
    ["doğum günü mesajları", 201000, "—", "informational", "★★★ yıl boyu"],
    ["güzel sözler", 165000, 11, "informational", "★★★"],
    ["iftar vakti", 165000, "—", "informational", "Ramazan'da 1M"],
    ["günaydın mesajları", 135000, 4, "informational", "★★★ EN DÜŞÜK KD"],
    ["sgk işe giriş", 135000, 20, "navigational", "★★★"],
    ["chatgpt türkçe", 135000, 16, "navigational", "★★"],
    ["anlamlı sözler", 110000, 4, "informational", "★★★ EN DÜŞÜK KD"],
    ["hgs bakiye sorgulama", 110000, 19, "navigational", "★★★"],
    ["bugün ayın kaçı", 90500, "—", "informational", "★★"],
    ["anneler günü mesajları (Mayıs 1M)", 74000, "—", "informational", "★★★"],
    ["ramazan bayramı mesajları (Mart 823K)", 74000, "—", "informational", "★★★"],
    ["dilekçe örneği", 74000, "—", "informational", "★★ KOBİ"],
    ["plaka sorgu", 74000, 4, "informational", "★★★ DÜŞÜK KD"],
    ["kalori hesaplama", 74000, "medium", "informational", "★★"],
    ["döviz çevirici", 74000, 32, "informational", "★★"],
    ["claude ai", 74000, 65, "informational", "Zor"],
    ["sgk hizmet dökümü", 49500, 28, "navigational", "★★"],
    ["iyi geceler mesajları", 40500, "—", "informational", "★★"],
    ["noktalama işaretleri", 40500, 4, "informational", "★★★ DÜŞÜK KD"],
    ["iban sorgulama", 33100, 23, "navigational", "★★★ Paycell"],
    ["sevgililer günü mesajları (Şubat 368K)", 33100, "—", "informational", "★★★"],
    ["cv hazırlama ücretsiz", 27100, 33, "transactional", "★★★ Vodafone modeli"],
    ["ihbar tazminatı hesaplama", 27100, 25, "commercial", "★★★"],
    ["yeşil pasaport ücreti", 27100, 3, "informational", "★★★ DÜŞÜK KD"],
    ["resmi tatiller 2026", 22200, 3, "informational", "★★★ +%8303 trend"],
    ["mtv sorgulama", 22200, 42, "informational", "★★"],
    ["pasaport ücreti 2026", 18100, 3, "informational", "★★★ DÜŞÜK KD"],
    ["WiFi şifresi değiştirme", 18100, "—", "informational", "★★★"],
    ["kelime sayacı", 18100, 12, "informational", "★★★"],
    ["regl hesaplama", 18100, "—", "informational", "★★★"],
    ["kahve falı", 18100, 12, "informational", "★★"],
    ["sinema bilet fiyatları", 18100, 14, "transactional", "★★"],
    ["karakter sayacı", 14800, 3, "informational", "★★★ EN DÜŞÜK KD"],
    ["iyelik eki", 14800, 5, "informational", "★★★ DÜŞÜK KD"],
    ["posta kodu sorgulama", 14800, 40, "informational", "★★"],
    ["bayram mesajları (Mart 110K)", 14800, "—", "informational", "★★★"],
    ["bordo pasaport", 12100, "—", "informational", "★★★"],
    ["google form oluşturma", 12100, 14, "informational", "★★"],
    ["fazla mesai hesaplama", 9900, "—", "informational", "★★★"],
    ["yıllık izin hesaplama", 9900, 4, "informational", "★★★ DÜŞÜK KD"],
    ["siber güvenlik", 9900, 10, "navigational", "★★★ DÜŞÜK KD"],
    ["5G nedir", 9900, 29, "informational", "★★★"],
    ["BMI hesaplama", 8100, "—", "informational", "★★"],
    ["brüt net hesaplama", 8100, 30, "informational", "★★ Vodafone modeli"],
    ["doğum günü hesaplama", 8100, 15, "informational", "★★"],
    ["deepfake", 8100, 49, "informational", "★★"],
    ["kira kontratı örneği", 8100, "—", "informational", "★★"],
    ["cezaevi kontör (+%9329 trend!)", 6600, "—", "informational", "★★★"],
    ["kira artış oranı hesaplama", 6600, "—", "informational", "★★"],
    ["ovulasyon hesaplama", 6600, 7, "informational", "★★★ DÜŞÜK KD"],
    ["isim anlamı", 6600, 3, "informational", "★★★ DÜŞÜK KD"],
    ["resmi tatil günleri", 6600, 8, "informational", "★★★"],
    ["sevgiliye şiir", 6600, "—", "informational", "★★"],
    ["eSIM nedir", 6600, 19, "informational", "★★★"],
    ["telefon IMEI sorgulama", 5400, 26, "informational", "★★★"],
    ["elektrik fatura hesaplama", 5400, "—", "informational", "★★"],
    ["emeklilik yaşı hesaplama", 3600, 18, "informational", "★★"],
    ["saat kaç türkiye", 2900, 37, "informational", "★★"],
    ["yakıt tüketim hesaplama", 2400, 8, "informational", "★★ Otomotiv"],
    ["hamilelik haftası hesaplama", 2400, 8, "informational", "★★★ DÜŞÜK KD"],
    ["askerlik hesaplama", 1900, "—", "informational", "★★"],
    ["iki faktörlü kimlik doğrulama", 1900, 36, "navigational", "★★"],
    ["yazım kılavuzu", 1900, 18, "informational", "★★"],
    ["taksit hesaplama", 1900, 37, "informational", "★"],
    ["TC kimlik no sorgulama", 1000, 10, "navigational", "★"],
    ["saat farkı", 880, "—", "informational", "★"],
    ["yaş günü mesajları", 880, "—", "informational", "★"],
    ["araç muayene süresi", 720, 3, "informational", "★★"],
    ["isim bulma", 480, "—", "informational", "★"],
    ["abd saat farkı", 390, "—", "informational", "★"],
    ["yıldız falı", 390, 6, "informational", "★"],
    ["alan kodları", 260, 3, "informational", "★ il bazlı sayfa fırsatı"],
    ["uyumlu burçlar", 260, "—", "informational", "★"],
    ["biletix etkinlik", 170, 24, "navigational", "★"],
    ["qr kod nasıl okunur", 170, 1, "informational", "★ niş"],
    ["iban hesaplama", 170, 30, "navigational", "★ sorgulama daha hacimli"],
    ["vergi iadesi hesaplama", 140, "—", "informational", "★"],
    ["soyadı anlamları", 90, 1, "informational", "★ niş"],
    ["doğal gaz fatura hesaplama", 70, "—", "transactional", "★ niş"],
], columns=["Keyword", "Aylik_Hacim", "KD", "Intent", "Notlar"])
write_df(ws, dfs_validation, col_widths={"A": 42, "B": 14, "C": 10, "D": 16, "E": 36})
for r_idx in range(2, len(dfs_validation) + 2):
    notlar = str(ws.cell(row=r_idx, column=5).value or "")
    if "★★★" in notlar:
        for c in range(1, 6):
            ws.cell(row=r_idx, column=c).fill = prio_fill
    elif "★★" in notlar:
        for c in range(1, 6):
            ws.cell(row=r_idx, column=c).fill = mid_fill

# ========= NICHE CLUSTER SHEET'LERI =========
cluster_order = [
    "01_HESAPLAMA_DUNYA", "03_TATIL_BAYRAM_OZEL_GUN", "11_TELEFON_AYAR_SORUN",
    "12_INTERNET_MODEM_WIFI", "13_EDEVLET_SORGULAMA", "04_MESAJ_SOZ_KART",
    "06_NASIL_YAPILIR", "07_NE_ZAMAN", "05_TANIM_NEDIR",
    "02_BIRIM_DONUSUM", "08_KAC", "20_KLAVYE_SEMBOL", "21_KELIME_YAZIM_DIL",
    "28_IS_BASVURU_CV", "30_HUKUK_KANUN", "34_SAGLIK_VUCUT", "26_KONSER_ETKINLIK",
    "29_FINANS_BANKA_YATIRIM", "16_BURC_ASTROLOJI", "24_RUYA_TABIRI", "15_EGITIM_SINAV",
    "14_OTOMOTIV", "17_SPOR", "18_OYUN_REHBER", "19_DIZI_FILM",
    "27_SEYAHAT_ULASIM", "33_COCUK_BEBEK", "22_HAYVANLAR", "23_BITKILER_BAHCE",
    "32_BEAUTY_KOZMETIK", "31_EV_DEKOR_MOBILYA", "35_YEMEK_TARIF",
    "09_SOSYAL_MEDYA_APP", "10_TELEFON_CIHAZ", "25_ASTRONOMI",
]

sheet_no = 9
for cname in cluster_order:
    sub = df_sorted[df_sorted["niche"] == cname].head(300)
    if len(sub) == 0:
        continue
    short_name = cname[3:][:24]
    sheet_name = f"{sheet_no:02d}_n_{short_name}"
    sheet_no += 1
    ws = wb.create_sheet(sheet_name)
    out = sub[["Keyword", "Volume", "KD", "CPC", "SERP features",
               "best_competitor", "best_position", "best_traffic", "best_url"]].copy()
    out.columns = ["Keyword", "Aylik_Hacim", "KD", "CPC", "SERP_Features",
                   "En_Iyi_Rakip", "Pozisyon", "Trafik", "Rakip_URL"]
    write_df(ws, out, col_widths={"A": 42, "B": 14, "C": 8, "D": 8, "E": 28,
                                   "F": 22, "G": 12, "H": 14, "I": 70})

# Diger sheet'i (top 500)
diger = df_sorted[df_sorted["niche"] == "ZZ_DIGER"].head(500)
ws = wb.create_sheet(f"{sheet_no:02d}_n_DIGER_top500")
sheet_no += 1
out = diger[["Keyword", "Volume", "KD", "best_competitor", "best_position", "best_traffic"]].copy()
out.columns = ["Keyword", "Aylik_Hacim", "KD", "En_Iyi_Rakip", "Pozisyon", "Trafik"]
write_df(ws, out, col_widths={"A": 42, "B": 14, "C": 8, "D": 22, "E": 12, "F": 14})

# Tum 60K firsat tek tablo
ws = wb.create_sheet(f"{sheet_no:02d}_TUM_60K_FIRSATLAR")
all_opp = df_sorted[["Keyword", "Volume", "KD", "CPC", "SERP features", "niche",
                       "best_competitor", "best_position", "best_traffic", "best_url"]].copy()
all_opp.columns = ["Keyword", "Aylik_Hacim", "KD", "CPC", "SERP_Features", "Niche_Cluster",
                    "En_Iyi_Rakip", "Pozisyon", "Trafik", "Rakip_URL"]
# Excel sınırı: 1M satır. 60K rahat sığar.
write_df(ws, all_opp, col_widths={"A": 42, "B": 14, "C": 8, "D": 8, "E": 28,
                                    "F": 28, "G": 22, "H": 12, "I": 14, "J": 70})

wb.save(EXCEL_PATH)
print(f"\nExcel v2 kaydedildi: {EXCEL_PATH}")
print(f"Toplam sheet sayisi: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"  - {s}")
