"""
Tum analizleri tek Excel dosyasinda birlestir.
20+ sheet: ozet, kendi arastirma, cluster ozet, top kw cluster bazinda, dfs dogrulama
"""
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

CSV_PATH = "/Users/Erdo/Desktop/Claude Projects/Turkcel/www.turkcell.com.tr-content-gap-subdomains-t_2026-05-13_14-14-31.csv"
OUTPUT_DIR = "/Users/Erdo/Desktop/Claude Projects/Turkcel/output"
EXCEL_PATH = f"{OUTPUT_DIR}/TURKCELL_TRAFIK_FIRSATLARI.xlsx"

# Pool'u tekrar yukle
df = pd.read_csv(CSV_PATH, low_memory=False)
competitors = ["turktelekom.com.tr", "vodafone.com.tr", "mediamarkt.com.tr",
               "vatanbilgisayar.com", "turk.net", "pttcell.com.tr"]

df = df[df["www.turkcell.com.tr/: URL"].isna()].copy()

def best_comp_info(row):
    best_pos = 999
    best_comp = None
    best_url = None
    best_traf = 0
    second_comp = None
    second_pos = None
    for comp in competitors:
        pos = row.get(f"{comp}/: Organic Position", None)
        if pd.notna(pos):
            if pos < best_pos:
                if best_comp:
                    second_comp = best_comp
                    second_pos = best_pos
                best_pos = pos
                best_comp = comp
                best_url = row.get(f"{comp}/: URL", None)
                best_traf = row.get(f"{comp}/: Organic Traffic", 0)
            elif second_pos is None or pos < second_pos:
                second_comp = comp
                second_pos = pos
    return pd.Series([best_comp, best_pos if best_pos < 999 else None, best_url,
                      best_traf if pd.notna(best_traf) else 0, second_comp, second_pos])

df[["best_competitor", "best_position", "best_url", "best_traffic",
    "second_competitor", "second_position"]] = df.apply(best_comp_info, axis=1)
df = df[df["best_competitor"].notna()].copy()
pool = df[df["Volume"] >= 200].copy()
print(f"Pool: {len(pool)} keyword")

# Cluster atama (04 script ile ayni)
CLUSTERS = {
    "01_MAAS_HESAPLAMA": {
        "include": [r"\b(brüt|brut|net|maaş|maas|asgari ücret|asgari ucret)\b",
                    r"\b(yıllık|aylık|yillik|aylik) maaş",
                    r"\bnet brüt\b", r"\bbrüt net\b",
                    r"\bbordro\b", r"\bsgk\b", r"\bemekli\b", r"\bişe giriş\b"],
        "exclude": [r"\bbayrami\b"],
    },
    "02_HESAPLAMA_ARACLARI": {
        "include": [r"\bhesapla(ma)?\b", r"\bhesap makinesi\b", r"\bhesab[ıi]\b",
                    r"\byüzde\b", r"\byuzde\b", r"\byaş hesap", r"\byas hesap",
                    r"\bkdv\b", r"\bötv\b", r"\botv\b", r"\bvergi\b",
                    r"\bnot ortalama\b", r"\bkredi hesap", r"\btaksit hesap",
                    r"\biban\b", r"\bbenzin hesap", r"\bmotorin hesap",
                    r"\bgöğüs hesap", r"\byükselen burç hesap"],
        "exclude": [r"\binstagram hesab", r"\bpubg hesab", r"\bgmail hesab",
                    r"\bgoogle hesab", r"\bsteam hesab", r"\btwitter hesab"],
    },
    "03_DONUSTURUCU_BIRIM": {
        "include": [r"\b(inc|inç)\b", r"\b(metre|cm|mm|km|gram|kg|miligram|ton|litre|ml|gallon|inch)\b",
                    r"\bkac (saat|dakika|saniye|kg|gram|metre|cm|gb|mb|tl|dolar|euro|km)\b",
                    r"\bçevirme\b", r"\bcevirme\b", r"\bdönüştürücü\b", r"\bdonusturucu\b",
                    r"\bbirim\b"],
        "exclude": [r"\bcumhuriyet bayram\b"],
    },
    "04_DOVIZ_FINANS": {
        "include": [r"\bdolar\b", r"\beuro\b", r"\bsterlin\b", r"\bdöviz\b", r"\bdoviz\b",
                    r"\baltın\b", r"\baltin\b", r"\bborsa\b", r"\bfaiz oran",
                    r"\bbist\b", r"\bbankamatik\b"],
        "exclude": [],
    },
    "05_RESMI_TATIL_BAYRAM": {
        "include": [r"\bresmî tatil\b", r"\bresmi tatil\b", r"\btatil mi\b",
                    r"\bbayram\b", r"\bkurban\b", r"\bramazan\b", r"\bayrefe\b", r"\barefe\b",
                    r"\byılbaş[ıi]\b", r"\byilbasi\b",
                    r"\bcumhuriyet bayram\b", r"\b29 ekim\b", r"\b23 nisan\b",
                    r"\b19 may[ıi]s\b", r"\b30 ağustos\b", r"\b30 agustos\b",
                    r"\bsömestr\b", r"\bsomestr\b", r"\byarı yıl\b", r"\byari yil\b",
                    r"\b15 tatil\b"],
        "exclude": [r"\byunan adaları\b"],
    },
    "06_OZEL_GUNLER": {
        "include": [r"\banneler g[uü]n[uü]\b", r"\bbabalar g[uü]n[uü]\b",
                    r"\bsevgililer g[uü]n[uü]\b", r"\b14 şubat\b", r"\b14 subat\b",
                    r"\böğretmenler g[uü]n[uü]\b", r"\bogretmenler g[uü]n[uü]\b",
                    r"\bçocuklar g[uü]n[uü]\b", r"\bcocuklar g[uü]n[uü]\b",
                    r"\bkad[ıi]nlar g[uü]n[uü]\b", r"\bkadinlar g[uü]n[uü]\b",
                    r"\bdoğum g[uü]n[uü]\b", r"\bdogum g[uü]n[uü]\b",
                    r"\bnewroz\b", r"\bnevruz\b",
                    r"\b8 mart\b", r"\bdoğum tarihi\b", r"\bdogum tarihi\b"],
        "exclude": [],
    },
    "07_MESAJ_SOZ_SIIR_KART": {
        "include": [r"\b(mesaj|söz|sozler|şiir|siir|yazi|notu|kart) ?(ı|i|ları|leri)?\b",
                    r"\b(en güzel|guzel|duygusal|romantik|kısa|uzun|anlamlı|anlamli)\b.*\b(mesaj|söz|sözleri|şiir|sözler)\b"],
        "exclude": [r"\bcarrefour\b", r"\bsd kart\b", r"\bbankkart\b", r"\bkart okuyucu\b",
                    r"\bparaf kart\b", r"\bsim kart bloke\b", r"\bwhatsapp.*mesaj\b",
                    r"\binstagram.*mesaj\b", r"\bsilinen mesaj\b", r"\bmesaj gelmiyor\b",
                    r"\bbiletix\b"],
    },
    "08_NE_ZAMAN_TARIH": {
        "include": [r"\bne zaman\b", r"\bhangi g[uü]n\b", r"\bhangi tarih\b",
                    r"\bhangi ay\b", r"\bkaçında\b", r"\bkacinda\b",
                    r"\bne zamand[ıi]\b", r"\bgeri sayım\b", r"\bgeri sayim\b",
                    r"\bkaç gün kaldı\b", r"\bkac gun kaldi\b"],
        "exclude": [],
    },
    "09_NEDIR_TANIM": {
        "include": [r"\bnedir\b", r"\bne demek\b", r"\bne anlama gelir\b",
                    r"\bne işe yarar\b", r"\bne ise yarar\b"],
        "exclude": [],
    },
    "10_NASIL_YAPILIR": {
        "include": [r"\bnas[ıi]l\b"],
        "exclude": [r"\bnas[ıi]l biri\b", r"\bnas[ıi]l adam\b"],
    },
    "11_KAC_INC_EKRAN": {
        "include": [r"\bkaç (inç|cm|ekran|gb|mb)\b", r"\b(inç|cm|ekran) kaç\b",
                    r"\b(ekran|televizyon|tv) (boyutu|olcusu)\b"],
        "exclude": [],
    },
    "12_WHATSAPP_INSTAGRAM_REHBER": {
        "include": [r"\bwhatsapp\b", r"\binstagram\b", r"\btiktok\b", r"\bfacebook\b",
                    r"\bsnapchat\b", r"\btwitter\b", r"\b\bx giriş\b", r"\btelegram\b"],
        "exclude": [r"\bnedir\b"],
    },
    "13_TELEFON_AYAR_SORUN": {
        "include": [r"\biphone (\d+|\w+) ne|kaç|nasıl\b",
                    r"\bgüvenli arama\b", r"\bnfc\b", r"\bairdrop\b",
                    r"\bairtag\b", r"\bekran goruntusu\b", r"\bekran görüntüsü\b",
                    r"\bekran kayd[ıi]\b", r"\btelefon\b", r"\bsim kart\b",
                    r"\besim\b", r"\bgsm\b", r"\bnumara öğrenme\b", r"\bnumara ogrenme\b"],
        "exclude": [r"\bmasaj salonu\b", r"\bsofa\b"],
    },
    "14_DIZI_FILM_REHBER": {
        "include": [r"\b(dizi|film|sezon|bolum|bölüm) (konusu|oyuncuları|oyuncular)\b",
                    r"\b(oyuncular[ıi]|oyuncular)\b",
                    r"\bnetflix\b", r"\bdisney\+ \b", r"\bexxen\b", r"\btabii\b", r"\bblutv\b",
                    r"\bmaç (hangi|nerede|saat)\b"],
        "exclude": [r"\bmaç tipi\b", r"\bkahve makinesi\b"],
    },
    "15_OYUN_REHBER": {
        "include": [r"\b(gta|pubg|lol|valorant|fortnite|minecraft|fifa|playstation|ps5|ps6|xbox|nintendo|steam|monopoly|wordle)\b"],
        "exclude": [],
    },
    "16_DEPREM_AFET_GUVENLIK": {
        "include": [r"\bdeprem\b", r"\bafad\b", r"\btsunami\b", r"\bafet\b",
                    r"\bguvenlik\b", r"\bgüvenlik\b"],
        "exclude": [r"\bgüvenli arama\b"],
    },
    "17_KIMLIK_EDEVLET_NUMARA": {
        "include": [r"\be-devlet\b", r"\bedevlet\b", r"\btc kimlik\b",
                    r"\bkimlik no\b", r"\bpasaport\b", r"\bvergi numaras\b",
                    r"\bbarkod sorgulama\b", r"\biban sorgu\b", r"\bborç sorgu\b",
                    r"\bplaka sorgu\b", r"\bmuayene\b", r"\btrafik cezas\b",
                    r"\bsgk\b", r"\bn[uü]fus\b"],
        "exclude": [],
    },
    "18_INTERNET_HIZ_MODEM": {
        "include": [r"\binternet h[ıi]z\b", r"\bh[ıi]z testi\b", r"\bspeed test\b",
                    r"\bmodem\b", r"\bwifi\b", r"\brouter\b", r"\bping\b",
                    r"\b192\.168\.\b"],
        "exclude": [],
    },
    "19_POSTA_KODU_ALAN": {
        "include": [r"\bposta kodu\b", r"\balan kodu\b", r"\bil kodu\b",
                    r"\bplaka kodu\b", r"\btelefon kodu\b"],
        "exclude": [],
    },
    "20_SAGLIK_VUCUT": {
        "include": [r"\bbmi\b", r"\bvücut\b", r"\bvucut\b", r"\bkalori\b",
                    r"\bhamilelik\b", r"\bgebelik\b", r"\bdoğum tarihi hesap\b",
                    r"\bregl\b", r"\başı\b", r"\basi takvim\b", r"\bilac\b",
                    r"\bilaç\b", r"\bsemptom\b", r"\bbelirti\b", r"\bhastalik\b",
                    r"\bhastalık\b"],
        "exclude": [],
    },
    "21_KELIME_DIL_YAZIM": {
        "include": [r"\bçeviri\b", r"\bceviri\b", r"\bdil\b", r"\bingilizce\b", r"\btürkçe\b",
                    r"\bnoktal[ıi] virgül\b", r"\bbüyük harf\b", r"\beşanlamlı\b", r"\bzit anlamli\b",
                    r"\bkelime say\b", r"\bkarakter say\b", r"\byaz[ıi]m\b"],
        "exclude": [r"\bturkce dublaj\b"],
    },
    "22_OTOMOTIV_TRAFIK": {
        "include": [r"\bplaka\b", r"\bmuayene\b", r"\btrafik\b", r"\baraç vergisi\b",
                    r"\barac vergisi\b", r"\botv\b", r"\bbenzin (fiyat|hesap)\b",
                    r"\bmotorin (fiyat|hesap)\b", r"\bsürücü kurs\b"],
        "exclude": [],
    },
}

def cluster_match(keyword):
    kw = str(keyword).lower()
    for cname, rules in CLUSTERS.items():
        if any(re.search(p, kw) for p in rules["exclude"]):
            continue
        if any(re.search(p, kw) for p in rules["include"]):
            return cname
    return "ZZ_DIGER"

pool["cluster"] = pool["Keyword"].apply(cluster_match)
pool_sorted = pool.sort_values("Volume", ascending=False)

# Excel olusturma
wb = Workbook()
wb.remove(wb.active)  # bos sheet'i sil

# Stiller
hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="1F4E78")
sub_hdr_font = Font(bold=True, size=11)
sub_hdr_fill = PatternFill("solid", fgColor="D9E1F2")
prio_fill = PatternFill("solid", fgColor="C6EFCE")  # green
mid_fill = PatternFill("solid", fgColor="FFEB9C")   # yellow
border = Border(left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"))

def write_df(ws, df, start_row=1, header_style=True, freeze=True, col_widths=None):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if r_idx == start_row and header_style:
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
    if freeze:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    if col_widths:
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

# ============== SHEET 1: OZET ==============
ws = wb.create_sheet("00_Yonetici_Ozeti")
summary_rows = [
    ["TURKCELL.COM.TR - YENİ TRAFİK FIRSATLARI", "", ""],
    ["Tarih", "13 Mayıs 2026", ""],
    ["Veri Kaynakları", "Ahrefs Content Gap (75.000 kw) + DataForSEO doğrulama (TR/tr)", ""],
    ["", "", ""],
    ["TEMEL BULGULAR", "", ""],
    ["Toplam Turkcell-yok rakip-var fırsat sayısı", 60715, ""],
    ["Volume ≥ 200 keyword", 29624, ""],
    ["Volume ≥ 500 keyword", 12826, ""],
    ["Volume ≥ 500 + KD ≤ 30 keyword", 11889, ""],
    ["", "", ""],
    ["EN BÜYÜK TEK-SAYFA FIRSAT", "", ""],
    ["'hesap makinesi' aylık arama", 2740000, "DfS doğrulama"],
    ["Vatan Bilgisayar /hesap-makinesi sayfası trafiği", 31802, "Tek bir sayfa, 48 kw"],
    ["", "", ""],
    ["KRİTİK CLUSTER'LAR (Volume ≥ 200)", "Keyword Sayısı", "Aylık Toplam Hacim"],
    ["Hesaplama Araçları", 100, 2273400],
    ["Resmi Tatil & Bayram", 425, 631650],
    ["İnternet Hız & Modem", 394, 615350],
    ["Mesaj/Söz/Şiir/Kart", 366, 430750],
    ["Maaş Hesaplama", 150, 409050],
    ["Özel Günler", 175, 271700],
    ["Posta Kodu/Alan", 119, 212650],
    ["Birim Çevirici", 730, 755400],
    ["", "", ""],
    ["RAPORDAKİ İKİ ANA BÖLÜM", "", ""],
    ["Bölüm A (sheet'ler 01-09)", "Kendi araştırma + DataForSEO doğrulama", ""],
    ["Bölüm B (sheet'ler 10-22)", "Ahrefs CSV rakip gap analizi", ""],
]
for r_idx, row in enumerate(summary_rows, 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == 1:
            cell.font = Font(bold=True, size=16, color="FFFFFF")
            cell.fill = hdr_fill
        elif row[0] in ["TEMEL BULGULAR", "EN BÜYÜK TEK-SAYFA FIRSAT", "KRİTİK CLUSTER'LAR (Volume ≥ 200)", "RAPORDAKİ İKİ ANA BÖLÜM"]:
            cell.font = sub_hdr_font
            cell.fill = sub_hdr_fill
ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 40
ws.merge_cells("A1:C1")

# ============== SHEET 2: KENDI ARASTIRMA - DfS DOGRULANMIS ============
ws = wb.create_sheet("01_A_Kendi_Arastirma_DfS")
own_research = pd.DataFrame([
    # A1.1 Genel hesaplayıcılar
    ["A1.1 Genel Matematik", "hesap makinesi (genel hub)", "/hesaplama-araclari/", 2740000, "düşük", "informational", "★★★", "Vatan tek başına 32K trafik alıyor; Turkcell brand çok daha güçlü"],
    ["A1.1 Genel Matematik", "yüzde hesaplama (MEVCUT)", "/yuzde-hesaplama (mevcut)", 673000, "12", "informational", "★★★", "Mevcut sayfa - güçlendir"],
    ["A1.1 Genel Matematik", "yaş hesaplama (MEVCUT)", "/yas-hesaplama (mevcut)", 301000, "9", "informational", "★★★", "Mevcut sayfa - kedi/köpek widget'ı ekle"],
    ["A1.1 Genel Matematik", "kedi yaşı hesaplama", "/kedi-yasi-hesaplama/", 12000, "0-1", "informational", "★★★", "Vodafone 1. sırada; Turkcell kapayabilir"],
    ["A1.1 Genel Matematik", "köpek yaşı hesaplama", "/kopek-yasi-hesaplama/", 6000, "düşük", "informational", "★★", "Tahmini hacim"],
    ["A1.1 Genel Matematik", "iki tarih arası gün hesaplama", "/iki-tarih-arasi-gun/", 5000, "düşük", "informational", "★★", "Çeşit yelpazesi"],
    # A1.2 Finans/maaş
    ["A1.2 Finans/Maaş", "kredi hesaplama", "/kredi-hesaplama/", 1220000, "45", "informational", "★★", "Yüksek KD ama brand otoritesiyle 3-5 sıra"],
    ["A1.2 Finans/Maaş", "brüt-net maaş hesaplama", "/brutten-nete-maas-hesaplama/", 8100, "30", "informational", "★★★", "Vodafone modeli + KDV/yıllık varyantlar"],
    ["A1.2 Finans/Maaş", "KDV hesaplama", "/kdv-hesaplama/", 201000, "19", "informational", "★★★", "Tek başına dev fırsat"],
    ["A1.2 Finans/Maaş", "kıdem tazminatı hesaplama", "/kidem-tazminati-hesaplama/", 201000, "27", "informational", "★★★", "HR kitlesi"],
    ["A1.2 Finans/Maaş", "ihbar tazminatı hesaplama", "/ihbar-tazminati-hesaplama/", 27100, "25", "commercial", "★★★", "Kıdem ile aynı sayfa veya alt"],
    ["A1.2 Finans/Maaş", "yıllık izin hesaplama", "/yillik-izin-hesaplama/", 9900, "4", "informational", "★★★", "Düşük zorluk - acil!"],
    ["A1.2 Finans/Maaş", "fazla mesai hesaplama", "/fazla-mesai-hesaplama/", 9900, "—", "informational", "★★", "HR cluster"],
    ["A1.2 Finans/Maaş", "emeklilik yaşı hesaplama", "/emeklilik-yasi-hesaplama/", 3600, "18", "informational", "★★", "Brand güveni kritik"],
    ["A1.2 Finans/Maaş", "askerlik hesaplama", "/askerlik-hesaplama/", 1900, "—", "informational", "★★", "Erkek 18-35 (core mobile)"],
    ["A1.2 Finans/Maaş", "taksit hesaplama", "/taksit-hesaplama/", 1900, "37", "informational", "★", "Paycell ile bağlantı"],
    ["A1.2 Finans/Maaş", "vade farkı hesaplama", "/vade-farki-hesaplama/", 2400, "0", "informational", "★★★", "KOBİ kitlesi - Ahrefs CSV"],
    ["A1.2 Finans/Maaş", "YouTube gelir hesaplama", "/youtube-gelir-hesaplama/", 2100, "2", "informational", "★★★", "İçerik üreticisi kitle"],
    ["A1.2 Finans/Maaş", "cayma bedeli hesaplama", "/cayma-bedeli-hesaplama/", 3000, "0", "informational", "★★★", "TELEKOM-ÖZEL: turk.net 1.427 trafik"],
    ["A1.2 Finans/Maaş", "iskonto hesaplama", "/iskonto-hesaplama/", 1000, "17", "informational", "★", "İşletme blog"],
    # A1.3 Sağlık
    ["A1.3 Sağlık/Yaşam", "kalori hesaplama", "/kalori-hesaplama/", 74000, "medium", "informational", "★★", "Lifestyle blog"],
    ["A1.3 Sağlık/Yaşam", "BMI hesaplama", "/bmi-hesaplama/", 8100, "—", "informational", "★★", "İdeal kilo ile"],
    ["A1.3 Sağlık/Yaşam", "hamilelik haftası hesaplama", "/hamilelik-haftasi-hesaplama/", 2400, "8", "informational", "★★★", "Düşük KD"],
    ["A1.3 Sağlık/Yaşam", "ovulasyon hesaplama", "/ovulasyon-hesaplama/", 6600, "7", "informational", "★★★", "Düşük KD"],
    ["A1.3 Sağlık/Yaşam", "regl hesaplama", "/regl-takvimi/", 18100, "—", "informational", "★★★", "Kadın kitle"],
    ["A1.3 Sağlık/Yaşam", "doğum tarihi hesaplama (gebelik)", "/dogum-tarihi-hesaplama/", 5000, "düşük", "informational", "★★", "Tahmin"],
    # A1.4 Astroloji
    ["A1.4 Astroloji", "yükselen burç hesaplama", "/yukselen-burc-hesaplama/", 201000, "10", "informational", "★★★", "DEV FIRSAT - düşük KD"],
    ["A1.4 Astroloji", "burç hesaplama (doğum tarihine göre)", "/burc-hesaplama/", 50000, "düşük", "informational", "★★★", "Tahmin"],
    ["A1.4 Astroloji", "Çin burcu hesaplama", "/cin-burcu-hesaplama/", 3000, "düşük", "informational", "★", "Niş"],
    # A1.5 Yazma
    ["A1.5 Yazma Araçları", "kelime sayacı", "/kelime-sayaci/", 18100, "12", "informational", "★★★", "Öğrenci+yazar"],
    ["A1.5 Yazma Araçları", "karakter sayacı", "/karakter-sayaci/", 14800, "3", "informational", "★★★", "EN DÜŞÜK KD"],
    ["A1.5 Yazma Araçları", "büyük harf küçük harf çevirici", "/buyuk-harf-kucuk-harf-cevirici/", 5000, "düşük", "informational", "★★", "Tahmin"],
    ["A1.5 Yazma Araçları", "klavye sembolleri (et işareti @ vb.)", "/klavye-sembolleri/", 72000, "düşük", "informational", "★★", "Vodafone freezone blog'dan"],
    # A1.6 Birim çevirici
    ["A1.6 Birim Çevirici", "döviz çevirici", "/doviz-cevirici/", 74000, "32", "informational", "★★", "Finans portföyü"],
    ["A1.6 Birim Çevirici", "inç-cm çevirici (TV)", "/inc-cm-cevirici/", 100000, "0-1", "informational", "★★", "MediaMarkt baskın"],
    ["A1.6 Birim Çevirici", "GB-MB çevirici (telekom)", "/gb-mb-cevirici/", 3000, "düşük", "informational", "★★", "BRAND UYUMU"],
    ["A1.6 Birim Çevirici", "kg-libre çevirici", "/kg-libre-cevirici/", 10000, "düşük", "informational", "★", "Tahmin"],
    # A1.7 Tarih/Zaman
    ["A1.7 Tarih/Zaman", "bugün ayın kaçı", "/bugun-ayin-kaci/", 90500, "—", "informational", "★★", "Sürekli hacim"],
    ["A1.7 Tarih/Zaman", "saat kaç Türkiye", "/saat-kac-turkiye/", 2900, "37", "informational", "★", "İl bazlı genişlet"],
    ["A1.7 Tarih/Zaman", "saat farkı (ABD, Avrupa)", "/saat-farki-abd/", 5000, "düşük", "informational", "★", "Yurtdışı TR"],
    # A2 Tatil hub
    ["A2 Resmi Tatil/Bayram Hub", "resmi tatiller 2026", "/resmi-tatiller/2026/", 22200, "3", "informational", "★★★", "Trend +%8303"],
    ["A2 Resmi Tatil/Bayram Hub", "resmi tatil günleri", "/resmi-tatiller/", 6600, "8", "informational", "★★★", "Evergreen ana hub"],
    ["A2 Resmi Tatil/Bayram Hub", "kurban bayramı ne zaman", "/kurban-bayrami-ne-zaman/", 1220000, "11", "informational", "★★★", "Mayıs zirve 6M"],
    ["A2 Resmi Tatil/Bayram Hub", "ramazan ne zaman başlıyor 2026", "/ramazan-ne-zaman/", 18000, "0", "informational", "★★★", "Mart spike"],
    ["A2 Resmi Tatil/Bayram Hub", "yarı yıl tatili ne zaman", "/yariyil-tatili/", 11000, "6", "informational", "★★★", "Veliler"],
    ["A2 Resmi Tatil/Bayram Hub", "yaz tatili ne zaman 2026", "/yaz-tatili/", 11000, "26", "informational", "★★", "Sezonsal"],
    ["A2 Resmi Tatil/Bayram Hub", "29 Ekim Cumhuriyet Bayramı", "/29-ekim-cumhuriyet-bayrami/", 450000, "—", "informational", "★★★", "Ekim 5M spike"],
    ["A2 Resmi Tatil/Bayram Hub", "30 Ağustos Zafer Bayramı", "/30-agustos-zafer-bayrami/", 150000, "tahmin", "informational", "★★", "Ağustos spike"],
    ["A2 Resmi Tatil/Bayram Hub", "23 Nisan Ulusal Egemenlik", "/23-nisan/", 200000, "tahmin", "informational", "★★", "Nisan spike"],
    ["A2 Resmi Tatil/Bayram Hub", "10 Kasım Atatürk", "/10-kasim/", 50000, "tahmin", "informational", "★★", "Kasım spike"],
    ["A2 Resmi Tatil/Bayram Hub", "31 aralık tatil mi", "/31-aralik-tatil-mi/", 13000, "0", "informational", "★★★", "Yılbaşı eşi"],
    ["A2 Resmi Tatil/Bayram Hub", "namaz vakitleri (il bazlı)", "/namaz-vakitleri/", 13600000, "15", "informational", "★", "Diyanet dominant"],
    # A3 Özel günler / mesaj
    ["A3 Özel Günler Hub", "anneler günü ne zaman 2026", "/anneler-gunu-ne-zaman/", 673000, "8", "informational", "★★★", "Mayıs 5M"],
    ["A3 Özel Günler Hub", "babalar günü ne zaman 2026", "/babalar-gunu-ne-zaman/", 450000, "—", "informational", "★★★", "Haziran 4M"],
    ["A3 Özel Günler Hub", "sevgililer günü 14 Şubat", "/sevgililer-gunu/", 200000, "tahmin", "informational", "★★★", "Şubat spike"],
    ["A3 Özel Günler Hub", "öğretmenler günü 24 Kasım", "/ogretmenler-gunu/", 150000, "tahmin", "informational", "★★", "Kasım spike"],
    ["A3 Özel Günler Hub", "kadınlar günü 8 Mart", "/8-mart-kadinlar-gunu/", 100000, "tahmin", "informational", "★★", "Mart spike"],
    ["A3 Mesaj/Söz/Kart", "doğum günü mesajları", "/dogum-gunu-mesajlari/", 201000, "—", "informational", "★★★", "Yıl boyu yüksek"],
    ["A3 Mesaj/Söz/Kart", "güzel sözler", "/guzel-sozler/", 165000, "11", "informational", "★★★", "Evergreen"],
    ["A3 Mesaj/Söz/Kart", "günaydın mesajları", "/gunaydin-mesajlari/", 135000, "4", "informational", "★★★", "Sabah trafiği"],
    ["A3 Mesaj/Söz/Kart", "anlamlı sözler", "/anlamli-sozler/", 110000, "4", "informational", "★★★", "Düşük KD!"],
    ["A3 Mesaj/Söz/Kart", "anneler günü mesajları", "/anneler-gunu-mesajlari/", 74000, "—", "informational", "★★★", "Mayıs 1M"],
    ["A3 Mesaj/Söz/Kart", "ramazan bayramı mesajları", "/ramazan-bayrami-mesajlari/", 74000, "—", "informational", "★★★", "Mart 823K"],
    ["A3 Mesaj/Söz/Kart", "iyi geceler mesajları", "/iyi-geceler-mesajlari/", 40500, "—", "informational", "★★", "Akşam"],
    ["A3 Mesaj/Söz/Kart", "sevgililer günü mesajları", "/sevgililer-gunu-mesajlari/", 33100, "—", "informational", "★★★", "Şubat 368K"],
    ["A3 Mesaj/Söz/Kart", "bayram mesajları", "/bayram-mesajlari/", 14800, "—", "informational", "★★★", "Mart 110K"],
    ["A3 Mesaj/Söz/Kart", "babalar günü mesajları", "/babalar-gunu-mesajlari/", 4300, "—", "informational", "★★", "Haziran spike"],
    ["A3 Mesaj/Söz/Kart", "öğretmenler günü güzel sözler", "/ogretmenler-gunu-sozler/", 5600, "1", "informational", "★★★", "Kasım spike"],
    # A4 Telekom yardım
    ["A4 Telekom Yardım", "WiFi şifresi değiştirme", "/wifi-sifresi-degistirme/", 18100, "—", "informational", "★★★", "Modemlere göre"],
    ["A4 Telekom Yardım", "modem arayüzü 192.168.1.1", "/modem-arayuzu-giris/", 106000, "19", "informational", "★★★", "TT 3."],
    ["A4 Telekom Yardım", "internet hız testi", "/hiz-testi/", 1000000, "34", "informational", "★★★", "Mevcutsa kontrol et"],
    ["A4 Telekom Yardım", "telefon IMEI sorgulama", "/imei-sorgulama/", 5400, "26", "informational", "★★★", "Brand güveni"],
    ["A4 Telekom Yardım", "eSIM nedir / nasıl kullanılır", "/esim-nedir/", 6600, "19", "informational", "★★★", "Turkcell eSIM ürünü"],
    ["A4 Telekom Yardım", "5G nedir / 5G destekleyen telefonlar", "/5g-nedir/", 9900, "29", "informational", "★★★", "5G otorite"],
    ["A4 Telekom Yardım", "NFC özelliği nedir/nasıl açılır", "/nfc-nedir/", 23000, "1-2", "informational", "★★★", "Vodafone baskın"],
    ["A4 Telekom Yardım", "güvenli arama nasıl kapatılır", "/guvenli-arama-kapatma/", 20000, "0", "informational", "★★★", "TT 4."],
    ["A4 Telekom Yardım", "konum nasıl atılır / paylaşılır", "/konum-paylasma/", 10000, "0", "informational", "★★★", "Vodafone 7."],
    ["A4 Telekom Yardım", "kendi numaramı öğrenme", "/kendi-numaramı-ogrenme/", 3600, "0", "informational", "★★★", "Turkcell için doğal"],
    ["A4 Telekom Yardım", "e-imza/mobil imza nasıl alınır", "/mobil-imza/", 10000, "2", "informational", "★★★", "GoldenSign ürünü"],
    ["A4 Telekom Yardım", "numara taşıma rehberi", "/numara-tasima/", 5000, "düşük", "informational", "★★★", "Rakipten kazanç"],
    # A4.2 Sorgulama
    ["A4 Sorgulama", "IBAN sorgulama", "/iban-sorgulama/", 33100, "23", "navigational", "★★★", "Paycell entegrasyon"],
    ["A4 Sorgulama", "posta kodu sorgulama", "/posta-kodu-sorgulama/", 14800, "40", "informational", "★★", "İl bazlı alt sayfalar"],
    ["A4 Sorgulama", "TC kimlik no sorgulama", "/tc-kimlik-sorgulama/", 1000, "10", "navigational", "★", "Düşük hacim"],
    ["A4 Sorgulama", "alan kodları (il bazlı)", "/alan-kodlari/", 5000, "3", "informational", "★★", "İl bazlı sayfa - TT model"],
    ["A4 Sorgulama", "il plaka kodları", "/plaka-kodlari/", 5000, "düşük", "informational", "★", "Otomotiv"],
    # A4.3 Dijital servis
    ["A4 Dijital Rehber", "WhatsApp silinen mesaj geri getirme", "/whatsapp-silinen-mesaj/", 16000, "1", "informational", "★★★", "BiP köprüsü"],
    ["A4 Dijital Rehber", "WhatsApp Web giriş", "/whatsapp-web/", 201000, "50", "navigational", "★★", "turk.net 3."],
    ["A4 Dijital Rehber", "Instagram hesabı silme/dondurma", "/instagram-hesap-yonetimi/", 20000, "24-38", "informational", "★★", "turk.net 1-3"],
    ["A4 Dijital Rehber", "Telegram Web kullanımı", "/telegram-web/", 170000, "85", "informational", "★", "Yüksek KD"],
    ["A4 Dijital Rehber", "QR kod nasıl okunur", "/qr-kod-okuma/", 3000, "1", "informational", "★★", "Düşük KD"],
    # A5 Deprem
    ["A5 Deprem/Afet", "deprem güvenlik bilgileri (hub)", "/deprem-guvenlik/", 158000, "3", "informational", "★★★", "turk.net 1., trafik 2571"],
    ["A5 Deprem/Afet", "deprem çantası içeriği", "/deprem-cantasi/", 10000, "düşük", "informational", "★★", "Tahmin"],
    ["A5 Deprem/Afet", "Android/iOS deprem uyarı uygulamaları", "/deprem-uyari-app/", 4300, "düşük", "informational", "★★", "Brand önemli"],
    ["A5 Deprem/Afet", "AFAD uyarı sistemi kayıt", "/afad-uyari/", 3000, "düşük", "informational", "★★", "Tahmin"],
], columns=["Kategori", "Onerilen_Keyword", "URL_Onerisi", "Aylik_Hacim_DfS",
            "Tahmini_KD", "Search_Intent", "Oncelik", "Notlar"])
write_df(ws, own_research, col_widths={"A": 22, "B": 42, "C": 38, "D": 14, "E": 12, "F": 14, "G": 10, "H": 55})

# Renklendirme oncelige gore
for r_idx in range(2, len(own_research) + 2):
    prio_cell = ws.cell(row=r_idx, column=7)
    if prio_cell.value == "★★★":
        for c in range(1, 9):
            ws.cell(row=r_idx, column=c).fill = prio_fill
    elif prio_cell.value == "★★":
        for c in range(1, 9):
            ws.cell(row=r_idx, column=c).fill = mid_fill

# ============== SHEET 3: CLUSTER OZET ==============
ws = wb.create_sheet("02_B_Cluster_Ozeti")
cluster_summary = pool.groupby("cluster").agg(
    Keyword_Sayisi=("Keyword", "count"),
    Toplam_Aylik_Hacim=("Volume", "sum"),
    Ort_Hacim=("Volume", "mean"),
    Ort_KD=("KD", "mean"),
    Max_Hacim=("Volume", "max"),
).reset_index().sort_values("Toplam_Aylik_Hacim", ascending=False)
cluster_summary["Ort_Hacim"] = cluster_summary["Ort_Hacim"].astype(int)
cluster_summary["Toplam_Aylik_Hacim"] = cluster_summary["Toplam_Aylik_Hacim"].astype(int)
cluster_summary["Max_Hacim"] = cluster_summary["Max_Hacim"].astype(int)
cluster_summary["Ort_KD"] = cluster_summary["Ort_KD"].round(1)
write_df(ws, cluster_summary, col_widths={"A": 32, "B": 16, "C": 22, "D": 14, "E": 12, "F": 14})

# ============== SHEET 4: RAKIP URL PATTERN ==============
ws = wb.create_sheet("03_B_Rakip_URL_Patternleri")
def extract_path_prefix(url, depth=2):
    if pd.isna(url):
        return None
    m = re.match(r"https?://[^/]+(/[^?#]*)", str(url))
    if not m:
        return None
    path = m.group(1)
    parts = [p for p in path.split("/") if p]
    if not parts:
        return "/"
    return "/" + "/".join(parts[:depth])

all_patterns = []
for comp in competitors:
    url_col = f"{comp}/: URL"
    traf_col = f"{comp}/: Organic Traffic"
    sub = df[df[url_col].notna() & df[traf_col].notna()].copy()
    sub = sub[sub["Volume"] >= 200]
    sub["path_prefix"] = sub[url_col].apply(lambda x: extract_path_prefix(x, 2))
    stats = sub.groupby("path_prefix").agg(
        kw_sayisi=("Keyword", "count"),
        toplam_trafik=(traf_col, "sum"),
        toplam_hacim=("Volume", "sum"),
    ).sort_values("toplam_trafik", ascending=False).head(15).reset_index()
    stats["Rakip"] = comp
    all_patterns.append(stats[["Rakip", "path_prefix", "kw_sayisi", "toplam_trafik", "toplam_hacim"]])

patterns_df = pd.concat(all_patterns, ignore_index=True)
patterns_df.columns = ["Rakip", "URL_Path_Prefix", "Keyword_Sayisi", "Aylik_Trafik", "Toplam_Hacim"]
patterns_df["Aylik_Trafik"] = patterns_df["Aylik_Trafik"].astype(int)
patterns_df["Toplam_Hacim"] = patterns_df["Toplam_Hacim"].astype(int)
write_df(ws, patterns_df, col_widths={"A": 25, "B": 58, "C": 14, "D": 14, "E": 14})

# ============== SHEET 5: DfS DOGRULAMA ==============
ws = wb.create_sheet("04_DfS_Dogrulama_TR")
dfs_validation = pd.DataFrame([
    ["hesap makinesi", 2740000, "—", "informational", "Vatan baskın - hub fırsatı"],
    ["yüzde hesaplama", 673000, 12, "informational", "MEVCUT - güçlendir"],
    ["yaş hesaplama", 301000, 9, "informational", "MEVCUT - genişlet"],
    ["kredi hesaplama", 1220000, 45, "informational", "Yüksek KD"],
    ["KDV hesaplama", 201000, 19, "informational", "★★★ DEV"],
    ["kıdem tazminatı hesaplama", 201000, 27, "informational", "★★★ HR"],
    ["yükselen burç hesaplama", 201000, 10, "informational", "★★★ DEV"],
    ["ihbar tazminatı hesaplama", 27100, 25, "commercial", "★★★"],
    ["karakter sayacı", 14800, 3, "informational", "★★★ EN DÜŞÜK KD"],
    ["kelime sayacı", 18100, 12, "informational", "★★★"],
    ["yıllık izin hesaplama", 9900, 4, "informational", "★★★ DÜŞÜK KD"],
    ["BMI hesaplama", 8100, "—", "informational", "★★"],
    ["brüt net hesaplama", 8100, 30, "informational", "★★ Vodafone modeli"],
    ["ovulasyon hesaplama", 6600, 7, "informational", "★★★ DÜŞÜK KD"],
    ["regl hesaplama", 18100, "—", "informational", "★★★ Kadın kitle"],
    ["hamilelik haftası hesaplama", 2400, 8, "informational", "★★★"],
    ["emeklilik yaşı hesaplama", 3600, 18, "informational", "★★"],
    ["askerlik hesaplama", 1900, "—", "informational", "★★"],
    ["taksit hesaplama", 1900, 37, "informational", "★"],
    ["fazla mesai hesaplama", 9900, "—", "informational", "★★★"],
    ["vergi iadesi hesaplama", 140, "—", "informational", "Düşük hacim"],
    ["kalori hesaplama", 74000, "medium", "informational", "★★"],
    ["döviz çevirici", 74000, 32, "informational", "★★ Sezonsal"],
    ["bugün ayın kaçı", 90500, "—", "informational", "★★ Sürekli"],
    ["saat kaç türkiye", 2900, 37, "informational", "★"],
    ["anneler günü ne zaman", 673000, 8, "informational", "★★★ Mayıs 5M"],
    ["babalar günü ne zaman", 450000, "—", "informational", "★★★ Haziran 4M"],
    ["kurban bayramı ne zaman", 1220000, 11, "informational", "★★★ Mart-Mayıs zirve"],
    ["29 ekim cumhuriyet bayramı", 450000, "—", "informational", "★★★ Ekim 5M"],
    ["resmi tatiller 2026", 22200, 3, "informational", "★★★ +8303% trend"],
    ["resmi tatil günleri", 6600, 8, "informational", "★★★"],
    ["doğum günü mesajları", 201000, "—", "informational", "★★★"],
    ["güzel sözler", 165000, 11, "informational", "★★★"],
    ["günaydın mesajları", 135000, 4, "informational", "★★★ DÜŞÜK KD"],
    ["anlamlı sözler", 110000, 4, "informational", "★★★ DÜŞÜK KD"],
    ["anneler günü mesajları", 74000, "—", "informational", "★★★ Mayıs 1M"],
    ["ramazan bayramı mesajları", 74000, "—", "informational", "★★★ Mart 823K"],
    ["iyi geceler mesajları", 40500, "—", "informational", "★★"],
    ["sevgililer günü mesajları", 33100, "—", "informational", "★★★ Şubat 368K"],
    ["bayram mesajları", 14800, "—", "informational", "★★★ Mart 110K"],
    ["doğum günü hesaplama", 8100, 15, "informational", "★★"],
    ["yaş günü mesajları", 880, "—", "informational", "★"],
    ["sevgiliye şiir", 6600, "—", "informational", "★★"],
    ["sevgililer günü mesajları", 33100, "—", "informational", "★★★"],
    ["internet hız testi", 1000000, 34, "informational", "★★★ MEVCUTSA KONTROL"],
    ["WiFi şifresi değiştirme", 18100, "—", "informational", "★★★"],
    ["telefon IMEI sorgulama", 5400, 26, "informational", "★★★"],
    ["IBAN sorgulama", 33100, 23, "navigational", "★★★ Paycell"],
    ["posta kodu sorgulama", 14800, 40, "informational", "★★"],
    ["TC kimlik no sorgulama", 1000, 10, "navigational", "★"],
    ["5G nedir", 9900, 29, "informational", "★★★ 5G otorite"],
    ["esim nedir", 6600, 19, "informational", "★★★ eSIM ürünü"],
    ["alan kodları", 260, 3, "informational", "Düşük hacim"],
    ["abd saat farkı", 390, "—", "informational", "★"],
    ["29 ekim mesajları", 880, "—", "informational", "Sezonsal"],
    ["qr kod nasıl okunur", 170, 1, "informational", "Düşük hacim"],
    ["namaz vakitleri", 13600000, 15, "informational", "Diyanet baskın"],
    ["iftar vakti", 165000, "—", "informational", "Sezonsal"],
    ["imsak vakti", 823000, 32, "informational", "Ramazan'da"],
], columns=["Keyword", "Aylik_Hacim", "KD", "Intent", "Notlar"])
write_df(ws, dfs_validation, col_widths={"A": 38, "B": 16, "C": 10, "D": 16, "E": 36})
# DfS sheet'inde de oncelige gore renklendirme
for r_idx in range(2, len(dfs_validation) + 2):
    notlar = str(ws.cell(row=r_idx, column=5).value or "")
    if "★★★" in notlar:
        for c in range(1, 6):
            ws.cell(row=r_idx, column=c).fill = prio_fill
    elif "★★" in notlar:
        for c in range(1, 6):
            ws.cell(row=r_idx, column=c).fill = mid_fill

# ============== SHEET'LER: HER CLUSTER ICIN TOP KEYWORD'LER ==============
cluster_order = [
    "02_HESAPLAMA_ARACLARI", "01_MAAS_HESAPLAMA", "05_RESMI_TATIL_BAYRAM",
    "06_OZEL_GUNLER", "07_MESAJ_SOZ_SIIR_KART", "08_NE_ZAMAN_TARIH",
    "09_NEDIR_TANIM", "10_NASIL_YAPILIR", "13_TELEFON_AYAR_SORUN",
    "18_INTERNET_HIZ_MODEM", "19_POSTA_KODU_ALAN", "17_KIMLIK_EDEVLET_NUMARA",
    "03_DONUSTURUCU_BIRIM", "11_KAC_INC_EKRAN", "16_DEPREM_AFET_GUVENLIK",
    "20_SAGLIK_VUCUT", "21_KELIME_DIL_YAZIM", "14_DIZI_FILM_REHBER",
    "15_OYUN_REHBER", "12_WHATSAPP_INSTAGRAM_REHBER", "22_OTOMOTIV_TRAFIK",
    "04_DOVIZ_FINANS",
]

sheet_counter = 5
for cname in cluster_order:
    sub = pool_sorted[pool_sorted["cluster"] == cname].head(100)
    if len(sub) == 0:
        continue
    short_name = cname[3:][:25]  # numerini at, kisalt
    sheet_name = f"{sheet_counter:02d}_{short_name}"
    sheet_counter += 1
    ws = wb.create_sheet(sheet_name)
    out = sub[["Keyword", "Volume", "KD", "CPC", "SERP features",
               "best_competitor", "best_position", "best_traffic", "best_url",
               "second_competitor", "second_position"]].copy()
    out.columns = ["Keyword", "Aylik_Hacim", "KD", "CPC", "SERP_Features",
                   "En_Iyi_Rakip", "Rakip_Pozisyon", "Rakip_Trafik", "Rakip_URL",
                   "2_Rakip", "2_Pozisyon"]
    write_df(ws, out, col_widths={"A": 38, "B": 14, "C": 8, "D": 8, "E": 28,
                                   "F": 22, "G": 14, "H": 14, "I": 70, "J": 22, "K": 12})

# ============== SHEET: TUM FIRSATLAR (V >= 500) ==============
ws = wb.create_sheet(f"{sheet_counter:02d}_TUM_FIRSATLAR_V500")
sheet_counter += 1
all_top = pool_sorted[pool_sorted["Volume"] >= 500].copy()
out = all_top[["Keyword", "Volume", "KD", "CPC", "SERP features", "cluster",
               "best_competitor", "best_position", "best_traffic", "best_url"]].copy()
out.columns = ["Keyword", "Aylik_Hacim", "KD", "CPC", "SERP_Features", "Cluster",
               "En_Iyi_Rakip", "Pozisyon", "Trafik", "URL"]
write_df(ws, out.head(2000), col_widths={"A": 40, "B": 14, "C": 8, "D": 8, "E": 28,
                                           "F": 28, "G": 22, "H": 12, "I": 14, "J": 70})

# ============== SHEET: YOL HARITASI ==============
ws = wb.create_sheet(f"{sheet_counter:02d}_Yol_Haritasi")
roadmap = pd.DataFrame([
    ["Sprint 1 (0-3 ay)", "/hesaplama-araclari/ ana hub", "2.74M umbrella", "düşük", "★★★", "Calculator hub kurulumu"],
    ["Sprint 1 (0-3 ay)", "/hesap-makinesi/", "2.17M", "3", "★★★", "Genel hesap makinesi (basic+bilimsel+fonksiyonlu)"],
    ["Sprint 1 (0-3 ay)", "/karakter-sayaci/", "14.8K", "3", "★★★", "EN DÜŞÜK KD - hızlı kazanım"],
    ["Sprint 1 (0-3 ay)", "/yillik-izin-hesaplama/", "9.9K", "4", "★★★", "HR cluster başlangıç"],
    ["Sprint 1 (0-3 ay)", "/yukselen-burc-hesaplama/", "201K", "10", "★★★", "DEV fırsat"],
    ["Sprint 1 (0-3 ay)", "/kdv-hesaplama/", "201K", "19", "★★★", "İş dünyası kitle"],
    ["Sprint 1 (0-3 ay)", "/kidem-tazminati-hesaplama/", "201K", "27", "★★★", "HR cluster"],
    ["Sprint 1 (0-3 ay)", "/kelime-sayaci/", "18.1K", "12", "★★★", "Öğrenci+yazar"],
    ["Sprint 1 (0-3 ay)", "/iban-sorgulama/", "33.1K", "23", "★★★", "Paycell entegrasyonu"],
    ["Sprint 1 (0-3 ay)", "/cayma-bedeli-hesaplama/", "3K", "0", "★★★", "TELEKOM-ÖZEL"],
    ["Sprint 2 (3-6 ay)", "/resmi-tatiller/ ana hub", "22.2K (ana) + 600K+ alt", "3", "★★★", "Vodafone modeli"],
    ["Sprint 2 (3-6 ay)", "Her tatil için ne zaman sayfası", "150K-1.22M per tatil", "8-26", "★★★", "12-15 alt sayfa"],
    ["Sprint 2 (3-6 ay)", "Geri sayım widget'lı bayram sayfaları", "13K+ (kaç gün kaldı)", "düşük", "★★★", "5 widget"],
    ["Sprint 2 (3-6 ay)", "Özel gün mesaj/söz/kart galerileri", "Toplam 700K+/ay", "4-12", "★★★", "8-10 galeri sayfa"],
    ["Sprint 2 (3-6 ay)", "29 Ekim hub + alt sayfa", "450K (Ekim 5M)", "—", "★★★", "Spike içerik"],
    ["Sprint 3 (6-9 ay)", "Telekom yardım hub'ı", "30+ sayfa cluster", "0-30", "★★★", "Doğal brand alanı"],
    ["Sprint 3 (6-9 ay)", "/wifi-sifresi-degistirme/", "18.1K", "—", "★★★", "Modem bazlı alt sayfa"],
    ["Sprint 3 (6-9 ay)", "/modem-arayuzu-giris/", "106K", "19", "★★★", "TT 3."],
    ["Sprint 3 (6-9 ay)", "/esim-nedir/ + alt", "6.6K + varyantlar", "19", "★★★", "Turkcell eSIM köprü"],
    ["Sprint 3 (6-9 ay)", "/5g-nedir/ + alt", "9.9K + varyantlar", "29", "★★★", "5G otoriter"],
    ["Sprint 3 (6-9 ay)", "/nfc-rehberi/", "23K toplam", "1-2", "★★★", "Vodafone alıyor"],
    ["Sprint 3 (6-9 ay)", "/mobil-imza/", "10K", "2", "★★★", "GoldenSign ürünü"],
    ["Sprint 4 (9-12 ay)", "Sağlık hesaplayıcı paketi (BMI, kalori, regl, ovulasyon)", "100K+ toplam", "0-12", "★★", "Kadın+sağlık kitle"],
    ["Sprint 4 (9-12 ay)", "Astroloji deep funnel", "200K+", "10-15", "★★", "Yükselen burç altyapısı"],
    ["Sprint 4 (9-12 ay)", "Dizi/film rehberleri (Turkcell TV+ köprü)", "Per dizi 5-100K", "0-23", "★★", "İçerik takvim"],
    ["Sprint 4 (9-12 ay)", "Birim çevirici paketi", "180K+ toplam", "0-32", "★★", "TV inç, GB-MB, döviz"],
    ["Sprint 4 (9-12 ay)", "Klavye sembolleri rehberi", "72K", "düşük", "★★", "Vodafone freezone alıyor"],
], columns=["Sprint", "Sayfa_Onerisi", "Aylik_Hacim", "KD", "Oncelik", "Notlar"])
write_df(ws, roadmap, col_widths={"A": 18, "B": 44, "C": 22, "D": 10, "E": 10, "F": 38})
for r_idx in range(2, len(roadmap) + 2):
    if ws.cell(row=r_idx, column=5).value == "★★★":
        for c in range(1, 7):
            ws.cell(row=r_idx, column=c).fill = prio_fill
    elif ws.cell(row=r_idx, column=5).value == "★★":
        for c in range(1, 7):
            ws.cell(row=r_idx, column=c).fill = mid_fill

# ============== SAVE ==============
wb.save(EXCEL_PATH)
print(f"\nExcel kaydedildi: {EXCEL_PATH}")
print(f"Toplam sheet sayisi: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"  - {s}")
