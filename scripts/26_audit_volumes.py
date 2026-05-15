"""
HTML'deki kart hacimlerini DfS Dogrulama sheet'iyle karsilastir.
Tutarsizliklari listele.
"""
import re
from pathlib import Path
from openpyxl import load_workbook

HTML = Path("/Users/Erdo/Desktop/Claude Projects/Turkcel/index.html")
XLSX = Path("/Users/Erdo/Desktop/Claude Projects/Turkcel/TURKCELL_TRAFIK_FIRSATLARI.xlsx")

text = HTML.read_text(encoding="utf-8")

# Excel'den DfS dogrulanmis kelimeleri ve hacimlerini cek
wb = load_workbook(XLSX)
ws = wb["09_DfS_Dogrulama_TR"]
dfs_data = {}
for r in range(2, ws.max_row + 1):
    kw = ws.cell(row=r, column=1).value
    vol = ws.cell(row=r, column=2).value
    if kw and vol and isinstance(vol, (int, float)):
        # Kelime adini normalize et
        kw_norm = str(kw).lower().strip()
        # '(MEVCUT)', '(*)' gibi suffixleri kaldir
        kw_norm = re.sub(r"\s*\(\*\).*", "", kw_norm).strip()
        kw_norm = re.sub(r"\s*\(mevcut\).*", "", kw_norm, flags=re.IGNORECASE).strip()
        kw_norm = re.sub(r"\s*\(.*?\)", "", kw_norm).strip()
        dfs_data[kw_norm] = int(vol)

print(f"DfS sheet'ten {len(dfs_data)} anahtar kelime hacmi okundu.\n")
print("=" * 80)
print("HTML KART HACIM DOGRULAMA")
print("=" * 80)

# HTML'deki tum kartlardaki hacim ifadelerini kontrol et
# Pattern 1: "anahtar kelime (XXX/ay)" veya "anahtar kelime (XX.XK)"
# Pattern 2: metric-badge coral'da "XXX/ay" benzeri ifadeler

# Her kart icin ana baslik + bullet listelerini al
card_pattern = re.compile(
    r'<div class="idea-card"[^>]*>.*?<h4>(.*?)</h4>.*?<div class="idea-metrics">(.*?)</div>.*?(?:<ul class="idea-bullets">(.*?)</ul>)?.*?(?:<div class="idea-kw">(.*?)</div>)?',
    re.DOTALL
)

# Hacim ifadelerini bul: (NNK/ay), (NNN/ay), N.NM, vb.
volume_in_text = re.compile(r"\(([0-9.,]+[KM]?[\s/+]*ay?)\)")
volume_kpi = re.compile(r"<div class=\"v\">([0-9.,KM]+)</div>")

problems = []
verifications = []

# Kart bazli kontrol
for m in card_pattern.finditer(text):
    title = re.sub(r"<[^>]+>", "", m.group(1)).strip()
    metrics = m.group(2)
    bullets = m.group(3) or ""
    kw_block = m.group(4) or ""

    # Metric badge'larindaki hacmi al
    coral_match = re.search(r'metric-badge coral">([^<]+)</span>', metrics)
    if coral_match:
        metric_str = coral_match.group(1)
        verifications.append({
            "card": title[:60],
            "metric": metric_str,
            "kw_block": (kw_block[:200] + "..") if len(kw_block) > 200 else kw_block,
        })

# Hava durumu kontrolu
print("\n--- A6.2 Hava Durumu Mega-Hub ---")
hava_kws = ["hava durumu", "yarınki hava durumu", "istanbul hava durumu",
             "ankara hava durumu", "izmir hava durumu", "5 günlük hava durumu",
             "10 günlük hava durumu", "saatlik hava durumu", "haftalık hava durumu"]
total = 0
for k in hava_kws:
    if k in dfs_data:
        total += dfs_data[k]
        print(f"  {k}: {dfs_data[k]:,}")
print(f"  TOPLAM: {total:,}/ay")

# Mesaj galerileri toplami
print("\n--- A3 Günlük Mesaj Galerileri ---")
msg_kws = ["günaydın mesajları", "iyi geceler mesajları", "güzel sözler", "anlamlı sözler"]
total = 0
for k in msg_kws:
    if k in dfs_data:
        total += dfs_data[k]
        print(f"  {k}: {dfs_data[k]:,}")
print(f"  TOPLAM: {total:,}/ay")

# Hesaplama hub toplam
print("\n--- A1 Calculator Hub (mevcut + onerilen) ---")
calc_kws = ["hesap makinesi", "yüzde hesaplama", "yaş hesaplama",
             "kdv hesaplama", "kıdem tazminatı hesaplama", "ihbar tazminatı hesaplama",
             "yıllık izin hesaplama", "fazla mesai hesaplama", "kalori hesaplama",
             "bmi hesaplama", "brüt net hesaplama", "kaç gün kaldı",
             "döviz çevirici"]
total = 0
for k in calc_kws:
    if k in dfs_data:
        total += dfs_data[k]
        print(f"  {k}: {dfs_data[k]:,}")
print(f"  TOPLAM: {total:,}/ay")

# Bayram + Tatil toplam
print("\n--- A2 Tatil + Bayram Hub ---")
tatil_kws = ["kurban bayramı ne zaman", "anneler günü ne zaman", "babalar günü ne zaman",
              "29 ekim cumhuriyet bayramı", "resmi tatiller 2026", "resmi tatil günleri"]
total = 0
for k in tatil_kws:
    if k in dfs_data:
        total += dfs_data[k]
        print(f"  {k}: {dfs_data[k]:,}")
print(f"  TOPLAM: {total:,}/ay")

# Tum DfS verilerini hizli liste
print("\n" + "=" * 80)
print("TUM DfS DOGRULANMIS KW'ler (volume sirasi)")
print("=" * 80)
for kw, v in sorted(dfs_data.items(), key=lambda x: -x[1])[:30]:
    print(f"  {v:>12,}  {kw}")

# HTML'deki kart hacim ifadelerini ozetle
print("\n\n" + "=" * 80)
print("HTML KART HACIM BADGE'LERI")
print("=" * 80)
for v in verifications[:50]:
    print(f"  [{v['metric']:>30}]  {v['card']}")
