"""
Excel'i HTML rapor akisina tam uygun hale getir:
1. A1 sheet: Siralamayi HTML'e gore yeniden duzenle + 'Kac Gun Kaldi Hesaplama' satirini ekle/koru
2. A4 sheet: Row 3 baslik = 'INTERNET YARDIM HUB'
3. A5 sheet: 'CICEK / HEDIYE / ETKINLIK' satirini sil, BURC'u yukari kaydir
4. A6 sheet: Row 2 Veri Calc'e '(konusuldu)' etiketi ekle
5. Yonetici Ozeti: Faz hedeflerini ekle + A6 isaretlerini guncelle
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

XLSX = Path("/Users/Erdo/Desktop/Claude Projects/Turkcel/TURKCELL_TRAFIK_FIRSATLARI.xlsx")
wb = load_workbook(XLSX)

star_fill = PatternFill("solid", fgColor="FFD966")
coral_fill = PatternFill("solid", fgColor="FFE3D8")
mint_fill = PatternFill("solid", fgColor="C8E6C9")
hdr_fill = PatternFill("solid", fgColor="10332F")

# ============================================================
# 1. A4 Sheet: Row 3 basligi "INTERNET YARDIM HUB" yap
# ============================================================
ws = wb["04_A4_Telekom_Sorgu_Fikirleri"]
cur = ws.cell(row=3, column=1).value
ws.cell(row=3, column=1, value="A4.2 INTERNET YARDIM HUB (Superonline)")
# URL onerisini de guncelle
url_val = ws.cell(row=3, column=2).value
if url_val:
    ws.cell(row=3, column=2, value="/superonline-yardim/ veya /internet-yardim/")
print(f"A4 Row 3: '{cur}' -> 'A4.2 INTERNET YARDIM HUB (Superonline)'")

# ============================================================
# 2. A5 Sheet: Cicek/Hediye/Etkinlik satirini sil
# ============================================================
ws = wb["05_A5_AI_Siber_Diger"]
# Row 9 = "A5.8 ÇİÇEK / HEDIYE / ETKİNLİK"
to_delete = None
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if v and ("ÇİÇEK" in str(v) or "Cicek" in str(v) or "ETKİNLİK" in str(v).upper()):
        to_delete = r
        break
if to_delete:
    ws.delete_rows(to_delete)
    print(f"A5: Cicek/Hediye/Etkinlik satiri silindi (row {to_delete})")
    # Burc satirinin numarasini guncelle: A5.9 -> A5.8
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and "BURÇ" in str(v).upper():
            new_v = str(v).replace("A5.9", "A5.8")
            ws.cell(row=r, column=1, value=new_v)
            print(f"A5: BURC satiri A5.8'e taşındı")
            break

# ============================================================
# 3. A6 Sheet: Row 2 Veri Calc'e "(konusuldu)" ekle + star_fill
# ============================================================
ws = wb["06_A6_2026_Plan_Perspektif"]
v = ws.cell(row=2, column=1).value
if v and "konuşuldu" not in str(v) and "konusuldu" not in str(v):
    new_v = str(v) + " (konuşuldu)"
    ws.cell(row=2, column=1, value=new_v)
    ws.cell(row=2, column=1).fill = star_fill
    print(f"A6 Row 2: '(konuşuldu)' etiketi eklendi")

# ============================================================
# 4. A1 Sheet: Sirayi HTML'e gore yeniden duzenle
#    HTML sirasi:
#    1. Yüzde Hesaplama, 2. Yaş, 3. Kaç Gün Kaldı, 4. Brütten Nete,
#    5. KDV, 6. Kıdem-İhbar, 7. Çalışan Hakları, 8. Yükselen Burç,
#    9. Sağlık, 10. Yazma, 11. Dil-Sözlük, 12. Birim, 13. Zaman,
#    14. Kredi/Taksit, 15. Vergi
# ============================================================
ws = wb["01_A1_Hesaplama_Fikirleri"]

# Once tum satirlari dict olarak topla
all_rows = {}
for r in range(2, ws.max_row + 1):
    title = str(ws.cell(row=r, column=1).value or "")
    if not title.strip():
        continue
    row_data = []
    for c in range(1, ws.max_column + 1):
        row_data.append(ws.cell(row=r, column=c).value)
    all_rows[title] = row_data

# Hedef siralamayi tanimla (key match)
target_order = [
    ("yuzde-hesaplama", "FIKIR 1: /yuzde-hesaplama genişletme (MEVCUT)"),
    ("yas-hesaplama", "FIKIR 2: /yas-hesaplama genişletme (MEVCUT)"),
    ("KAC_GUN_KALDI", "FIKIR 3: /kac-gun-kaldi-hesaplama/ (Markaya iletilmiş)"),
    ("brutten-nete", "FIKIR 4: /brutten-nete-maas-hesaplama (Vodafone modeli)"),
    ("kdv-hesaplama", "FIKIR 5: /kdv-hesaplama/"),
    ("kidem-ve-ihbar", "FIKIR 6: /kidem-ve-ihbar-tazminati-hesaplama/"),
    ("calisan-haklari", "FIKIR 7: /calisan-haklari-hesaplayicilari/ mini-hub"),
    ("burc-ve-astroloji", "FIKIR 8: /burc-ve-astroloji-hesaplayicilari/ -> Ticari yonlendirme (paket + Pasaj)"),
    ("saglik-hesaplayicilari", "FIKIR 9: /saglik-hesaplayicilari/ mini-hub"),
    ("yazma-araclari", "FIKIR 10: /yazma-araclari/ mini-hub"),
    ("dil-ve-sozluk", "FIKIR 11: /dil-ve-sozluk-araclari/"),
    ("birim-cevirici", "FIKIR 12: /birim-cevirici/ mini-hub"),
    ("zaman-araclari", "FIKIR 13: /zaman-araclari/ mini-hub"),
    ("kredi-ve-taksit", "FIKIR 14: /kredi-ve-taksit-hesaplama/"),
    ("vergi-hesaplama", "FIKIR 15: /vergi-hesaplama-hub/"),
]

# Mevcut satirlari yeni siraya tasi
# Once tum mevcut data satirlarini sil
while ws.max_row >= 2:
    ws.delete_rows(2)

# Kac Gun Kaldi yeni eklenecek bir satir, manuel ekle
target_data = {
    "KAC_GUN_KALDI": [
        "FIKIR 3: /kac-gun-kaldi-hesaplama/ (Markaya iletilmiş)",
        "/kac-gun-kaldi-hesaplama/",
        "kaç gün kaldı + 1.4M küme",
        "33.100/ay (ana) + küme 1.4M",
        "düşük",
        "informational",
        "★★★ MARKAYA İLETİLMİŞ",
        "kaç gün kaldı, kaç gün kaldı hesaplama (3.600), bayrama kaç gün kaldı (2.600), kurban bayramına kaç gün kaldı (3.200+), ramazan'a kaç gün kaldı (13K), iki tarih arası gün hesaplama (22.200), tarih arası süre hesap, doğum gününe kaç gün kaldı, tatile kaç gün kaldı, 2026 yılbaşına kaç gün, geri sayım widget'lı tüm tarihler",
        "Tek hub'tan dinamik geri sayım, geri sayım'ı pop-up değil URL parametresiyle paylaşılabilir yap."
    ]
}

# Hedef sirayla satirlari ekle
def find_match(all_rows, key):
    for title, data in all_rows.items():
        if key.lower() in title.lower():
            return data
    return None

new_rows = []
for key, new_title in target_order:
    if key == "KAC_GUN_KALDI":
        # Manuel ekle
        d = target_data["KAC_GUN_KALDI"]
        new_rows.append(d)
    else:
        d = find_match(all_rows, key)
        if d:
            # Title'i guncelle (yeni numarayla)
            d = list(d)
            d[0] = new_title
            new_rows.append(d)
        else:
            print(f"  UYARI: {key} bulunamadı")

# Yaz
for row_data in new_rows:
    ws.append(row_data)

print(f"A1: {len(new_rows)} satir yeniden duzenlendi")
for i, row in enumerate(new_rows, 1):
    title = row[0]
    if "(*)" in str(title) or "MEVCUT" in str(title) or "Markaya" in str(title) or "konuşuldu" in str(title):
        ws.cell(row=i+1, column=1).fill = star_fill

# ============================================================
# 5. Yonetici Ozeti: Faz hedeflerini ekle
# ============================================================
ws = wb["00_Yonetici_Ozeti"]
print(f"\nYonetici Ozeti sheet'i kontrol ediliyor...")
print(f"  max_row: {ws.max_row}")

# Mevcut son satirin altina FAZ HEDEFLERI ekle
last_row = ws.max_row + 2

ws.cell(row=last_row, column=1, value="FAZ YOL HARİTASI (Bölüm 12)").font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
ws.cell(row=last_row, column=1).fill = hdr_fill
ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=3)

faz_data = [
    ["FAZ", "Dönem", "Hedef Tahmini Click"],
    ["FAZ 1 - Hızlı Kazanım", "0-3 ay", "150-250K ek aylık tık"],
    ["FAZ 2 - Yeni Yetkinlik", "3-6 ay", "350-500K ek aylık tık"],
    ["FAZ 3 - Otorite Genişletme", "6-12 ay", "300-450K ek aylık tık"],
]

for i, row in enumerate(faz_data):
    r = last_row + 1 + i
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        if i == 0:
            cell.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
            cell.fill = hdr_fill
        else:
            cell.font = Font(bold=False, size=11, name="Calibri")
            if c == 1:
                cell.fill = coral_fill if i == 1 else (PatternFill("solid", fgColor="FFF8E1") if i == 2 else mint_fill)

print(f"Yonetici Ozeti'ne Faz hedefleri eklendi (rows {last_row} - {last_row + 4})")

# ============================================================
# Kayit et
# ============================================================
wb.save(XLSX)
print(f"\n=== Kayit edildi: {XLSX} ===")
