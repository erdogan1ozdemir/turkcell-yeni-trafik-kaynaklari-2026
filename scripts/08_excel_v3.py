"""
Excel v3: 2026 Plan perspektifleri eklendi (A6), markaya iletilen fikirler (*) ile isaretlendi.
"""
import pandas as pd
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

EXCEL_PATH = "/Users/Erdo/Desktop/Claude Projects/Turkcel/output/TURKCELL_TRAFIK_FIRSATLARI.xlsx"

# Mevcut workbook'u yukle
wb = load_workbook(EXCEL_PATH)

# Stil
hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="1F4E78")
sub_hdr_font = Font(bold=True, size=11)
sub_hdr_fill = PatternFill("solid", fgColor="D9E1F2")
prio_fill = PatternFill("solid", fgColor="C6EFCE")  # green - acil
mid_fill = PatternFill("solid", fgColor="FFEB9C")   # yellow - hizli
star_fill = PatternFill("solid", fgColor="FFD966")  # orange - markaya iletilmis
border = Border(left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"))

def write_df(ws, dataf, start_row=1, freeze=True, col_widths=None):
    for r_idx, row in enumerate(dataframe_to_rows(dataf, index=False, header=True), start_row):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if r_idx == start_row:
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
    if freeze:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    if col_widths:
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

# Mevcut sheet'lerin sırasını koru, A6 sheet'ini araya ekle (5 ve 6 arasına)
# 01-05 mevcut A1-A5 fikirleri, 06 cluster ozet, 07 rakip url, 08 dfs dogrulama
# Hedef sıra: 00 ozet, 01-05 A1-A5, 06_A6_2026_Plan (YENİ), 07 cluster, 08 rakip, 09 dfs, 10+ niche'ler

# Mevcut sheet'leri listele
print("Mevcut sheet'ler:", wb.sheetnames)

# Yöneticik özetini güncelle - A6 maddesini ekle
ws_summary = wb["00_Yonetici_Ozeti"]
# Mevcut içeriği temizleyelim, yeniden yazalım
wb.remove(ws_summary)
ws = wb.create_sheet("00_Yonetici_Ozeti", 0)  # ilk sheet olarak

rows = [
    ["TURKCELL.COM.TR - YENİ TRAFİK FIRSATLARI v3 (2026 Plan Perspektifleri Eklendi)", "", ""],
    ["Tarih", "13 Mayıs 2026", ""],
    ["Veri", "Ahrefs Content Gap (75K kw) + DfS (150+ kw doğrulandı) + Markaya iletilmiş Excel + Sunum görselleri", ""],
    ["", "", ""],
    ["TEMEL SAYILAR", "", ""],
    ["Toplam Turkcell-yok rakip-var fırsat", 60715, ""],
    ["Volume ≥ 10", 60715, "Hepsi"],
    ["Volume ≥ 100", 48087, ""],
    ["Volume ≥ 500", 12826, ""],
    ["", "", ""],
    ["2026 PLAN İLE HİZALAMA", "", ""],
    ["Hesaplama sayfaları (A1 + A6.1)", "5.5M hacim hedefi", "500K+ click"],
    ["Mass trafik LP'ler (A6.2 Hava Durumu, A2, A3)", "44M hacim hedefi", "500K+ click"],
    ["Paket listeleme içerikler (A6.8)", "Mevcut sayfalar", "%7-10 click artış"],
    ["Aylık blog (A4/A5 cluster'larını besler)", "—", "%5-7 click artış"],
    ["Schema markup (her sayfa tipi)", "—", "%6-10 click artış"],
    ["Persona tarifeler (A6.5)", "Düşük hacim, yüksek CR", "50K+ click"],
    ["Sınırsız mecra paketleri (A6.4)", "Facet sayfalar", "%8 click artış"],
    ["Teknoloji sözlüğü (A6.6)", "200K+/ay direkt", "AI Overview ideal"],
    ["4.5G karşılaştırma fark (A6.7)", "UX/CR odaklı", "Pop-up genişletme"],
    ["Coverage Map (A6.3 - Verizon modeli)", "5G trend +%1614 quarterly", "TR launch ile patlayacak"],
    ["", "", ""],
    ["EN BÜYÜK DOĞRULANMIŞ FIRSATLAR", "Aylık Hacim", "KD/Not"],
    ["hava durumu (*)", 55600000, "KD 39 (mega LP A6.2)"],
    ["yarınki hava durumu (*)", 16600000, "KD 23"],
    ["istanbul hava durumu", 7480000, "KD 19"],
    ["hesap makinesi", 2740000, "Vatan modeli"],
    ["kurban bayramı ne zaman", 1220000, "KD 11"],
    ["kredi hesaplama", 1220000, "KD 45"],
    ["internet hız testi", 1000000, "KD 34"],
    ["yüzde hesaplama (MEVCUT) (*)", 673000, "KD 12"],
    ["anneler günü ne zaman", 673000, "KD 8"],
    ["babalar günü ne zaman", 450000, "Haziran 4M"],
    ["29 ekim cumhuriyet bayramı", 450000, "Ekim 5M"],
    ["asgari ücret 2026", 368000, "KD 9, +%210K trend"],
    ["IMEI Sorgulama (*)", 550000, "KD 14"],
    ["yaş hesaplama (MEVCUT) (*)", 301000, "KD 9"],
    ["KDV hesaplama (*)", 201000, "KD 19"],
    ["kıdem tazminatı hesaplama", 201000, "KD 27"],
    ["yükselen burç hesaplama", 201000, "KD 10"],
    ["doğum günü mesajları", 201000, "yıl boyu"],
    ["güzel sözler", 165000, "KD 11"],
    ["sgk işe giriş", 135000, "KD 20"],
    ["günaydın mesajları", 135000, "KD 4"],
    ["güvenli arama kapatma (*)", 142000, "küme"],
    ["anlamlı sözler", 110000, "KD 4"],
    ["hgs bakiye sorgulama", 110000, "KD 19"],
    ["İndirim Hesaplama (*)", 49500, "DfS doğrulandı"],
    ["Kaç Gün Kaldı (*)", 33100, "küme 1.4M"],
    ["Gizli Numara Kapatma (*)", 27100, "Turkcell URL'ler dağınık"],
    ["", "", ""],
    ["(*) İŞARETLİ FİKİRLER MARKAYA İLETİLMİŞ / KONUŞULMUŞ", "", ""],
    ["", "", ""],
    ["RAPOR YAPISI", "", ""],
    ["Sheet 01-05: A1-A5 Kendi araştırma fikirleri", "", ""],
    ["Sheet 06: A6 - 2026 PLAN PERSPEKTİFLERİ (YENİ)", "8 yeni fikir", ""],
    ["Sheet 07: B1 Cluster Özeti (35 niche)", "", ""],
    ["Sheet 08: B2 Rakip URL Patternleri", "", ""],
    ["Sheet 09: DfS Doğrulama (150+ kw)", "", ""],
    ["Sheet 10+: Niche cluster detaylar (top 200-300 her cluster)", "", ""],
    ["Son sheet: TUM_60K_FIRSATLAR (filtrelenebilir)", "", ""],
]
for r_idx, row in enumerate(rows, 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == 1:
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.fill = hdr_fill
        elif val in ["TEMEL SAYILAR", "2026 PLAN İLE HİZALAMA", "EN BÜYÜK DOĞRULANMIŞ FIRSATLAR", "RAPOR YAPISI", "(*) İŞARETLİ FİKİRLER MARKAYA İLETİLMİŞ / KONUŞULMUŞ"]:
            cell.font = sub_hdr_font
            cell.fill = sub_hdr_fill
        elif "(*)" in str(val):
            cell.fill = star_fill
ws.column_dimensions["A"].width = 60
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 40
ws.merge_cells("A1:C1")

# A6 YENİ SHEET ekle (5 ve 6 arası, yani index 6)
sheet_idx = wb.sheetnames.index("06_B1_Cluster_Ozeti")
ws_a6 = wb.create_sheet("06_A6_2026_Plan_Perspektif", sheet_idx)

a6_data = pd.DataFrame([
    # A6.1 DATA CALCULATOR
    ["A6.1 Veri/Data Kullanım Hesaplayıcı (Vodafone UK Modeli)",
     "/data-hesaplayicim/", "ne kadar internet lazım + youtube/netflix/instagram ne kadar veri",
     "~10-15K + dolaylı CR", "düşük", "informational", "★★★ AI ideal",
     "ne kadar internet lazım, kaç gb internet yeterli, internet kullanım hesaplama, youtube ne kadar veri tüketir, netflix ne kadar veri, instagram ne kadar veri, tiktok ne kadar veri, spotify ne kadar veri, 1 saat film kaç gb, 1 saat youtube ne kadar veri, 1 gb kaç saat youtube, 5 gb ne kadar yeter, 1 gb kaç mb (5.4K), mb gb dönüştürücü, kb to gb dönüştürücü (*), data calculator turkcell",
     "Vodafone UK örneği: 469 KW, 56.6K trafik, tüm informational - AI Overview için ideal"],
    # A6.2 HAVA DURUMU
    ["A6.2 Hava Durumu Mega-Hub (*)",
     "/hava-durumu/ + 81 il alt sayfa", "hava durumu + il bazlı + yarınki + 5 günlük",
     "95M+ (DfS doğrulandı)", "19-39", "informational", "★★★ MARKAYA İLETİLMİŞ",
     "hava durumu (55.6M), yarınki hava durumu (16.6M), istanbul hava durumu (7.5M), ankara (6.1M), izmir (3.4M), 5 günlük (110K), 10 günlük (90K), saatlik (49K), haftalık (33K), yarın hava nasıl olacak (18K), tüm 81 il + ilçeleri",
     "Markaya iletilmiş (referans Excel 41.6M). ICT desteği gerekli. MGM API."],
    # A6.3 COVERAGE MAP
    ["A6.3 5G Kapsama Haritası (Verizon Modeli)",
     "/kapsama-haritasi/", "5g kapsama + kapsama alanı sorgulama",
     "~1K + mega trend", "düşük-29", "informational/navigational", "★★★",
     "5g kapsama alanı (480, +%1025), 5g kapsama haritası (170, +%2500 quarterly), 5g hızı (260, +%1344), 5g nedir (9.9K), turkcell kapsama alanı (50), kapsama alanı sorgulama (30), 5g destekleyen telefonlar, 5g hangi illerde, 4.5g 5g farkı, gerçek 5g, 5g için telefon değiştirmek gerekir mi",
     "Verizon: 7041 KW / 109K trafik / %1 toplam. TR 5G launch ile patlayacak. Tahmini 30-100K aylık trafik 12 ay sonra."],
    # A6.4 SINIRSIZ MECRA
    ["A6.4 Sınırsız Mecra Paketleri (Plan Maddesi 7)",
     "/sinirsiz-paketler/ veya facet", "sınırsız [mecra] paketi",
     "~10-15K küme + dolaylı CR", "29 (ana)", "commercial", "★★★",
     "sınırsız internet paketi (3.6K), sınırsız sosyal medya paketi (1.6K), sınırsız youtube paketi (590), sınırsız tiktok paketi (70), sınırsız whatsapp paketi (90), sınırsız instagram paketi (70), sınırsız netflix paketi, sınırsız spotify paketi, sınırsız bip paketi, sınırsız konuşma, sınırsız dakika, gençlere özel paket, oyuncu tarifesi",
     "Plan hedefi: %8 click artışı (150K+). Facet sayfaları her mecraya özel landing page."],
    # A6.5 PERSONA
    ["A6.5 Persona Odaklı Tarife Sayfaları (Plan Maddesi 6)",
     "/tarife-personalari/", "öğrenci/asker/emekli/kobi tarifesi",
     "~3-5K küme + CR", "1-56", "commercial/informational", "★★★",
     "öğrenci tarifeleri (320, KD 43), öğrenci tarifesi, asker tarifesi, emekli tarifesi (140, KD 1), kobi tarifesi, işletme paketi, engelli tarifesi, aile paketi, ikinci hat tarifesi, gençler için tarife, genç paketi, yabancı uyruklu hat, expat tarife, oyuncu tarifesi, gamer paketi",
     "Plan hedefi: 50K+ click. Her persona için tipik kullanım + tarife öneri + FAQ."],
    # A6.6 TEKNOLOJI SOZLUGU
    ["A6.6 Teknoloji Sözlüğü Hub (Plan Maddesi 7)",
     "/teknoloji-sozlugu/", "100+ teknik terim 'nedir' sayfası",
     "200K+/ay direkt", "2-36", "informational", "★★★ AI ideal",
     "gsm nedir (14.8K), 5g nedir (9.9K), ram nedir (9.9K, KD 5), esim nedir (6.6K, KD 19), ip adresi nedir (6.6K), dns nedir (5.4K), router nedir (5.4K), fiber internet nedir (1.9K, KD 2), modem nedir (1K), vlan nedir (720), wifi nedir (320), fttx nedir (90), threads nedir (22K), podcast nedir (21K), grok nedir (19K), muud nedir (11K, Turkcell kendi markası!), muud premium nedir (15K), nfc nedir (11K), network nedir (10K), airtag nedir (10K), spam ne demek (14K), canva nedir (14K), repost ne demek (13K), deepfake nedir (9.1K)",
     "Schema: DefinedTerm + FAQPage. Muud nedir/Muud Premium TT'ye gidiyor - Turkcell kendi açıklaması yok!"],
    # A6.7 TARIFE KARSILASTIRMA FARK
    ["A6.7 Tarife Karşılaştırma 'Fark' Detayı (Plan Maddesi 4)",
     "/paket-ve-tarifeler/4-5-g-hizinda pop-up", "Mevcut karşılaştırmaya detay",
     "Direkt arama yok - UX/CR", "—", "transactional", "★★",
     "Fark alanı: GB miktarı, dakika her yöne, SMS adet, hotspot izni + GB limit + kısıtlamalar, 5G destek + hız sınırı, roaming ülke listesi, dijital servisler (BiP/fizy/TV+/lifebox), cihaz desteği (eSIM uyumlu), sözleşme taahhüt + cayma, avantajlar (genç indirimi, mağaza puanı)",
     "Pop-up'ta görsel + eksik/fark detayı. ComparisonTable schema. A6.1 (Data Calculator) ile bağlantı."],
    # A6.8 PAKET LISTELEME ICERIK
    ["A6.8 Paket Listeleme Kategori İçerikleri (Plan Maddesi 3)",
     "Mevcut /paket-ve-tarifeler/* sayfaların üstüne içerik",
     "Her kategori sayfası", "Plan: %7-10 artış (250K+)", "düşük-orta", "commercial", "★★★",
     "Önerilen sayfalar: /paket-ve-tarifeler/ (ana), /4-5-g-hizinda (4.5G ve 5G karşılaştırma), /superonline-fiber-internet/ (fiber paketler), /genc-tarifeleri (sosyal medya & eğlence), /aile-paketleri (çift+ek hat). İçerik: kategori tanımı 200-300 kelime + karşılaştırma tablosu + 5-10 FAQ + internal link + schema (BreadcrumbList, FAQPage, ItemList)",
     "Plan hedefi: %7-10 click artışı (250K+). AI Overview için FAQ optimizasyonu kritik."],
], columns=["Fikir", "URL_Onerisi", "Ana_Keyword", "Aylik_Hacim", "KD", "Intent", "Oncelik",
            "Oynanabilecek_KW_Listesi", "Notlar"])
write_df(ws_a6, a6_data, col_widths={"A": 48, "B": 38, "C": 38, "D": 22, "E": 14, "F": 22, "G": 26, "H": 90, "I": 60})
# Hepsi acil (yeni perspektif olduğu için star_fill)
for r_idx in range(2, len(a6_data) + 2):
    for c in range(1, 10):
        ws_a6.cell(row=r_idx, column=c).fill = prio_fill

# YENİ DfS doğrulama sheet'i ekle - eksikleri tamamlamak için update et
# 08_DfS_Dogrulama_TR sheet'ine ek doğrulama ekleyelim
ws_dfs = wb["08_DfS_Dogrulama_TR"]
# Mevcut satır sayısını bul ve sonuna ekle
last_row = ws_dfs.max_row + 1
new_dfs_rows = [
    # A6 doğrulamaları (yeni)
    ["--- A6.2 HAVA DURUMU (YENİ) ---", "", "", "", ""],
    ["hava durumu (*)", 55600000, 39, "informational", "★★★ MASS LP (markaya iletilmiş)"],
    ["yarınki hava durumu (*)", 16600000, 23, "informational", "★★★"],
    ["istanbul hava durumu", 7480000, 19, "informational", "★★★ il bazlı"],
    ["ankara hava durumu", 6120000, 28, "informational", "★★★ il bazlı"],
    ["izmir hava durumu", 3350000, 23, "informational", "★★★ il bazlı"],
    ["5 günlük hava durumu", 110000, 20, "informational", "★★★"],
    ["10 günlük hava durumu", 90500, 27, "informational", "★★"],
    ["saatlik hava durumu", 49500, 16, "informational", "★★★"],
    ["haftalık hava durumu", 33100, 30, "informational", "★★"],
    ["yarın hava nasıl olacak", 18100, 12, "informational", "★★★ DÜŞÜK KD"],
    ["--- A6.3 5G KAPSAMA ---", "", "", "", ""],
    ["5g kapsama alanı", 480, 29, "informational", "+%1025 yıllık trend"],
    ["5g kapsama haritası", 170, "düşük", "informational", "+%2500 quarterly!"],
    ["5g hızı", 260, 14, "informational", "+%1344 quarterly"],
    ["turkcell kapsama alanı", 50, 62, "navigational", "+%200 yıllık"],
    ["kapsama alanı sorgulama", 30, "düşük", "navigational", "+%350 yıllık"],
    ["--- A6.4 SINIRSIZ MECRA ---", "", "", "", ""],
    ["sınırsız internet paketi", 3600, 29, "commercial", "★★"],
    ["sınırsız sosyal medya paketi", 1600, "—", "commercial", "★★"],
    ["sınırsız youtube paketi", 590, "—", "commercial", "★★"],
    ["sınırsız tiktok paketi", 70, "—", "commercial", "★ +%250 yıllık trend"],
    ["sınırsız whatsapp paketi", 90, "—", "commercial", "★"],
    ["sınırsız instagram paketi", 70, "—", "commercial", "★"],
    ["--- A6.5 PERSONA TARİFELER ---", "", "", "", ""],
    ["öğrenci tarifeleri", 320, 43, "informational", "★★"],
    ["emekli tarifesi", 140, 1, "informational", "★ KD çok düşük"],
    ["mobil tarifeler", 880, 56, "commercial", "★"],
    ["--- A6.6 TEKNOLOJİ SÖZLÜĞÜ ---", "", "", "", ""],
    ["gsm nedir", 14800, "düşük", "informational", "★★★"],
    ["ram nedir", 9900, 5, "informational", "★★★ DÜŞÜK KD"],
    ["esim nedir (tekrar)", 6600, 19, "informational", "★★★"],
    ["ip adresi nedir", 6600, 36, "informational", "★★"],
    ["dns nedir", 5400, 13, "informational", "★★★"],
    ["router nedir", 5400, "düşük", "informational", "★★★"],
    ["fiber internet nedir", 1900, 2, "informational", "★★★ EN DÜŞÜK KD"],
    ["modem nedir", 1000, "düşük", "informational", "★★★"],
    ["vlan nedir", 720, "düşük", "informational", "★ niş"],
    ["wifi nedir", 320, 12, "informational", "★★"],
    ["fttx nedir", 90, "düşük", "informational", "★ teknik niş"],
    ["--- DİĞER YENİ DOĞRULAMALAR ---", "", "", "", ""],
    ["IMEI Sorgulama (*)", 550000, 14, "informational", "★★★ markaya iletilmiş"],
    ["gizli numara kapatma (*)", 27100, "düşük", "navigational", "★★★ markaya iletilmiş"],
    ["gizli numara nasıl aranır", 6600, "düşük", "informational", "★★"],
    ["kaç gün kaldı (*)", 33100, "düşük", "informational", "★★★ markaya iletilmiş"],
    ["iki tarih arası gün hesaplama", 22200, 20, "informational", "★★"],
    ["indirim hesaplama (*)", 49500, "—", "commercial", "★★★ markaya iletilmiş"],
    ["1 gb kaç mb", 5400, "düşük", "informational", "★★ telekom data"],
]
for r_idx_offset, row in enumerate(new_dfs_rows):
    actual_row = last_row + r_idx_offset
    for c_idx, val in enumerate(row, 1):
        cell = ws_dfs.cell(row=actual_row, column=c_idx, value=val)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if str(val).startswith("---"):
            cell.font = sub_hdr_font
            cell.fill = sub_hdr_fill
        elif "(*)" in str(val):
            cell.fill = star_fill
        elif "★★★" in str(ws_dfs.cell(row=actual_row, column=5).value or ""):
            cell.fill = prio_fill
        elif "★★" in str(ws_dfs.cell(row=actual_row, column=5).value or ""):
            cell.fill = mid_fill

# Mevcut A1-A5 fikir sheet'lerinde (*) işareti vurgusu ekle
for sname in ["01_A1_Hesaplama_Fikirleri", "02_A2_Tatil_Bayram_Fikirleri",
              "03_A3_OzelGun_Mesaj_Fikirleri", "04_A4_Telekom_Sorgu_Fikirleri",
              "05_A5_AI_Siber_Diger"]:
    if sname not in wb.sheetnames:
        continue
    ws_x = wb[sname]
    for r_idx in range(2, ws_x.max_row + 1):
        kw_val = str(ws_x.cell(row=r_idx, column=1).value or "")
        kw_val2 = str(ws_x.cell(row=r_idx, column=3).value or "")
        kw_val_all = kw_val + " " + kw_val2
        # MEVCUT yüzde/yaş hesaplama veya markaya iletilmiş referansları işaretle
        if "MEVCUT" in kw_val_all or "kbtogb" in kw_val_all.lower() or "kb to gb" in kw_val_all.lower():
            # Yıldız ile işaretle (header dışı)
            for c in range(1, 10):
                cell = ws_x.cell(row=r_idx, column=c)
                if cell.fill.start_color.rgb in [None, "00000000", "FFFFFFFF"]:  # boş ya da beyaz
                    cell.fill = star_fill

# Yeni A6 sheet'ini doğru pozisyona yerleştirmek için (sheet sıralamasını ayarla)
# Sheet sırasını manuel ayarla
new_order = [
    "00_Yonetici_Ozeti",
    "01_A1_Hesaplama_Fikirleri",
    "02_A2_Tatil_Bayram_Fikirleri",
    "03_A3_OzelGun_Mesaj_Fikirleri",
    "04_A4_Telekom_Sorgu_Fikirleri",
    "05_A5_AI_Siber_Diger",
    "06_A6_2026_Plan_Perspektif",  # YENİ
    "06_B1_Cluster_Ozeti",
    "07_B2_Rakip_URL_Patternleri",
    "08_DfS_Dogrulama_TR",
]
# Geri kalanları sıralamadan al
existing = list(wb.sheetnames)
remaining = [s for s in existing if s not in new_order]
final_order = new_order + remaining

# Sheet sıralaması için openpyxl'da _sheets attribute'unu kullanmak
sheet_map = {s.title: s for s in wb._sheets}
wb._sheets = [sheet_map[s] for s in final_order if s in sheet_map]

# Cluster özet sheet'inin numarasını güncelle
if "06_B1_Cluster_Ozeti" in wb.sheetnames:
    wb["06_B1_Cluster_Ozeti"].title = "07_B1_Cluster_Ozeti"
if "07_B2_Rakip_URL_Patternleri" in wb.sheetnames:
    wb["07_B2_Rakip_URL_Patternleri"].title = "08_B2_Rakip_URL_Patternleri"
if "08_DfS_Dogrulama_TR" in wb.sheetnames:
    wb["08_DfS_Dogrulama_TR"].title = "09_DfS_Dogrulama_TR"

wb.save(EXCEL_PATH)
print(f"\nExcel v3 kaydedildi: {EXCEL_PATH}")
print(f"Toplam sheet sayisi: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"  - {s}")
