"""
Excel'i HTML revizyonlarina paralel guncelle:
1. A1: 'Genel Hesap Makinesi Hub' satirini cikar
2. A1: Yukselen Burc satirini ticari yonlendirme odakli guncelle
3. A1: Saglik satirinda ovulasyon/regl kaldir
4. A4: Sorgulama satirinda IBAN kelime listesini cikar
5. A4: 'Cezaevi Iletisim Rehberi' satirini cikar
6. A5: 'Cicek/Hediye/Etkinlik' satirini cikar
7. A6 sheet: Bu raporun fikirleri konusuldu olarak isaretle (Veri Calc, Coverage, Sinirsiz, Persona, Sozluk)
8. Olarak refleksiyon Excel'in mete-sheet'leri (00 Ozet) hizmet kaldirildigini yansitsin
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

XLSX = Path("/Users/Erdo/Desktop/Claude Projects/Turkcel/TURKCELL_TRAFIK_FIRSATLARI.xlsx")
wb = load_workbook(XLSX)

# Stil yardimcilari
hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
hdr_fill = PatternFill("solid", fgColor="10332F")
star_fill = PatternFill("solid", fgColor="FFD966")
mint_fill = PatternFill("solid", fgColor="C8E6C9")

def col_idx(ws, header_name, row=1):
    """Ilk satirdaki kolon indeksini bul"""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=c).value
        if v and header_name.lower() in str(v).lower():
            return c
    return None

def delete_rows_matching(ws, search_col, keywords, header_row=1):
    """Belirli kolonda kelime icerien satirlari sil"""
    to_delete = []
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=search_col).value
        if v is None:
            continue
        v_str = str(v).lower()
        for kw in keywords:
            if kw.lower() in v_str:
                to_delete.append(r)
                break
    # Sondan basa silmek satir indekslerinde kayma olmasin
    for r in reversed(to_delete):
        ws.delete_rows(r)
    return len(to_delete)


changes = []

# ============================================================
# 1. A1: Genel Hesap Makinesi Hub kaldir
# ============================================================
if "01_A1_Hesaplama_Fikirleri" in wb.sheetnames:
    ws = wb["01_A1_Hesaplama_Fikirleri"]
    # Fikir kolonunu bul
    col = col_idx(ws, "Fikir") or 1
    deleted = delete_rows_matching(ws, col, ["Genel Hesap Makinesi Hub", "Genel Calculator Hub"])
    changes.append(f"A1: Genel Hesap Makinesi Hub satiri silindi ({deleted})")

# ============================================================
# 2. A1: Yukselen Burc -> Ticari yonlendirme
# ============================================================
if "01_A1_Hesaplama_Fikirleri" in wb.sheetnames:
    ws = wb["01_A1_Hesaplama_Fikirleri"]
    col_fikir = col_idx(ws, "Fikir") or 1
    col_url = col_idx(ws, "URL") or 2
    col_aciklama = col_idx(ws, "Notlar") or col_idx(ws, "Detay") or col_idx(ws, "Aciklama")
    col_kw = col_idx(ws, "Oynanabilecek") or col_idx(ws, "Kelime")

    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col_fikir).value
        if v and "Yükselen Burç" in str(v) or v and "Yukselen Burc" in str(v):
            ws.cell(row=r, column=col_fikir, value="Fikir 11: /burc-ve-astroloji-hesaplayicilari/ -> Ticari yonlendirme (paket + Pasaj)")
            if col_kw:
                ws.cell(row=r, column=col_kw, value="yukselen burc hesaplama, burc hesaplama, dogum haritasi, '[burc adi] icin tarife', '[burc adi] uyumlu paket', '[burc adi] hediye onerileri', '[burc adi] icin urun koleksiyonu' (Pasaj)")
            changes.append(f"A1: Yukselen Burc -> ticari yonlendirme (satir {r})")
            break

# ============================================================
# 3. A1: Saglik satirinda ovulasyon/regl kaldir
# ============================================================
if "01_A1_Hesaplama_Fikirleri" in wb.sheetnames:
    ws = wb["01_A1_Hesaplama_Fikirleri"]
    col_fikir = col_idx(ws, "Fikir") or 1
    col_kw = col_idx(ws, "Oynanabilecek") or col_idx(ws, "Kelime")
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col_fikir).value
        if v and ("Saglik Hesaplayicilari" in str(v) or "Sağlık Hesaplayıcı" in str(v)):
            if col_kw:
                ws.cell(row=r, column=col_kw, value="bmi hesaplama (8.1K), kalori (74K), ideal kilo, hamilelik haftasi (2.4K, KD 8), tahmini dogum tarihi, su tuketim hesaplama")
            changes.append(f"A1: Saglik satirindan ovulasyon/regl kaldirildi (satir {r})")
            break

# ============================================================
# 4. A4: IBAN sorgulama kelimelerini cikar (Sorgulama Rehberi Hub)
# ============================================================
if "04_A4_Telekom_Sorgu_Fikirleri" in wb.sheetnames:
    ws = wb["04_A4_Telekom_Sorgu_Fikirleri"]
    col_fikir = col_idx(ws, "Fikir") or 1
    col_kw = col_idx(ws, "Oynanabilecek") or col_idx(ws, "Kelime")
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col_fikir).value
        if v and ("Sorgulama Rehberi" in str(v) or "Sorgulama Hub" in str(v)):
            kw_val = ws.cell(row=r, column=col_kw).value if col_kw else None
            if kw_val:
                # IBAN ile baslayan virgule kadar olan kismi cikar
                new_kw = kw_val
                import re
                new_kw = re.sub(r'iban sorgulama[^,]*,\s*', '', new_kw, flags=re.IGNORECASE)
                new_kw = re.sub(r'iban hesaplama[^,]*,\s*', '', new_kw, flags=re.IGNORECASE)
                new_kw = re.sub(r'iban [oöuü]ğrenme[^,]*,\s*', '', new_kw, flags=re.IGNORECASE)
                new_kw = re.sub(r'iban nedir[^,]*,\s*', '', new_kw, flags=re.IGNORECASE)
                ws.cell(row=r, column=col_kw, value=new_kw)
            changes.append(f"A4: Sorgulama Rehberi'nden IBAN kelimeleri kaldirildi (satir {r})")
            break

# ============================================================
# 5. A4: Cezaevi Iletisim satirini cikar
# ============================================================
if "04_A4_Telekom_Sorgu_Fikirleri" in wb.sheetnames:
    ws = wb["04_A4_Telekom_Sorgu_Fikirleri"]
    col_fikir = col_idx(ws, "Fikir") or 1
    deleted = delete_rows_matching(ws, col_fikir, ["Cezaevi", "cezaevi-iletisim"])
    changes.append(f"A4: Cezaevi Iletisim satiri silindi ({deleted})")

# ============================================================
# 6. A5: Cicek/Hediye/Etkinlik satirini cikar
# ============================================================
if "05_A5_AI_Siber_Diger" in wb.sheetnames:
    ws = wb["05_A5_AI_Siber_Diger"]
    col_fikir = col_idx(ws, "Fikir") or 1
    deleted = delete_rows_matching(ws, col_fikir, ["Cicek", "Çiçek", "Hediye / Etkinlik"])
    changes.append(f"A5: Cicek/Hediye/Etkinlik satiri silindi ({deleted})")

# ============================================================
# 7. A6 sheet: Tum fikirleri 'konusuldu' olarak isaretle
# ============================================================
if "06_A6_2026_Plan_Perspektif" in wb.sheetnames:
    ws = wb["06_A6_2026_Plan_Perspektif"]
    # 'Markaya iletildi mi?' veya 'Konusuldu' sutunu var mi?
    col_fikir = col_idx(ws, "Fikir") or 1

    # Hangi fikirler konusuldu olarak isaretlenecek
    konusuldu_fikirleri = [
        "Veri / Data Kullanim",
        "Veri / Data Kullanım",
        "Coverage Map",
        "5G Kapsama",
        "Sinirsiz Mecra",
        "Sınırsız Mecra",
        "Persona Odakli",
        "Persona Odaklı",
        "Teknoloji Sozlugu",
        "Teknoloji Sözlüğü",
        "Hava Durumu",  # zaten konusuldu
        "Tarife Karsilastirma",
        "Tarife Karşılaştırma",
        "Paket Listeleme"
    ]

    # Yeni bir kolonu kullanmak yerine, fikir basligini "(konusuldu)" eki ile guncelle
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col_fikir).value
        if not v:
            continue
        v_str = str(v)
        for kw in konusuldu_fikirleri:
            if kw.lower() in v_str.lower():
                if "(konuşuldu)" not in v_str and "(konusuldu)" not in v_str:
                    new_v = v_str + " (konuşuldu)"
                    ws.cell(row=r, column=col_fikir, value=new_v)
                    ws.cell(row=r, column=col_fikir).fill = star_fill
                    changes.append(f"A6: '{v_str[:40]}...' konusuldu olarak isaretlendi (satir {r})")
                break

# ============================================================
# 8. A6 sheet: Teknoloji Sozlugu icin blog notu ekle
# ============================================================
if "06_A6_2026_Plan_Perspektif" in wb.sheetnames:
    ws = wb["06_A6_2026_Plan_Perspektif"]
    col_fikir = col_idx(ws, "Fikir") or 1
    col_notlar = col_idx(ws, "Notlar") or ws.max_column
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col_fikir).value
        if v and ("Teknoloji Sozlugu" in str(v) or "Teknoloji Sözlüğü" in str(v) or "Sözlük" in str(v)):
            cur_notlar = ws.cell(row=r, column=col_notlar).value or ""
            if "blog" not in cur_notlar.lower():
                new_notlar = "Blog altyapisi uzerinden ilerlenebilir; mevcut blog sisteminde her terim icin ayri bir post acilarak hizlica uretilebilir, ekstra teknik yatirim gerekmez. " + cur_notlar
                ws.cell(row=r, column=col_notlar, value=new_notlar)
            changes.append(f"A6: Teknoloji Sozlugu'ne blog notu eklendi (satir {r})")
            break

wb.save(XLSX)
print(f"\n=== Excel guncellendi ({XLSX}) ===")
for c in changes:
    print("  -", c)
print(f"\nToplam degisiklik: {len(changes)}")
