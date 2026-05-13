"""
Excel'deki 02_A2_Tatil_Bayram_Fikirleri sheet'ini yeni A2 yaklasimina gore guncelle.
Eski 7 alt fikir yerine: 1 mega fikir (mevcut sayfa genisletme) + icerik mimarisi bloklari.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

XLSX = Path("/Users/Erdo/Desktop/Claude Projects/Turkcel/TURKCELL_TRAFIK_FIRSATLARI.xlsx")
wb = load_workbook(XLSX)

SH = "02_A2_Tatil_Bayram_Fikirleri"
if SH not in wb.sheetnames:
    print("ERROR: sheet not found")
    raise SystemExit(1)

# Eski sheet'i sil ve yenisini olustur (ayni indekste)
old_idx = wb.sheetnames.index(SH)
del wb[SH]
ws = wb.create_sheet(SH, old_idx)

# Stiller
hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
hdr_fill = PatternFill("solid", fgColor="10332F")
sub_hdr_font = Font(bold=True, size=11, name="Calibri")
sub_hdr_fill = PatternFill("solid", fgColor="FFE3D8")
prio_fill = PatternFill("solid", fgColor="C6EFCE")
star_fill = PatternFill("solid", fgColor="FFD966")
border = Border(left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"))

# Header
ws["A1"] = "A2 — RESMİ TATİL & BAYRAM HUB (MEVCUT SAYFA GENİŞLETME)"
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
ws["A1"].fill = hdr_fill
ws.merge_cells("A1:G1")
ws.row_dimensions[1].height = 26

ws["A2"] = "Karar: Yeni sayfa AÇILMIYOR. Mevcut /kampanya/resmi-tatil-gunleri/ sayfası tatil+bayram cluster'ı için tek hub olarak genişletilecek."
ws["A2"].font = Font(bold=True, italic=True, size=11, color="E85F36", name="Calibri")
ws["A2"].fill = star_fill
ws.merge_cells("A2:G2")
ws.row_dimensions[2].height = 22

# Bos satir
ws.row_dimensions[3].height = 6

# Sayfa metadata
meta_rows = [
    ["URL", "https://www.turkcell.com.tr/kampanya/resmi-tatil-gunleri/ (DEĞİŞMEZ)"],
    ["Strateji", "Mevcut sayfaya 10 yapısal blok ekleyerek mega-hub haline getir"],
    ["Ana KW (DfS)", "resmi tatiller 2026 = 22.200/ay (KD 3, yıllık trend +%8.303); resmi tatil günleri = 6.600/ay (KD 8)"],
    ["Referans Model", "Vodafone /resmi-tatiller = 30.896 trafik / 579 keyword (tek sayfa)"],
    ["Toplam Adreslenebilir Hacim", "10M+ /ay (tüm tatil/bayram varyasyonları)"],
    ["Brand Uyumu", "★★★ Telekom'un yıllık spike LP modeli için ideal"],
    ["Markaya İletildi mi?", "Mevcut sayfa zaten var (*) - genişletme önerisi yeni"],
]
r = 4
for k, v in meta_rows:
    ws.cell(row=r, column=1, value=k).font = Font(bold=True, size=11, color="10332F", name="Calibri")
    ws.cell(row=r, column=1).fill = sub_hdr_fill
    ws.cell(row=r, column=2, value=v).font = Font(size=11, name="Calibri")
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = border
    r += 1

# Bos satir
r += 1

# Icerik Mimarisi tablosu basligi
ws.cell(row=r, column=1, value="SAYFA İÇERİK MİMARİSİ — Mevcut Sayfaya Eklenecek Bloklar").font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
for c in range(1, 8):
    ws.cell(row=r, column=c).fill = hdr_fill
    ws.cell(row=r, column=c).border = border
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws.row_dimensions[r].height = 22
r += 1

# Tablo basliklari
headers = ["Blok #", "Blok Adı", "İçerik", "Ana Hedef KW", "DfS Hacim", "KD", "Notlar"]
for c, h in enumerate(headers, 1):
    ws.cell(row=r, column=c, value=h).font = hdr_font
    ws.cell(row=r, column=c).fill = hdr_fill
    ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(row=r, column=c).border = border
ws.row_dimensions[r].height = 30
r += 1

# Icerik bloklari
bloklar = [
    ["1", "Yıllık Resmi Tatil Takvimi (dinamik)", "2026, 2025, 2027 toggle. Her gün için: tarih, ad, tatil mi, kaç gün.",
     "resmi tatiller 2026; resmi tatil günleri", "22.200 + 6.600", "3-8", "Yıl değiştikçe canlı güncellenmeli"],
    ["2", "Dini Bayramlar (Kurban + Ramazan)", "Kurban Bayramı, Ramazan Bayramı, kandiller - her biri accordion",
     "kurban bayramı ne zaman; ramazan ne zaman 2026", "1.220.000 + 18.000", "11 / 0", "Mart-Mayıs spike 6M"],
    ["3", "Milli Bayramlar Bölümü", "29 Ekim, 23 Nisan, 19 Mayıs, 30 Ağustos, 10 Kasım, 18 Mart, 1 Mayıs",
     "29 ekim cumhuriyet bayramı", "450.000 (Ekim 5M)", "—", "Spike ay kapsamı"],
    ["4", "Okul Tatilleri Bölümü", "Yarı yıl, sömestr, yaz tatili, MEB takvim",
     "yarı yıl tatili 2026; sömestr ne zaman; yaz tatili", "11.000 + 11.000 + 11.000", "6 / 7 / 26", "Veli kitlesi"],
    ["5", "Yılbaşı ve Yeni Yıl", "31 Aralık, yeni yıl, yılbaşı tatili",
     "31 aralık tatil mi; yılbaşı hangi güne", "13.000 + 1.700", "0", "Aralık spike"],
    ["6", "Uzun Tatil Planı (9-15 gün)", "Bağlı tatiller, izin alarak uzayan tatiller",
     "2026 15 tatil ne zaman", "14.000", "7", "Yıl başı + ortası planlama"],
    ["7", "Bugün Tatil mi? Hızlı Sorgu", "Dinamik widget (bugün, X tarihinde tatil mi)",
     "bugün tatil mi; 29 ekim tatil mi; 18 mart resmi tatil mi", "2.000 + 10.000 + 6.800", "0-30", "JS widget gerek"],
    ["8", "Kaç Gün Kaldı? Geri Sayım", "Önümüzdeki resmi tatillere geri sayım widget",
     "bayrama kaç gün kaldı; kurban bayramına kaç gün; ramazan'a kaç gün", "2.600 + 3.200 + 13.000", "0-1", "Dinamik countdown"],
    ["9", "Black Friday & Alışveriş Günleri", "Kara Cuma, Cyber Monday, 11.11",
     "black friday ne zaman; kara cuma ne zaman", "31.000 + 13.000", "11", "Kasım-Aralık spike"],
    ["10", "SSS / FAQ Bloku (AI Overview için)", "30-40 soru: Kurban hangi gün?, 29 Ekim resmi tatil mi? vb.",
     "uzun kuyruk varyasyonları", "Toplam 200K+", "0-15", "FAQPage schema kritik"],
]
for blok in bloklar:
    for c, val in enumerate(blok, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border
        cell.font = Font(size=10, name="Calibri")
    ws.row_dimensions[r].height = 42
    r += 1

# Bos satir
r += 1

# Kelime listesi basligi
ws.cell(row=r, column=1, value="MEVCUT SAYFAYA EKLENECEK TAM KELİME LİSTESİ").font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
for c in range(1, 8):
    ws.cell(row=r, column=c).fill = hdr_fill
    ws.cell(row=r, column=c).border = border
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws.row_dimensions[r].height = 22
r += 1

# KW tablosu basliklari
kw_headers = ["Kategori", "Anahtar Kelime", "DfS / Ahrefs Hacim", "KD", "Sezonsal Spike", "Hedef Blok", "Notlar"]
for c, h in enumerate(kw_headers, 1):
    ws.cell(row=r, column=c, value=h).font = hdr_font
    ws.cell(row=r, column=c).fill = hdr_fill
    ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(row=r, column=c).border = border
ws.row_dimensions[r].height = 28
r += 1

# Tam kelime listesi
kelimeler = [
    # Resmi Tatil Hub
    ["Resmi Tatil Hub", "resmi tatiller 2026", "22.200", "3", "Kasım-Şubat zirve", "Blok 1", "Yıllık trend +%8.303"],
    ["Resmi Tatil Hub", "resmi tatiller 2026 listesi", "—", "—", "—", "Blok 1", "Varyant"],
    ["Resmi Tatil Hub", "resmi tatil günleri", "6.600", "8", "Yıl boyu", "Blok 1", ""],
    ["Resmi Tatil Hub", "2026 15 tatil ne zaman", "14.000", "7", "Ocak", "Blok 6", "Uzun tatil planı"],
    ["Resmi Tatil Hub", "2026 15 tatil", "1.800", "2", "Ocak", "Blok 6", ""],
    ["Resmi Tatil Hub", "2025 15 tatil ne zaman", "5.700", "40", "Ocak 2025", "Blok 6", ""],
    ["Resmi Tatil Hub", "9 gün tatil 2026", "tahmin ~3K", "düşük", "Spike ay öncesi", "Blok 6", ""],
    ["Resmi Tatil Hub", "10 gün tatil", "tahmin ~3K", "düşük", "—", "Blok 6", ""],
    ["Resmi Tatil Hub", "bağlı tatil 2026", "tahmin ~3K", "düşük", "—", "Blok 6", ""],
    ["Resmi Tatil Hub", "okul tatilleri 2026", "tahmin ~5K", "düşük", "—", "Blok 4", "Veli kitlesi"],
    ["Resmi Tatil Hub", "meb tatil takvimi 2026", "tahmin ~3K", "düşük", "—", "Blok 4", ""],
    ["Resmi Tatil Hub", "yarı yıl tatili 2026", "11.000", "6", "Ocak-Şubat", "Blok 4", ""],
    ["Resmi Tatil Hub", "yariyil tatili", "varyant", "—", "—", "Blok 4", ""],
    ["Resmi Tatil Hub", "sömestr ne zaman", "11.000", "7", "Ocak-Şubat", "Blok 4", ""],
    ["Resmi Tatil Hub", "2026 yarıyıl tatili", "7.200", "3", "Ocak-Şubat", "Blok 4", ""],
    ["Resmi Tatil Hub", "yaz tatili ne zaman 2026", "11.000", "26", "Mayıs-Haziran", "Blok 4", ""],
    ["Resmi Tatil Hub", "kasım tatili 2025", "16.000", "45", "Kasım", "Blok 4", ""],
    # Kurban Bayramı
    ["Kurban Bayramı", "kurban bayramı ne zaman", "10.000+", "1", "Mart-Mayıs (Mayıs 6M)", "Blok 2", "Ana KW"],
    ["Kurban Bayramı", "2025 kurban bayramı ne zaman", "77.000", "2", "Mart-Mayıs 2025", "Blok 2", "Mevcut yıl"],
    ["Kurban Bayramı", "kurba bayram ne zaman 2025", "41.000", "5", "Mart-Mayıs 2025", "Blok 2", "Yazım varyantı"],
    ["Kurban Bayramı", "bayram ne zaman 2025", "13.000", "3", "Mart-Mayıs 2025", "Blok 2", ""],
    ["Kurban Bayramı", "2026 kurbanlık bayramı diyanet", "7.300", "4", "Mart-Mayıs 2026", "Blok 2", ""],
    ["Kurban Bayramı", "kurbann bayramı 2025", "6.300", "18", "—", "Blok 2", "Tipografik varyant"],
    ["Kurban Bayramı", "kurban bayramı hangi gün", "—", "—", "—", "Blok 2", ""],
    ["Kurban Bayramı", "kurban bayramı tarihi", "1.600", "2", "—", "Blok 2", ""],
    ["Kurban Bayramı", "kurban bayramı hangi ayda", "3.400", "1", "—", "Blok 2", ""],
    ["Kurban Bayramı", "kurban bayramı kaçında", "1.400", "1", "—", "Blok 2", ""],
    ["Kurban Bayramı", "kurban bayramı ayın kaçına", "1.900", "1", "—", "Blok 2", ""],
    ["Kurban Bayramı", "arefe günü tatil mi", "5.800 + 2.500", "1", "Mart-Mayıs", "Blok 7", "Hızlı sorgu"],
    ["Kurban Bayramı", "arefe ne zaman", "5.400", "28", "—", "Blok 2", ""],
    ["Kurban Bayramı", "arefe hangi gün", "2.400", "17", "—", "Blok 2", ""],
    ["Kurban Bayramı", "bayrama kaç gün kaldı", "2.600", "0", "Bayram öncesi", "Blok 8", "Geri sayım"],
    ["Kurban Bayramı", "kurban bayramı geri sayım", "—", "—", "—", "Blok 8", ""],
    # Ramazan
    ["Ramazan", "ramazan ne zaman başlıyor 2026", "18.000", "0", "Şubat-Mart", "Blok 2", ""],
    ["Ramazan", "2026 ramazan ne zaman başlıyor", "8.700", "—", "Şubat-Mart 2026", "Blok 2", ""],
    ["Ramazan", "2025 ramazan'a kaç gün kaldı", "13.000", "—", "—", "Blok 8", "Geri sayım"],
    ["Ramazan", "ramazan bayramı kaçında", "2.400", "17", "—", "Blok 2", ""],
    ["Ramazan", "ramazan bayramı hangi gün 2025", "1.300", "19", "—", "Blok 2", ""],
    # Kandiller
    ["Kandiller", "mevlit kandili ne zaman", "tahmin ~5K", "düşük", "Yıllık değişken", "Blok 2", ""],
    ["Kandiller", "berat kandili", "tahmin ~3K", "düşük", "—", "Blok 2", ""],
    ["Kandiller", "kadir gecesi", "tahmin ~10K", "düşük", "Ramazan'da", "Blok 2", ""],
    ["Kandiller", "regaip kandili", "tahmin ~2K", "düşük", "—", "Blok 2", ""],
    ["Kandiller", "miraç kandili", "tahmin ~2K", "düşük", "—", "Blok 2", ""],
    ["Kandiller", "kandil günleri 2026", "tahmin ~5K", "düşük", "—", "Blok 2", ""],
    # Milli Bayramlar
    ["Milli Bayramlar", "29 ekim cumhuriyet bayramı", "450.000", "—", "Ekim (5M)", "Blok 3", "Ekim ayı spike"],
    ["Milli Bayramlar", "29 ekim 1923 hangi gün", "2.600", "0", "Ekim", "Blok 3", ""],
    ["Milli Bayramlar", "29 ekim hangi gün", "3.600", "13", "Ekim", "Blok 7", ""],
    ["Milli Bayramlar", "29 ekim tatil mi", "10.000", "20", "Ekim", "Blok 7", "Hızlı sorgu"],
    ["Milli Bayramlar", "29 ekim resmi tatil mi", "9.500", "30", "Ekim", "Blok 7", ""],
    ["Milli Bayramlar", "29 ekim cumhuriyet bayramı tatil mi", "1.800", "20", "Ekim", "Blok 7", ""],
    ["Milli Bayramlar", "cumhuriyet bayramı resmi", "1.800", "0", "Ekim", "Blok 3", ""],
    ["Milli Bayramlar", "23 nisan ulusal egemenlik ve çocuk bayramı", "tahmin ~10K", "düşük", "Nisan", "Blok 3", ""],
    ["Milli Bayramlar", "23 nisan tatil mi", "tahmin ~5K", "düşük", "Nisan", "Blok 7", ""],
    ["Milli Bayramlar", "19 mayıs atatürk'ü anma", "tahmin ~8K", "düşük", "Mayıs", "Blok 3", ""],
    ["Milli Bayramlar", "19 mayıs tatil mi", "tahmin ~3K", "düşük", "Mayıs", "Blok 7", ""],
    ["Milli Bayramlar", "30 ağustos zafer bayramı", "tahmin ~8K", "düşük", "Ağustos", "Blok 3", ""],
    ["Milli Bayramlar", "30 ağustos tatil mi", "tahmin ~3K", "düşük", "Ağustos", "Blok 7", ""],
    ["Milli Bayramlar", "18 mart çanakkale şehitlerini anma günü", "6.800", "1", "Mart", "Blok 3", ""],
    ["Milli Bayramlar", "18 mart resmi tatil mi", "6.800", "1", "Mart", "Blok 7", ""],
    ["Milli Bayramlar", "on kasım resmi tatil mi", "1.200", "4", "Kasım", "Blok 7", "10 Kasım"],
    ["Milli Bayramlar", "10 kasım anma töreni", "tahmin ~3K", "düşük", "Kasım", "Blok 3", ""],
    ["Milli Bayramlar", "1 mayıs emek ve dayanışma günü", "1.300", "—", "Mayıs", "Blok 3", ""],
    # Yılbaşı
    ["Yılbaşı", "yılbaşı ne zaman", "tahmin ~10K", "düşük", "Aralık", "Blok 5", ""],
    ["Yılbaşı", "31 aralık tatil mi", "13.000", "0", "Aralık", "Blok 7", "Hızlı sorgu"],
    ["Yılbaşı", "yılbaşı hangi güne denk geliyor", "1.700", "14", "Aralık", "Blok 5", ""],
    ["Yılbaşı", "yılbaşı 2026", "tahmin ~15K", "düşük", "Aralık", "Blok 5", ""],
    ["Yılbaşı", "yeni yıl tatili", "tahmin ~5K", "düşük", "Aralık", "Blok 5", ""],
    # Genel sorgu
    ["Bugün Tatil mi?", "bugün tatil mi", "2.000", "20", "Yıl boyu", "Blok 7", "JS widget"],
    ["Bugün Tatil mi?", "pazartesi günü okullar tatil mi", "2.100", "0", "Yıl boyu", "Blok 7", ""],
    ["Bugün Tatil mi?", "9 haziran okullar tatil mi", "1.400", "16", "Haziran", "Blok 7", ""],
    # Black Friday
    ["Black Friday", "black friday ne zaman", "31.000", "11", "Kasım", "Blok 9", "Vodafone 3."],
    ["Black Friday", "black friday 2026", "tahmin ~10K", "düşük", "Kasım", "Blok 9", ""],
    ["Black Friday", "black friday türkiye", "tahmin ~5K", "düşük", "Kasım", "Blok 9", ""],
    ["Black Friday", "kara cuma ne zaman", "13.000", "10", "Kasım", "Blok 9", "MediaMarkt 2."],
    ["Black Friday", "kara cuma indirimleri", "tahmin ~5K", "düşük", "Kasım", "Blok 9", ""],
    ["Black Friday", "cyber monday", "tahmin ~3K", "düşük", "Kasım", "Blok 9", ""],
    ["Black Friday", "11.11 singles day", "tahmin ~3K", "düşük", "Kasım", "Blok 9", ""],
]

for kw in kelimeler:
    for c, val in enumerate(kw, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border
        cell.font = Font(size=10, name="Calibri")
    r += 1

# Teknik Uygulama Notlari
r += 2
ws.cell(row=r, column=1, value="TEKNİK UYGULAMA NOTLARI").font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
for c in range(1, 8):
    ws.cell(row=r, column=c).fill = hdr_fill
    ws.cell(row=r, column=c).border = border
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws.row_dimensions[r].height = 22
r += 1

teknik_notlar = [
    "1. Schema markup: Event (her tatil için), FAQPage (SSS bloku), BreadcrumbList - AI Overview odaklı",
    "2. İçerik dinamik: 'Bugün tatil mi?' widget'ı server-time veya client-time ile gerçek zamanlı",
    "3. Yıllık güncellenmeli: Yıl değiştikçe canlı tablo otomatik (CMS field). 2026, 2027 toggle",
    "4. Internal link: A3 (özel gün mesajları), A6.2 (hava durumu), A1 (kaç gün kaldı hesaplama) sayfalarına",
    "5. Mevcut sayfa kanibalizasyon kontrolü: Tatil ile ilgili dağınık destek/SSS sayfaları varsa 301 ile birleştirilmeli",
    "6. Spike aylarında (Mart-Mayıs, Ekim, Aralık) ek SEM/social destek değerlendirilebilir",
    "7. Yeni URL açılmıyor: /kampanya/resmi-tatil-gunleri/ tek hub, alt sayfa yok",
    "8. Mevcut sayfa içeriği KORUNUR, üzerine 10 yapısal blok eklenir",
]
for note in teknik_notlar:
    ws.cell(row=r, column=1, value=note).font = Font(size=11, name="Calibri")
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = border
    ws.row_dimensions[r].height = 20
    r += 1

# Kolon genislikleri
widths = {"A": 18, "B": 36, "C": 36, "D": 32, "E": 22, "F": 12, "G": 40}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

wb.save(XLSX)
print(f"Updated: {XLSX}")
print(f"02_A2 sheet replaced with new mega-hub structure ({len(kelimeler)} keywords)")
