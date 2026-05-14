"""
Excel'de kalan duzeltmeler:
- A1 Row 10 (Saglik): KW listesinde ovulasyon/regl kaldir
- A1 Row 11 (Burc): KW listesini ticari yonlendirme yap
- A4 Row 4 (Sorgulama Rehberi): KW listesinde IBAN kaldir
"""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

XLSX = "/Users/Erdo/Desktop/Claude Projects/Turkcel/TURKCELL_TRAFIK_FIRSATLARI.xlsx"
wb = load_workbook(XLSX)
star_fill = PatternFill("solid", fgColor="FFD966")

# A1 Sheet
ws = wb["01_A1_Hesaplama_Fikirleri"]
# Headers: Fikir(1), URL(2), Ana_Keyword(3), Hacim(4), KD(5), Intent(6), Oncelik(7), KW_Listesi(8), Notlar(9)

# Row 10 - Saglik mini-hub
r = 10
cur_kw = str(ws.cell(row=r, column=8).value or "")
new_kw = cur_kw
new_kw = re.sub(r"ovulasyon hesaplama \(6\.6K, KD 7\)[,;]?\s*", "", new_kw)
new_kw = re.sub(r"regl hesaplama \(18\.1K\)[,;]?\s*", "", new_kw)
new_kw = re.sub(r"ovulasyon \(6\.6K, KD 7\)[,;]?\s*", "", new_kw)
new_kw = re.sub(r"regl hesaplama \(18\.1K\)[,;]?\s*", "", new_kw)
# Eger su kelimeler hala varsa
new_kw = re.sub(r"\bovulasyon[^,]*,?\s*", "", new_kw, flags=re.IGNORECASE)
new_kw = re.sub(r"\bregl[^,]*,?\s*", "", new_kw, flags=re.IGNORECASE)
new_kw = new_kw.strip().strip(',').strip()
ws.cell(row=r, column=8, value=new_kw)
# Ana_Keyword da degisecek
ws.cell(row=r, column=3, value="bmi/kalori/hamilelik/su tuketim")
print(f"A1 Row 10 (Saglik): KW listesi guncellendi")
print(f"  -> {new_kw}")

# Row 11 - Burc/Astroloji ticari yonlendirme
r = 11
ws.cell(row=r, column=1, value="FIKIR 11: /burc-ve-astroloji-hesaplayicilari/ -> Ticari yonlendirme (paket + Pasaj)")
ws.cell(row=r, column=3, value="yukselen burc hesaplama (ticari sonuc sayfasi)")
ws.cell(row=r, column=8, value="yukselen burc hesaplama, burc hesaplama, ay/gunes burcu, cin burcu, dogum haritasi, '[burc adi] icin tarife', '[burc adi] uyumlu paket', '[burc adi] hediye onerileri', '[burc adi] icin urun koleksiyonu' (Pasaj)")
ws.cell(row=r, column=9, value="Eglence trafigini ticari hub'a kanalize eden bir kopru gorevi gorur. Sonuc sayfasinda burcuna gore tarife onerisi + Pasaj urun-kategori vitrini acilir.")
print(f"A1 Row 11 (Burc): ticari yonlendirme olarak guncellendi")

# A4 Sheet - Sorgulama Rehberi (Row 4)
ws = wb["04_A4_Telekom_Sorgu_Fikirleri"]
r = 4
cur_kw = str(ws.cell(row=r, column=8).value or "")
new_kw = cur_kw
new_kw = re.sub(r"iban sorgulama \(33\.1K\)[,;]?\s*", "", new_kw, flags=re.IGNORECASE)
new_kw = re.sub(r"iban hesaplama \(170\)[,;]?\s*", "", new_kw, flags=re.IGNORECASE)
new_kw = re.sub(r"iban [oö]grenme[,;]?\s*", "", new_kw, flags=re.IGNORECASE)
new_kw = re.sub(r"iban nedir[,;]?\s*", "", new_kw, flags=re.IGNORECASE)
new_kw = re.sub(r"\biban [^,]+,?\s*", "", new_kw, flags=re.IGNORECASE)
new_kw = new_kw.strip().strip(',').strip()
ws.cell(row=r, column=8, value=new_kw)
print(f"A4 Row 4 (Sorgulama Rehberi): IBAN kelimeleri kaldirildi")
print(f"  -> {new_kw[:100]}...")

wb.save(XLSX)
print(f"\nKayit edildi: {XLSX}")
